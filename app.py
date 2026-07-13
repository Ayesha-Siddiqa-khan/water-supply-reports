import csv
import io
import json
import zipfile
import os
import re
import shutil
import sqlite3
import sys
import threading
import webbrowser
from datetime import datetime

import pandas as pd
from flask import (
    Flask,
    Response,
    flash,
    has_request_context,
    redirect,
    render_template,
    request,
    send_from_directory,
    session,
    url_for,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    BaseDocTemplate,
    KeepTogether,
    Frame,
    PageBreak,
    PageTemplate,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from werkzeug.utils import secure_filename

def resource_path(*parts: str) -> str:
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, *parts)
    return os.path.join(os.path.abspath(os.path.dirname(__file__)), *parts)


def app_data_dir() -> str:
    if os.environ.get("VERCEL") == "1":
        return os.path.join("/tmp", "water-supply-report")
    if getattr(sys, "frozen", False):
        return os.path.abspath(os.path.dirname(sys.executable))
    return os.path.abspath(os.path.dirname(__file__))


BASE_DIR = app_data_dir()
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
RESULTS_CACHE = os.path.join(UPLOAD_FOLDER, "last_results.json")
SAVED_DASHBOARD_CSV = os.path.join(UPLOAD_FOLDER, "saved_dashboard_data.csv")
SAVED_DASHBOARD_META = os.path.join(UPLOAD_FOLDER, "saved_dashboard_meta.json")
# Persistent cache for the Consumer Sector Report summary so it survives across
# serverless function invocations on Vercel (global state is not kept between requests)
CONSUMER_REPORT_CACHE = os.path.join(UPLOAD_FOLDER, "consumer_report_cache.json")
BILL_LIST_DB = os.path.join(BASE_DIR, "bill_list.sqlite3")
SEED_ASSIGNMENTS_JSON = resource_path("seed", "staff_assignments.json")
SEED_ASSIGNMENTS_CSV = resource_path("seed", "staff_assignments.csv")
ALLOWED_EXTENSIONS = {".csv", ".xlsx"}

app = Flask(
    __name__,
    template_folder=resource_path("templates"),
    static_folder=resource_path("static"),
)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200 MB


# Handle 413 Payload Too Large errors gracefully instead of crashing
@app.errorhandler(413)
def request_entity_too_large(error):
    msg = "File is too large. Maximum upload size is 200 MB."
    if is_ajax():
        return ajax_error(msg)
    flash(msg)
    return redirect(request.referrer or url_for("consumer_report"))
app.secret_key = "dev-secret"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
if not os.path.exists(BILL_LIST_DB):
    bundled_db = resource_path("bill_list.sqlite3")
    if os.path.exists(bundled_db):
        shutil.copy2(bundled_db, BILL_LIST_DB)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

STAFF_PAIRS = {
    "MUHAMMAD ILYAS": "MUHAMMAD ILYAS\nMUHAMMAD NAEEM",
    "MUHAMMAD IRFAN": "MUHAMMAD IRFAN\nZULFIQAR",
    "MUHAMMAD KHALID": "MUHAMMAD KHALID\nMUHAMMAD NAZAR",
    "MUHAMMAD KASHIF": "MUHAMMAD KASHIF\nMUHAMMAD SHAHID",
    "TAUSEEF": "TAUSEEF\nGOSHI IBRAHIM",
}

AUTO_ASSIGNMENT_RULES = [
    {
        "sector": "29 - Mahboob Colony (SHARQI)",
        "locality": "02 - Mehboob Colony SHARQI Zone B",
        "connection_min": "12020001",
        "connection_max": "12020560",
        "staff_name": "MUHAMMAD KASHIF",
    },
    {
        "sector": "29 - Mahboob Colony (SHARQI)",
        "locality": "02 - Mehboob Colony SHARQI Zone B",
        "connection_min": "12020561",
        "connection_max": None,
        "staff_name": "MUHAMMAD ILYAS",
    },
    {
        "sector": "51 - NoorPura",
        "locality": "01 - Noor Pura Zone C",
        "connection_min": "04010001",
        "connection_max": "04010428",
        "staff_name": "MUHAMMAD MURTAZA",
    },
    {
        "sector": "51 - NoorPura",
        "locality": "01 - Noor Pura Zone C",
        "connection_min": "04010429",
        "connection_max": None,
        "staff_name": "ABDUL LATIF",
    },
]

_AAR_SECTOR = AUTO_ASSIGNMENT_RULES[0]["sector"]
_AAR_LOCALITY = AUTO_ASSIGNMENT_RULES[0]["locality"]

def _normalize_sector_locality(value: str) -> str:
    text = (value or "").strip().lower()
    text = re.sub(r"^\d+\s*[-–—]\s*", "", text)
    text = text.replace("mehbob", "mehboob")
    text = text.replace("noorpura", "noor pura")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


_UNMATCHED_LOG: list[dict] = []


def get_unmatched_log() -> list[dict]:
    return list(_UNMATCHED_LOG)


def clear_unmatched_log() -> None:
    _UNMATCHED_LOG.clear()


def _deep_normalize_sector(value: str) -> str:
    """Aggressively normalize a sector/locality name for robust matching."""
    text = (value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[,;:.()\[\]\{\}'\"`!@#$%^&*=+<>?/\\|~]", " ", text)
    text = re.sub(r"[–—\-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    # Known spelling/canonical variants (both directions)
    pairs = [
        ("gulbarg", "gulberg"),
        ("ghaziani", "g"),
        ("ghazianai", "g"),
        ("sabzi", "sabz"),
        ("sabaz", "sabz"),
        ("bagh", "bagh"),
        ("mohallah", "mohalla"),
        ("ghareeb", "ghareeb"),
        ("sharqi", "sharqi"),
        ("shumali", "shumali"),
        ("gharbi", "gharbi"),
        ("janubi", "janubi"),
        ("janub", "janubi"),
        ("colony", "colony"),
        ("society", "society"),
        ("private", "private"),
        ("phase", "phase"),
        ("town", "town"),
        ("chak", "chak"),
        ("nager", "nagar"),
        ("road", "road"),
        ("zone", "zone"),
        ("basti", "basti"),
        ("pura", "pura"),
        ("puraa", "pura"),
        ("pur", "pura"),
        ("mehboob", "mehboob"),
        ("mehbob", "mehboob"),
        ("mahboob", "mehboob"),
        ("mohajar", "mohajar"),
        ("muhajir", "mohajar"),
        ("mohajir", "mohajar"),
        ("noorpura", "noor pura"),
        ("noorpuraa", "noor pura"),
        ("noorpura", "noor pura"),
        ("kashmir", "kashmir"),
        ("zimindara", "zimindara"),
        ("zimandara", "zimindara"),
        ("zimidara", "zimindara"),
        ("zimmindara", "zimindara"),
        ("purani", "purani"),
        ("purany", "purani"),
        ("chishtian", "chishtian"),
        ("chistian", "chishtian"),
        ("feeder", "feeder"),
        ("fider", "feeder"),
        ("qadrabbad", "qadrabbad"),
        ("qadirabad", "qadrabbad"),
        ("qadir abad", "qadrabbad"),
        ("taqwa", "taqwa"),
        ("takwa", "taqwa"),
        ("takiya", "taqwa"),
        ("nasirabad", "nasirabbad"),
        ("nasirababad", "nasirabbad"),
        ("satellite", "satellite"),
        ("yasrab", "yasrab"),
        ("yasir", "yasrab"),
        ("tableghi", "tableghi"),
        ("tablegi", "tableghi"),
        ("markaz", "markaz"),
        ("markez", "markaz"),
        ("taj pura", "taj pura"),
        ("sarwar", "sarwar"),
        ("sarwer", "sarwar"),
        ("stadium", "stadium"),
    ]

    for a, b in pairs:
        # Replace word boundaries only to avoid partial matches
        text = re.sub(rf"\b{re.escape(a)}\b", b, text)
        text = re.sub(rf"\b{re.escape(b)}\b", b, text)

    text = re.sub(r"\s+", " ", text).strip()
    return text


def _keyword_set(value: str) -> set[str]:
    """Extract significant keywords from a sector/locality name."""
    text = (value or "").strip().lower()
    text = re.sub(r"[,;:.()\[\]\{\}'\"`!@#$%^&*=+<>?/\\|~\-–—]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    words = text.split()
    # Apply basic variant normalization to keywords
    norm_pairs = [
        ("sabzi", "sabz"), ("sabaz", "sabz"),
        ("ghaziani", "g"), ("gulbarg", "gulberg"),
        ("nager", "nagar"), ("mehbob", "mehboob"), ("mahboob", "mehboob"),
        ("mohajir", "mohajar"), ("muhajir", "mohajar"),
        ("zimandara", "zimindara"), ("zimidara", "zimindara"),
        ("chistian", "chishtian"),
    ]
    normalized = []
    for w in words:
        for a, b in norm_pairs:
            if w == a:
                w = b
                break
        normalized.append(w)
    stopwords = {"the", "a", "an", "and", "or", "of", "to", "in", "on", "at", "by", "for",
                 "with", "without", "near", "adjacent", "main", "road", "zone", "area",
                 "colony", "town", "city", "society", "private", "phase", "line", "connection"}
    return {w for w in normalized if w not in stopwords and len(w) > 1 and not w.isdigit()}


def _normalize_staff_name(name: str) -> str:
    return " ".join(name.split()).upper()

def _levenshtein(a: str, b: str) -> int:
    n, m = len(a), len(b)
    if n < m:
        a, b = b, a
        n, m = m, n
    prev = list(range(m + 1))
    for i in range(1, n + 1):
        curr = [i] * (m + 1)
        for j in range(1, m + 1):
            cost = 0 if a[i - 1] == b[j - 1] else 1
            curr[j] = min(curr[j - 1] + 1, prev[j] + 1, prev[j - 1] + cost)
        prev = curr
    return prev[m]

def _closest_staff_key(name: str, max_dist: int = 2):
    if not name:
        return None
    best_key = None
    best_dist = max_dist + 1
    for key in STAFF_PAIRS:
        d = _levenshtein(name, key)
        if d < best_dist:
            best_dist = d
            best_key = key
    return best_key

def fmt_staff_name(name: str) -> str:
    """Return display name: paired staff on separate lines, else as-is."""
    if not name:
        return name
    norm = _normalize_staff_name(name)
    key = _closest_staff_key(norm)
    if key:
        return STAFF_PAIRS[key]
    return name.replace(" / ", "\n")

def fmt_staff_name_html(name: str) -> str:
    """Like fmt_staff_name but returns HTML with <br> for display."""
    if not name:
        return name
    norm = _normalize_staff_name(name)
    key = _closest_staff_key(norm)
    display = STAFF_PAIRS.get(key) if key else name.replace(" / ", "\n")
    from markupsafe import Markup
    return Markup(display.replace("\n", "<br>"))

app.jinja_env.filters["staff_name"] = fmt_staff_name
app.jinja_env.filters["staff_name_html"] = fmt_staff_name_html

def allowed_file(filename: str) -> bool:
    _, ext = os.path.splitext(filename)
    return ext.lower() in ALLOWED_EXTENSIONS


def is_ajax() -> bool:
    return request.headers.get("X-Requested-With") == "XMLHttpRequest"


def ajax_ok(message: str = "", redirect_url: str = ""):
    from flask import jsonify
    payload = {"ok": True, "message": message}
    if redirect_url:
        payload["redirect"] = redirect_url
    return jsonify(payload)


def ajax_error(message: str):
    from flask import jsonify
    return jsonify({"ok": False, "error": message})


def read_dataframe(path: str) -> pd.DataFrame:
    _, ext = os.path.splitext(path)
    if ext.lower() == ".csv":
        return pd.read_csv(path)
    return pd.read_excel(path)


def read_uploaded_dataframe(file) -> pd.DataFrame:
    _, ext = os.path.splitext(file.filename)
    if ext.lower() == ".csv":
        return pd.read_csv(file)
    return pd.read_excel(file)


def normalize_column_name(name: str) -> str:
    name = str(name).strip().lower().replace("_", " ")
    return " ".join(name.split())


def normalize_sector_key(value: str) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\bprivate\b|\bsociety\b|\bphase\b", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


SECTOR_VALUE_DELIMITER = "|||"


def split_sector_values(values: list[str]) -> list[str]:
    sectors: list[str] = []
    for value in values:
        for sector in str(value).split(SECTOR_VALUE_DELIMITER):
            sector = sector.strip()
            if sector:
                sectors.append(sector)
    return sorted(set(sectors))


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [normalize_column_name(col) for col in df.columns]
    return df


def _dedupe_value(value) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return ""
    compact_number = text.replace(",", "")
    numeric_chars = compact_number.replace(".", "", 1).replace("-", "", 1)
    if numeric_chars.isdigit() and len(numeric_chars) <= 15:
        number = pd.to_numeric(pd.Series([compact_number]), errors="coerce").iloc[0]
        if pd.notna(number):
            return str(int(number)) if float(number).is_integer() else f"{float(number):.6f}".rstrip("0").rstrip(".")
    parsed_date = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.notna(parsed_date):
        return parsed_date.strftime("%Y-%m-%d")
    return " ".join(text.lower().split())


def drop_duplicate_bills(df: pd.DataFrame) -> tuple[pd.DataFrame, int, list[str]]:
    """Remove duplicate uploaded bills without collapsing different bills for one connection."""
    if df.empty:
        return df, 0, []

    normalized_columns = [normalize_column_name(col) for col in df.columns]
    column_lookup = dict(zip(normalized_columns, df.columns))
    key_candidates = [
        ["bill no"],
        ["reference no", "due date", "total bill"],
        ["reference no", "received date", "amount received"],
        ["connection no", "due date", "total bill"],
        ["connection no", "received date", "amount received"],
    ]

    key_columns = []
    for candidate in key_candidates:
        if all(col in column_lookup for col in candidate):
            key_columns = [column_lookup[col] for col in candidate]
            break

    if not key_columns:
        key_columns = list(df.columns)

    key_frame = df[key_columns].apply(lambda col: col.map(_dedupe_value))
    complete_key = key_frame.ne("").all(axis=1)
    duplicate_mask = pd.Series(False, index=df.index)

    if complete_key.any():
        duplicate_mask.loc[complete_key] = key_frame.loc[complete_key].duplicated(keep="last")

    if (~complete_key).any():
        full_row_key = df.apply(lambda col: col.map(_dedupe_value))
        duplicate_mask.loc[~complete_key] = full_row_key.loc[~complete_key].duplicated(keep="last")

    deduped = df.loc[~duplicate_mask].copy()
    return deduped, int(duplicate_mask.sum()), [normalize_column_name(col) for col in key_columns]


def pick_column(columns: list[str], candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in columns:
            return candidate
    return None


def fmt(value) -> str:
    if value is None:
        return "0"
    return f"{value:,.0f}"


def format_mobile(value) -> str:
    if value is None or value == "" or value == 0 or value == "0" or value == "0.0":
        return "-"
    mobile = str(value).strip()
    mobile = mobile.replace(".0", "")
    mobile = re.sub(r"\D", "", mobile)
    if not mobile or mobile == "0":
        return "-"
    if len(mobile) == 10 and mobile.startswith("3"):
        mobile = "0" + mobile
    if len(mobile) == 12 and mobile.startswith("92"):
        mobile = "0" + mobile[2:]
    return mobile


def parse_number(value) -> float:
    return clean_amount_value(value)


def format_fiscal_month(period: pd.Period) -> str:
    month_name = period.strftime("%B")
    fiscal_year = period.year + 1 if period.month >= 7 else period.year
    return f"{month_name} {str(fiscal_year)[-2:]}"


def format_calendar_month(period: pd.Period) -> str:
    return f"{period.strftime('%B')} {str(period.year)[-2:]}"


def fiscal_label_to_calendar_label(label: str) -> str:
    try:
        month_name, year_suffix = str(label).rsplit(" ", 1)
        month_num = datetime.strptime(month_name, "%B").month
        year = 2000 + int(year_suffix)
    except (TypeError, ValueError):
        return str(label)
    if month_num >= 7:
        year -= 1
    return f"{month_name} {str(year)[-2:]}"


def fiscal_label_to_calendar_full_label(label: str) -> str:
    short_label = fiscal_label_to_calendar_label(label)
    try:
        month_name, year_suffix = short_label.rsplit(" ", 1)
        return f"{month_name} {2000 + int(year_suffix)}"
    except (TypeError, ValueError):
        return str(label)


def format_report_date(value) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return str(value or "")
    return parsed.strftime("%d-%m-%Y")


def remove_pdf_column(headers: list[str], rows: list[list], column_name: str, grand: list | None = None):
    if column_name not in headers:
        return headers, rows, grand
    remove_idx = headers.index(column_name)
    pdf_headers = [value for idx, value in enumerate(headers) if idx != remove_idx]
    pdf_rows = [
        [value for idx, value in enumerate(row) if idx != remove_idx]
        for row in rows
    ]
    pdf_grand = None
    if grand is not None:
        pdf_grand = [value for idx, value in enumerate(grand) if idx != remove_idx]
    return pdf_headers, pdf_rows, pdf_grand


def clean_cell(value) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def match_key(value) -> str:
    return re.sub(r"\s+", " ", clean_cell(value)).casefold()


# ---------------------------------------------------------------------------
# Date parsing (DD-MM swap fix)
# ---------------------------------------------------------------------------

def swap_day_month(dt):
    """Swap day and month of a datetime to fix Excel DD-MM misinterpretation."""
    if pd.isna(dt):
        return pd.NaT
    try:
        if 1 <= dt.day <= 12:
            return dt.replace(month=dt.day, day=dt.month)
        return dt
    except ValueError:
        return pd.NaT


def parse_received_dates(series: pd.Series) -> pd.Series:
    """Parse received dates handling mixed formats and Excel DD-MM swap issues."""
    if pd.api.types.is_datetime64_any_dtype(series):
        return series.apply(swap_day_month)

    results = []
    for val in series:
        if pd.isna(val):
            results.append(pd.NaT)
        elif isinstance(val, (pd.Timestamp, datetime)):
            results.append(swap_day_month(val))
        else:
            str_val = str(val).strip()
            parsed = pd.to_datetime(str_val, errors="coerce", format="%d/%m/%Y", dayfirst=True)
            if pd.isna(parsed):
                parsed = pd.to_datetime(str_val, errors="coerce", format="%d/%m/%y", dayfirst=True)
            if pd.isna(parsed):
                parsed = pd.to_datetime(str_val, errors="coerce", dayfirst=True)
            results.append(parsed)
    return pd.Series(results, index=series.index, dtype="datetime64[ns]")


def get_fiscal_window() -> tuple[pd.Period, pd.Period]:
    current_period = pd.Period(datetime.now(), freq="M")
    start_year = current_period.year if current_period.month >= 7 else current_period.year - 1
    fiscal_start = pd.Period(year=start_year, month=7, freq="M")
    return fiscal_start, current_period


# ---------------------------------------------------------------------------
# Data builders
# ---------------------------------------------------------------------------

def build_daily_rows(dates: pd.Series, amounts: pd.Series | None, arrears: pd.Series | None) -> list[dict]:
    metric_df = pd.DataFrame({"date": dates})
    if amounts is not None:
        metric_df["amount"] = amounts.reindex(dates.index).fillna(0)
    if arrears is not None:
        metric_df["arrears"] = arrears.reindex(dates.index).fillna(0)
    # Filter to fiscal window
    fiscal_start, current_period = get_fiscal_window()
    fs_date = fiscal_start.start_time.date()
    cp_date = current_period.end_time.date()
    metric_df = metric_df[(metric_df["date"].dt.date >= fs_date) & (metric_df["date"].dt.date <= cp_date)]
    if metric_df.empty:
        return []
    grouped = metric_df.groupby(metric_df["date"].dt.date).agg(
        count=("date", "size"),
        **({} if amounts is None else {"amount": ("amount", "sum")}),
        **({} if arrears is None else {"arrears": ("arrears", "sum")}),
    ).sort_index()
    rows = []
    for label, row in grouped.iterrows():
        item = {"label": label.strftime("%d-%m-%Y"), "count": int(row["count"])}
        if "amount" in row:
            item["amount_total"] = float(row["amount"])
        if "arrears" in row:
            item["arrears_total"] = float(row["arrears"])
        rows.append(item)
    return rows


def build_monthly_rows(dates, amounts, arrears=None):
    metric_df = pd.DataFrame({"date": dates})
    if amounts is not None:
        metric_df["amount"] = amounts.reindex(dates.index).fillna(0)
    if arrears is not None:
        metric_df["arrears"] = arrears.reindex(dates.index).fillna(0)
    agg_dict = {"count": ("date", "size")}
    if amounts is not None:
        agg_dict["amount"] = ("amount", "sum")
    if arrears is not None:
        agg_dict["arrears"] = ("arrears", "sum")
    grouped = metric_df.groupby(metric_df["date"].dt.to_period("M")).agg(**agg_dict).sort_index()
    fiscal_start, current_period = get_fiscal_window()
    grouped = grouped.loc[(grouped.index >= fiscal_start) & (grouped.index <= current_period)]
    rows = []
    for label, row in grouped.iterrows():
        item = {"label": format_fiscal_month(label), "count": int(row["count"])}
        if "amount" in row:
            item["amount_total"] = float(row["amount"])
        if "arrears" in row:
            item["arrears_total"] = float(row["arrears"])
        rows.append(item)
    return rows


def build_commercial_mask(df: pd.DataFrame) -> pd.Series:
    columns = list(df.columns)
    commercial_cols = [
        col
        for col in [
            pick_column(columns, ["connection type", "bill type", "type"]),
            pick_column(columns, ["sector"]),
        ]
        if col
    ]
    if not commercial_cols:
        return pd.Series(False, index=df.index)
    mask = pd.Series(False, index=df.index)
    for col in commercial_cols:
        mask = mask | df[col].astype(str).str.lower().str.contains("commercial", na=False)
    return mask


def build_commercial_rows(df: pd.DataFrame, dates: pd.Series, amount_col: str | None, arrears_col: str | None):
    """Build commercial sector breakdown by locality — total and month-wise."""
    locality_col = pick_column(list(df.columns), ["locality"])
    if not locality_col:
        return None, None

    mask = build_commercial_mask(df)
    comm = df[mask].copy()
    if comm.empty:
        return [], []

    comm_dates = dates.reindex(comm.index).dropna()
    comm = comm.loc[comm_dates.index]
    comm_dates = comm_dates.loc[comm.index]

    comm["_locality"] = comm[locality_col].fillna("Unknown").astype(str).str.strip()
    if amount_col:
        comm["_amount"] = comm[amount_col].apply(clean_amount_value)
    if arrears_col:
        comm["_arrears"] = comm[arrears_col].apply(clean_amount_value)
    comm["_period"] = comm_dates.dt.to_period("M")

    fiscal_start, current_period = get_fiscal_window()
    comm = comm[(comm["_period"] >= fiscal_start) & (comm["_period"] <= current_period)]

    # Total by locality
    agg_t = {"count": ("_locality", "size")}
    if amount_col:
        agg_t["amount"] = ("_amount", "sum")
    if arrears_col:
        agg_t["arrears"] = ("_arrears", "sum")
    total_grouped = comm.groupby("_locality").agg(**agg_t).sort_index()
    total_rows = []
    for loc, row in total_grouped.iterrows():
        item = {"label": loc, "count": int(row["count"])}
        if "amount" in row:
            item["amount_total"] = float(row["amount"])
        if "arrears" in row:
            item["arrears_total"] = float(row["arrears"])
        total_rows.append(item)

    # Month-wise by locality
    agg_m = {"count": ("_locality", "size")}
    if amount_col:
        agg_m["amount"] = ("_amount", "sum")
    if arrears_col:
        agg_m["arrears"] = ("_arrears", "sum")
    monthly_grouped = comm.groupby(["_period", "_locality"]).agg(**agg_m).reset_index()
    monthly_data = {}
    for _, row in monthly_grouped.iterrows():
        month_label = format_fiscal_month(row["_period"])
        if month_label not in monthly_data:
            monthly_data[month_label] = []
        item = {"label": row["_locality"], "count": int(row["count"])}
        if "amount" in row:
            item["amount_total"] = float(row["amount"])
        if "arrears" in row:
            item["arrears_total"] = float(row["arrears"])
        monthly_data[month_label].append(item)

    return total_rows, monthly_data


def build_commercial_month_wise_summary(df: pd.DataFrame, dates: pd.Series, amount_col: str | None, arrears_col: str | None) -> list[dict]:
    """Build commercial month-wise summary — one row per fiscal month (July to June)."""
    mask = build_commercial_mask(df)
    comm = df[mask].copy()
    if comm.empty:
        return []

    comm_dates = dates.reindex(comm.index).dropna()
    comm = comm.loc[comm_dates.index]
    comm_dates = comm_dates.loc[comm.index]

    metric_df = pd.DataFrame({"date": comm_dates})
    if amount_col:
        metric_df["amount"] = comm[amount_col].apply(clean_amount_value).reindex(comm_dates.index).fillna(0)
    if arrears_col:
        metric_df["arrears"] = comm[arrears_col].apply(clean_amount_value).reindex(comm_dates.index).fillna(0)

    agg_dict = {"count": ("date", "size")}
    if amount_col is not None:
        agg_dict["amount"] = ("amount", "sum")
    if arrears_col is not None:
        agg_dict["arrears"] = ("arrears", "sum")
    grouped = metric_df.groupby(metric_df["date"].dt.to_period("M")).agg(**agg_dict).sort_index()

    fiscal_start, current_period = get_fiscal_window()
    grouped = grouped.loc[(grouped.index >= fiscal_start) & (grouped.index <= current_period)]

    rows = []
    for label, row in grouped.iterrows():
        item = {"label": format_fiscal_month(label), "count": int(row["count"])}
        if "amount" in row:
            item["amount_total"] = float(row["amount"])
        if "arrears" in row:
            item["arrears_total"] = float(row["arrears"])
        rows.append(item)
    return rows


def build_commercial_daily_income_rows(
    df: pd.DataFrame,
    dates: pd.Series,
    amount_col: str | None,
    arrears_col: str | None,
    areas_col: str | None,
) -> tuple[list[dict], str] | tuple[None, str]:
    columns = list(df.columns)
    required_cols = {
        "consumer_name": pick_column(columns, ["consumer name / f/h name", "consumer name", "name"]),
        "connection_no": pick_column(columns, ["connection no", "connection number", "old connection no"]),
        "sector": pick_column(columns, ["sector"]),
        "locality": pick_column(columns, ["locality"]),
    }
    if not required_cols["sector"] or not required_cols["locality"]:
        return None, "Arrears Received" if arrears_col else "Areas Received"

    metric_col = arrears_col or areas_col
    metric_label = "Arrears Received" if arrears_col else "Areas Received"
    mask = build_commercial_mask(df)
    comm = df[mask].copy()
    if comm.empty:
        return [], metric_label

    comm_dates = dates.reindex(comm.index).dropna()
    comm = comm.loc[comm_dates.index].copy()
    comm_dates = comm_dates.loc[comm.index]

    fiscal_start, current_period = get_fiscal_window()
    periods = comm_dates.dt.to_period("M")
    comm = comm[(periods >= fiscal_start) & (periods <= current_period)].copy()
    comm_dates = comm_dates.loc[comm.index]
    if comm.empty:
        return [], metric_label

    rows = []
    for idx in sorted(comm.index, key=lambda key: (comm_dates.loc[key].date(), str(key))):
        row = comm.loc[idx]
        rows.append(
            {
                "date": comm_dates.loc[idx].strftime("%d-%m-%Y"),
                "consumer_name": clean_cell(row.get(required_cols["consumer_name"])) if required_cols["consumer_name"] else "",
                "connection_no": clean_cell(row.get(required_cols["connection_no"])) if required_cols["connection_no"] else "",
                "sector": clean_cell(row.get(required_cols["sector"])),
                "locality": clean_cell(row.get(required_cols["locality"])),
                "metric_total": parse_number(row.get(metric_col)) if metric_col else 0,
                "amount_total": parse_number(row.get(amount_col)) if amount_col else 0,
            }
        )
    return rows, metric_label


def load_staff_assignment_rows() -> tuple[list[dict], list[dict]]:
    init_bill_list_db()
    with get_db() as conn:
        staff = [dict(row) for row in conn.execute("SELECT id, name FROM staff ORDER BY name").fetchall()]
        assignments = [
            dict(row)
            for row in conn.execute(
                """
                SELECT
                    sa.staff_id,
                    s.name AS staff_name,
                    sa.zone,
                    COALESCE(sa.sector, '') AS sector,
                    COALESCE(sa.locality, '') AS locality
                FROM staff_assignments sa
                JOIN staff s ON s.id = sa.staff_id
                ORDER BY s.name, sa.zone, sa.sector, sa.locality
                """
            ).fetchall()
        ]
    return staff, assignments


def get_auto_staff_override(sector: str, locality: str, connection_no: str, staff_list: list[dict]) -> dict | None:
    if not connection_no:
        return None
    conn_str = str(connection_no).strip()
    sector_norm = _normalize_sector_locality(sector)
    locality_norm = _normalize_sector_locality(locality)
    with get_db() as conn:
        rules = conn.execute(
            """
            SELECT staff_name, connection_min, connection_max, sector, locality
            FROM auto_assignment_rules
            ORDER BY connection_min
            """,
        ).fetchall()
    for rule in rules:
        rule_sector_norm = _normalize_sector_locality(rule["sector"])
        rule_locality_norm = _normalize_sector_locality(rule["locality"])
        if rule_sector_norm != sector_norm or rule_locality_norm != locality_norm:
            continue
        if conn_str >= rule["connection_min"]:
            if rule["connection_max"] is None or conn_str <= rule["connection_max"]:
                for s in staff_list:
                    if _normalize_staff_name(s["name"]) == _normalize_staff_name(rule["staff_name"]):
                        return {
                            "staff_id": s["id"],
                            "staff_name": s["name"],
                            "zone": "",
                            "sector": sector,
                            "locality": locality,
                        }
    return None


def match_staff_assignment(zone: str, sector: str, locality: str, assignments: list[dict], connection_no: str | None = None) -> dict | None:
    if connection_no:
        with get_db() as conn:
            staff_list = conn.execute("SELECT id, name FROM staff").fetchall()
        override = get_auto_staff_override(sector, locality, connection_no, staff_list)
        if override:
            return override
    zone = clean_cell(zone)
    sector = clean_cell(sector)
    locality = clean_cell(locality)
    zone_key = match_key(zone)
    sector_key = match_key(sector)
    locality_key = match_key(locality)

    deep_sector = _deep_normalize_sector(sector)
    deep_locality = _deep_normalize_sector(locality) if locality else ''
    sector_kw = _keyword_set(sector)

    alias_match = match_by_alias(sector, locality)
    if alias_match:
        for assignment in assignments:
            if (
                int(assignment.get("staff_id") or 0) == int(alias_match["staff_id"])
                and match_key(assignment["zone"]) == match_key(alias_match["zone"])
            ):
                return assignment
        return alias_match

    # Strategy 1: Exact zone+sector+locality match
    for assignment in assignments:
        if (
            match_key(assignment["zone"]) == zone_key
            and match_key(assignment["sector"]) == sector_key
            and match_key(assignment["locality"]) == locality_key
        ):
            return assignment

    # Strategy 2: Exact zone+sector match (assignment has no locality)
    for assignment in assignments:
        if (
            match_key(assignment["zone"]) == zone_key
            and match_key(assignment["sector"]) == sector_key
            and not clean_cell(assignment["locality"])
        ):
            return assignment

    # Strategy 3: Zone-only match (assignment has no sector or locality)
    for assignment in assignments:
        if (
            match_key(assignment["zone"]) == zone_key
            and not clean_cell(assignment["sector"])
            and not clean_cell(assignment["locality"])
        ):
            return assignment

    # Strategy 4: Deep-normalized sector match (with or without zone, ignore locality)
    deep_zone = _deep_normalize_sector(zone)
    for assignment in assignments:
        assign_deep_sector = _deep_normalize_sector(assignment["sector"]) if assignment["sector"] else ''
        if not assign_deep_sector:
            continue
        if deep_sector == assign_deep_sector:
            if not clean_cell(assignment["zone"]) or _deep_normalize_sector(assignment["zone"]) == deep_zone:
                return assignment

    # Strategy 5: Deep-normalized match swapping sector and locality fields
    if locality:
        for assignment in assignments:
            assign_deep_sector = _deep_normalize_sector(assignment["sector"]) if assignment["sector"] else ''
            if not assign_deep_sector:
                continue
            if deep_locality == assign_deep_sector:
                if not clean_cell(assignment["zone"]) or _deep_normalize_sector(assignment["zone"]) == deep_zone:
                    return assignment

    # Strategy 6: Same-zone keyword overlap on sector
    best_sector = None
    best_overlap = 0
    if sector_kw:
        for assignment in assignments:
            assign_sector = clean_cell(assignment["sector"])
            if not assign_sector:
                continue
            if clean_cell(assignment["zone"]) and match_key(assignment["zone"]) != zone_key:
                continue
            assign_kw = _keyword_set(assign_sector)
            if not assign_kw:
                continue
            overlap = len(sector_kw & assign_kw)
            if overlap > best_overlap:
                best_overlap = overlap
                best_sector = assign_sector
            elif overlap == best_overlap and overlap > 0 and best_sector:
                if len(assign_kw) < len(_keyword_set(best_sector)):
                    best_sector = assign_sector
        if best_overlap >= 2 and best_sector:
            for assignment in assignments:
                if match_key(assignment["sector"]) == match_key(best_sector):
                    return assignment

    # Strategy 7: Levenshtein fuzzy match on sector (same zone)
    for assignment in assignments:
        assign_sector = clean_cell(assignment["sector"])
        if not assign_sector:
            continue
        if clean_cell(assignment["zone"]) and match_key(assignment["zone"]) != zone_key:
            continue
        sector_len = max(len(sector_key), len(match_key(assign_sector)))
        max_dist = max(2, sector_len // 5)
        if _levenshtein(sector_key, match_key(assign_sector)) <= max_dist:
            return assignment

    # Strategy 8: Levenshtein fuzzy match using deep-normalized forms
    for assignment in assignments:
        assign_deep = _deep_normalize_sector(assignment["sector"]) if assignment["sector"] else ''
        if not assign_deep:
            continue
        if clean_cell(assignment["zone"]) and match_key(assignment["zone"]) != zone_key:
            continue
        max_dist = max(2, len(deep_sector) // 5)
        if _levenshtein(deep_sector, assign_deep) <= max_dist:
            return assignment

    best_assignment = None
    best_distance = 999
    for assignment in assignments:
        assign_sector = clean_cell(assignment["sector"])
        if not assign_sector:
            continue
        max_dist = max(2, len(deep_sector) // 4)
        distance = _levenshtein(deep_sector, _deep_normalize_sector(assign_sector))
        if distance <= max_dist and distance < best_distance:
            best_distance = distance
            best_assignment = assignment
    if best_assignment is not None:
        return best_assignment

    if sector_kw:
        best_assignment = None
        best_overlap = 0
        for assignment in assignments:
            assign_sector = clean_cell(assignment["sector"])
            if not assign_sector:
                continue
            assign_kw = _keyword_set(assign_sector)
            if not assign_kw:
                continue
            overlap = len(sector_kw & assign_kw)
            if overlap > best_overlap:
                best_overlap = overlap
                best_assignment = assignment
        if best_overlap >= 2 and best_assignment is not None:
            return best_assignment

    _UNMATCHED_LOG.append({
        "connection_no": connection_no or "",
        "zone": zone,
        "sector": sector,
        "locality": locality,
        "deep_sector": deep_sector,
    })

    if os.environ.get("FLASK_DEBUG") == "1" or os.environ.get("VERCEL") == "1":
        print(
            f"UNMATCHED_STAFF: conn={connection_no or ''} zone={zone!r} sector={sector!r} "
            f"locality={locality!r} deep_sector={deep_sector!r}"
        )

    return None


def get_staff_by_connection_rule(
    sector: str,
    locality: str,
    connection_no: str,
    staff_list: list[dict],
) -> dict | None:
    if not connection_no:
        return None
    conn_str = str(connection_no).strip()
    if not conn_str:
        return None
    sector_norm = _normalize_sector_locality(sector)
    locality_norm = _normalize_sector_locality(locality)
    for rule in AUTO_ASSIGNMENT_RULES:
        rule_sector_norm = _normalize_sector_locality(rule["sector"])
        rule_locality_norm = _normalize_sector_locality(rule["locality"])
        if rule_sector_norm != sector_norm or rule_locality_norm != locality_norm:
            continue
        if conn_str >= rule["connection_min"]:
            if rule["connection_max"] is None or conn_str <= rule["connection_max"]:
                for s in staff_list:
                    if _normalize_staff_name(s["name"]) == _normalize_staff_name(rule["staff_name"]):
                        return {"staff_id": s["id"], "staff_name": s["name"], "zone": "", "sector": sector, "locality": locality}
    return None


def build_daily_staff_receive_report(
    df: pd.DataFrame,
    dates: pd.Series,
    amount_col: str | None,
    arrears_col: str | None,
    areas_col: str | None,
) -> dict:
    columns = list(df.columns)
    sector_col = pick_column(columns, ["sector"])
    locality_col = pick_column(columns, ["locality"])
    if not sector_col or not locality_col:
        return {"summary_rows": [], "detail_rows": [], "metric_label": "Area Received", "date_range": ""}

    clear_unmatched_log()
    metric_col = arrears_col or areas_col
    connection_no_col = pick_column(columns, ["connection no", "connection number", "old connection no", "consumer no", "customer no", "connection id", "conn id"])
    staff, assignments = load_staff_assignment_rows()
    with get_db() as conn:
        locality_zones = {
            (row["sector"], row["locality"]): row["zone"]
            for row in conn.execute("SELECT sector, locality, zone FROM localities").fetchall()
        }
        sector_zones = {
            row["name"]: row["zone"]
            for row in conn.execute(
                f"""
                SELECT name, zone
                FROM sectors
                ORDER BY {zone_sort_expr('zone')}
                """
            ).fetchall()
        }

    staff_totals = {
        int(row["id"]): {
            "staff_id": int(row["id"]),
            "staff_name": row["name"],
            "zone": "",
            "bills": 0,
            "metric_total": 0.0,
            "amount_total": 0.0,
        }
        for row in staff
    }
    detail_groups: dict[int, dict] = {}

    valid_dates = dates.dropna()
    date_range = ""
    if not valid_dates.empty:
        min_date = valid_dates.min().strftime("%d-%m-%Y")
        max_date = valid_dates.max().strftime("%d-%m-%Y")
        date_range = min_date if min_date == max_date else f"{min_date} to {max_date}"

    for idx in valid_dates.index:
        row = df.loc[idx]
        amount_received = parse_number(row.get(amount_col)) if amount_col else 0.0
        metric_total = parse_number(row.get(metric_col)) if metric_col else 0.0
        if amount_col and amount_received <= 0:
            continue

        sector = clean_cell(row.get(sector_col)) or "Unknown"
        locality = clean_cell(row.get(locality_col)) or "Unknown"
        zone = locality_zones.get((sector, locality)) or sector_zones.get(sector) or infer_zone(sector, locality, row.to_dict())
        connection_no = clean_cell(row.get(connection_no_col)) if connection_no_col else None
        assignment = get_staff_by_connection_rule(sector, locality, connection_no, staff)
        if not assignment:
            unmatched_before = len(get_unmatched_log())
            assignment = match_staff_assignment(zone, sector, locality, assignments, connection_no)
            unmatched_after = len(get_unmatched_log())
            if unmatched_after > unmatched_before:
                for entry in get_unmatched_log()[unmatched_before:]:
                    consumer_name_col = pick_column(columns, ["consumer name / f/h name", "consumer name", "name"])
                    entry["consumer_name"] = clean_cell(row.get(consumer_name_col)) if consumer_name_col else ""
                    entry["amount_received"] = amount_received
                    entry["metric_total"] = metric_total
                    entry["date"] = str(valid_dates.loc[idx])
        if assignment:
            staff_id = int(assignment["staff_id"])
            staff_name = assignment["staff_name"]
            assigned_zone = assignment["zone"]
            staff_totals.setdefault(
                staff_id,
                {
                    "staff_id": staff_id,
                    "staff_name": staff_name,
                    "zone": assigned_zone,
                    "bills": 0,
                    "metric_total": 0.0,
                    "amount_total": 0.0,
                },
            )
        else:
            staff_id = 0
            staff_name = "Unassigned"
            assigned_zone = zone
            staff_totals.setdefault(
                staff_id,
                {
                    "staff_id": staff_id,
                    "staff_name": staff_name,
                    "zone": assigned_zone,
                    "bills": 0,
                    "metric_total": 0.0,
                    "amount_total": 0.0,
                },
            )

        total = staff_totals[staff_id]
        total["zone"] = total["zone"] or assigned_zone
        total["bills"] += 1
        total["metric_total"] += metric_total
        total["amount_total"] += amount_received

        group = detail_groups.setdefault(
            staff_id,
            {
                "staff_id": staff_id,
                "staff_name": staff_name,
                "zone": assigned_zone,
                "bills": 0,
                "metric_total": 0.0,
                "amount_total": 0.0,
                "sub_rows": {},
            },
        )
        group["zone"] = group["zone"] or assigned_zone
        group["bills"] += 1
        group["metric_total"] += metric_total
        group["amount_total"] += amount_received
        sub_key = normalise_sector(sector)
        sub = group["sub_rows"].setdefault(
            sub_key,
            {
                "zone": assigned_zone,
                "sector": sector,
                "locality": locality,
                "bills": 0,
                "metric_total": 0.0,
                "amount_total": 0.0,
            },
        )
        sub["bills"] += 1
        sub["metric_total"] += metric_total
        sub["amount_total"] += amount_received

    summary_rows = sorted(staff_totals.values(), key=lambda item: (item["zone"] or "ZZZ", item["staff_name"]))
    grouped_detail = sorted(
        (
            {
                "staff_id": g["staff_id"],
                "staff_name": g["staff_name"],
                "zone": g["zone"],
                "bills": g["bills"],
                "metric_total": g["metric_total"],
                "amount_total": g["amount_total"],
                "sub_rows": sorted(g["sub_rows"].values(), key=lambda s: (s["zone"], s["sector"], s["locality"])),
            }
            for g in detail_groups.values()
        ),
        key=lambda g: (g["staff_name"], g["zone"]),
    )
    detail_rows = [
        {
            "staff_id": g["staff_id"],
            "staff_name": g["staff_name"],
            "zone": sub["zone"],
            "sector": sub["sector"],
            "locality": sub["locality"],
            "bills": sub["bills"],
            "metric_total": sub["metric_total"],
            "amount_total": sub["amount_total"],
            "_group_bills": g["bills"],
            "_group_metric": g["metric_total"],
            "_group_amount": g["amount_total"],
        }
        for g in grouped_detail
        for sub in g["sub_rows"]
    ]
    return {
        "summary_rows": summary_rows,
        "detail_rows": detail_rows,
        "grouped_detail": grouped_detail,
        "metric_label": "Area Received",
        "date_range": date_range,
        "unmatched_log": get_unmatched_log(),
    }


# ---------------------------------------------------------------------------
# Summarize
# ---------------------------------------------------------------------------

MONTH_NAMES = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]


def clean_amount_value(val) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    s = str(val).strip()
    if not s or s.lower() in ("nan", "null", "none", ""):
        return 0.0
    s = re.sub(r"(?i)Rs\.?\s*", "", s)
    s = s.replace(",", "").replace('"', "").replace("'", "")
    s = s.strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def parse_date_dd_mm_yyyy(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "null", "none", ""):
        return None
    if "water bills receipts" in s.lower():
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    try:
        dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return dt.to_pydatetime()
    except Exception:
        pass
    return None


def build_receipt_monthly_rows(df: pd.DataFrame, date_col_name: str, arrears_col_name: str, amount_col_name: str) -> dict | None:
    month_map = {}
    total_rows = 0
    valid_rows = 0
    total_arrears = 0
    total_amount = 0

    for idx, row_ref in df.iterrows():
        total_rows += 1
        raw_date = row_ref[date_col_name]
        if raw_date is None or (isinstance(raw_date, float) and pd.isna(raw_date)):
            continue
        raw_date_str = str(raw_date).strip()
        if not raw_date_str or raw_date_str.lower() in ("nan", "null", "none", ""):
            continue
        if "water bills receipts" in raw_date_str.lower():
            continue

        dt = parse_date_dd_mm_yyyy(raw_date)
        if dt is None:
            continue

        valid_rows += 1
        month_key = f"{dt.year}-{dt.month:02d}"
        arrears_val = clean_amount_value(row_ref[arrears_col_name])
        amount_val = clean_amount_value(row_ref[amount_col_name])

        total_arrears += arrears_val
        total_amount += amount_val

        if month_key not in month_map:
            month_map[month_key] = {
                "date": dt,
                "label": f"{MONTH_NAMES[dt.month - 1]} {str(dt.year)[-2:]}",
                "_sort_key": dt.replace(day=1),
                "count": 0,
                "arrears_total": 0.0,
                "amount_total": 0.0,
            }
        month_map[month_key]["count"] += 1
        month_map[month_key]["arrears_total"] += arrears_val
        month_map[month_key]["amount_total"] += amount_val

    rows = sorted(month_map.values(), key=lambda r: r["_sort_key"])

    if not rows:
        return None

    min_date = rows[0]["_sort_key"]
    max_date = rows[-1]["_sort_key"]

    return {
        "rows": rows,
        "period_start": f"{MONTH_NAMES[min_date.month - 1]} {str(min_date.year)[-2:]}",
        "period_end": f"{MONTH_NAMES[max_date.month - 1]} {str(max_date.year)[-2:]}",
        "total_arrears": total_arrears,
        "total_amount": total_amount,
        "debug": {
            "total_rows_processed": total_rows,
            "valid_rows_used": valid_rows,
        },
    }


def summarize_dataframe(df: pd.DataFrame) -> dict:
    df = normalize_dataframe(df)
    columns = [str(col) for col in df.columns]

    date_col = pick_column(columns, ["received date", "receiving date", "date"])
    amount_col = pick_column(columns, ["amount received", "amount", "total amount", "total bill"])
    arrears_col = pick_column(columns, ["arrears"])
    areas_col = pick_column(columns, ["areas", "area"])
    sector_col = "sector" if "sector" in columns else None

    total_amount = None
    if amount_col:
        total_amount = float(df[amount_col].apply(clean_amount_value).sum())

    total_arrears = None
    if arrears_col:
        total_arrears = float(df[arrears_col].apply(clean_amount_value).sum())

    total_areas = None
    if areas_col:
        total_areas = float(pd.to_numeric(df[areas_col], errors="coerce").sum(skipna=True))

    clean_amount_debug = {}
    daily_rows, monthly_rows, date_note = [], [], None
    fiscal_row_count = None
    commercial_total, commercial_monthly = None, None
    commercial_daily_income, commercial_daily_metric_label = None, "Arrears Received" if arrears_col else "Areas Received"
    commercial_month_wise_summary = []
    daily_staff_receive = {"summary_rows": [], "detail_rows": [], "metric_label": "Area Received", "date_range": ""}
    dates = None

    if date_col:
        dates = parse_received_dates(df[date_col]).dropna()
        if dates.empty:
            date_note = f"No valid dates found in '{date_col}'."
        else:
            amounts_s = df[amount_col].apply(clean_amount_value) if amount_col else None
            arrears_s = df[arrears_col].apply(clean_amount_value) if arrears_col else None

            # Debug: compare wrong vs clean totals
            if amount_col:
                _wrong_amount = float(pd.to_numeric(df[amount_col], errors="coerce").sum(skipna=True))
                _clean_amount = float(df[amount_col].apply(clean_amount_value).sum())
                clean_amount_debug["amount_wrong"] = _wrong_amount
                clean_amount_debug["amount_clean"] = _clean_amount
                clean_amount_debug["amount_diff"] = _clean_amount - _wrong_amount
                if abs(_wrong_amount - _clean_amount) > 1:
                    print(f"CLEAN_AMOUNT_DEBUG: pd.to_numeric sum={_wrong_amount:.0f}, clean_amount_value sum={_clean_amount:.0f} (diff={_clean_amount - _wrong_amount:.0f})")
            if arrears_col:
                _wrong_arrears = float(pd.to_numeric(df[arrears_col], errors="coerce").sum(skipna=True))
                _clean_arrears = float(df[arrears_col].apply(clean_amount_value).sum())
                clean_amount_debug["arrears_wrong"] = _wrong_arrears
                clean_amount_debug["arrears_clean"] = _clean_arrears
                clean_amount_debug["arrears_diff"] = _clean_arrears - _wrong_arrears
                if abs(_wrong_arrears - _clean_arrears) > 1:
                    print(f"CLEAN_AMOUNT_DEBUG: pd.to_numeric sum={_wrong_arrears:.0f}, clean_amount_value sum={_clean_arrears:.0f} (diff={_clean_arrears - _wrong_arrears:.0f})")

            daily_rows = build_daily_rows(dates, amounts_s, arrears_s)
            monthly_rows = build_monthly_rows(dates, amounts_s, arrears_s)
            fiscal_row_count = sum(r.get("count", 0) for r in monthly_rows) if monthly_rows else 0
            commercial_total, commercial_monthly = build_commercial_rows(df, dates, amount_col, arrears_col)
            commercial_daily_income, commercial_daily_metric_label = build_commercial_daily_income_rows(
                df,
                dates,
                amount_col,
                arrears_col,
                areas_col,
            )
            commercial_month_wise_summary = build_commercial_month_wise_summary(df, dates, amount_col, arrears_col)
            daily_staff_receive = build_daily_staff_receive_report(
                df,
                dates,
                amount_col,
                arrears_col,
                areas_col,
            )
    else:
        date_note = "No received date column found."

    # Sector breakdown
    sector_rows, sector_note = [], None
    if sector_col and (amount_col or areas_col):
        sector_series = df[sector_col].fillna("Unknown").astype(str).str.strip().replace({"": "Unknown"})
        sector_df = pd.DataFrame({"sector": sector_series})
        if amount_col:
            sector_df["amount"] = df[amount_col].apply(clean_amount_value)
        if arrears_col:
            sector_df["arrears"] = df[arrears_col].apply(clean_amount_value)
        if areas_col:
            sector_df["areas"] = pd.to_numeric(df[areas_col], errors="coerce")
        sector_grouped = sector_df.groupby("sector").sum(numeric_only=True).sort_index()
        sector_rows = []
        for label, row in sector_grouped.iterrows():
            item = {"label": label, "count": 0}
            if "amount" in row:
                item["amount_total"] = float(row["amount"])
            if "arrears" in row:
                item["arrears_total"] = float(row["arrears"])
            if "areas" in row:
                item["areas_total"] = float(row["areas"])
            sector_rows.append(item)
    elif not sector_col:
        sector_note = "No sector column found."

    # Receipt month-wise report (new format: Received Date + Arrears + Amount Received)
    receipt_info = None
    if "received date" in columns and arrears_col is not None and amount_col is not None:
        receipt_info = build_receipt_monthly_rows(df, "received date", arrears_col, amount_col)

    has_receipt_format = receipt_info is not None

    return {
        "row_count": int(len(df)),
        "columns": columns,
        "date_column": date_col,
        "amount_column": amount_col,
        "arrears_column": arrears_col,
        "areas_column": areas_col,
        "sector_column": sector_col,
        "total_amount": total_amount,
        # The dashboard cards animate from these raw values after heavy uploads.
        # Keep them in the result object so generated pages do not display Rs. 0.
        "total_amount_raw": total_amount or 0,
        "total_amount_formatted": fmt(total_amount),
        "total_arrears": total_arrears,
        "total_arrears_raw": total_arrears or 0,
        "total_arrears_formatted": fmt(total_arrears),
        "total_areas": total_areas,
        "total_areas_raw": total_areas or 0,
        "total_areas_formatted": fmt(total_areas),
        "daily_rows": daily_rows,
        "monthly_rows": monthly_rows,
        "sector_rows": sector_rows,
        "date_note": date_note,
        "sector_note": sector_note,
        "fiscal_row_count": fiscal_row_count,
        "has_amount": amount_col is not None,
        "has_arrears": arrears_col is not None,
        "has_areas": areas_col is not None,
        "has_date": date_col is not None,
        "commercial_total": commercial_total,
        "commercial_monthly": commercial_monthly,
        "commercial_daily_income": commercial_daily_income,
        "commercial_daily_metric_label": commercial_daily_metric_label,
        "commercial_month_wise_summary": commercial_month_wise_summary,
        "daily_staff_receive": daily_staff_receive,
        "has_receipt_format": has_receipt_format,
        "receipt_monthly_rows": (receipt_info or {}).get("rows", []),
        "receipt_period_start": (receipt_info or {}).get("period_start", ""),
        "receipt_period_end": (receipt_info or {}).get("period_end", ""),
        "receipt_total_arrears": (receipt_info or {}).get("total_arrears", 0),
        "receipt_total_amount": (receipt_info or {}).get("total_amount", 0),
        "receipt_total_arrears_fmt": fmt((receipt_info or {}).get("total_arrears", 0)),
        "receipt_total_amount_fmt": fmt((receipt_info or {}).get("total_amount", 0)),
        "receipt_debug": (receipt_info or {}).get("debug", {}),
        "clean_amount_debug": clean_amount_debug,
    }


# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------

ACCENT = colors.HexColor("#2f6f6d")
ACCENT2 = colors.HexColor("#d28b62")
HEADER_BG = colors.HexColor("#2f6f6d")
HEADER_FG = colors.white
ALT_ROW = colors.HexColor("#f4f1ea")
BORDER_CLR = colors.HexColor("#d9d2c6")


def _make_pdf_table(
    data_rows,
    col_widths=None,
    first_col_left=False,
    left_cols=None,
    header_font_size=12,
    body_font_size=11,
    cell_padding=8,
):
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
        ("TEXTCOLOR", (0, 0), (-1, 0), HEADER_FG),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), header_font_size),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), body_font_size),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_CLR),
        ("TOPPADDING", (0, 0), (-1, -1), cell_padding),
        ("BOTTOMPADDING", (0, 0), (-1, -1), cell_padding),
        ("LEFTPADDING", (0, 0), (-1, -1), max(3, cell_padding - 2)),
        ("RIGHTPADDING", (0, 0), (-1, -1), max(3, cell_padding - 2)),
    ])
    # Alternate row colors
    for i in range(1, len(data_rows)):
        if i % 2 == 0:
            style.add("BACKGROUND", (0, i), (-1, i), ALT_ROW)
    # Bold last row (grand total)
    if len(data_rows) > 2:
        last = len(data_rows) - 1
        style.add("FONTNAME", (0, last), (-1, last), "Helvetica-Bold")
        style.add("FONTSIZE", (0, last), (-1, last), body_font_size)
        style.add("BACKGROUND", (0, last), (-1, last), colors.HexColor("#e6d8c8"))

    for idx, row in enumerate(data_rows[1:], start=1):
        row_text = " ".join(str(cell) for cell in row)
        if " Total" in row_text or "Grand Total" in row_text:
            style.add("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold")
            style.add("FONTSIZE", (0, idx), (-1, idx), body_font_size)
            style.add("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#e6d8c8"))
            style.add("TEXTCOLOR", (0, idx), (-1, idx), colors.black)
            style.add("BOTTOMPADDING", (0, idx), (-1, idx), cell_padding + 6)

    if first_col_left:
        style.add("ALIGN", (0, 0), (0, -1), "LEFT")
    for col in left_cols or []:
        style.add("ALIGN", (col, 1), (col, -1), "LEFT")

    t = Table(data_rows, colWidths=col_widths, repeatRows=1)
    t.setStyle(style)
    return t


# ---------------------------------------------------------------------------
# Export column selection helpers
# ---------------------------------------------------------------------------

def parse_export_cols(cols_param, col_map, headers, rows):
    if not cols_param:
        return headers, rows
    selected_keys = [k.strip() for k in cols_param.split(",") if k.strip()]
    selected_indices = []
    filtered_headers = []
    for key in selected_keys:
        if key in col_map:
            idx = col_map[key]
            selected_indices.append(idx)
            filtered_headers.append(headers[idx])
    if not selected_indices or selected_indices == list(range(len(headers))):
        return headers, rows
    filtered_rows = [[row[i] for i in selected_indices] for row in rows]
    return filtered_headers, filtered_rows


def _export_row_selection() -> tuple[str, set[str]]:
    """Read bill-list row checkbox selection from export query params."""
    if not has_request_context():
        return "", set()
    mode = (request.args.get("row_mode") or "").strip().lower()
    if mode not in ("include", "exclude"):
        return "", set()
    keys = {str(key) for key in request.args.getlist("row_key") if str(key).strip()}
    return mode, keys


def _filter_rows_by_selection(rows: list, key_func) -> list:
    """Keep only bill-list rows selected by the page checkboxes before exporting."""
    mode, keys = _export_row_selection()
    if not mode:
        return rows
    if mode == "include":
        return [row for row in rows if key_func(row) in keys]
    return [row for row in rows if key_func(row) not in keys]


def _staff_row_key(row: list) -> str:
    return f"{row[1]}|||{row[2]}|||{row[3]}|||{row[4]}"


def _selection_has_filter() -> bool:
    mode, _ = _export_row_selection()
    return bool(mode)


# Column key maps for dashboard cards (index.html)
CARD_COL_MAPS = {
    "monthly": {"month": 0, "count": 1, "arrearsReceived": 2, "currentAmountReceived": 3, "amountReceived": 4},
    "daily": {"date": 0, "count": 1, "arrearsReceived": 2, "amountReceived": 3},
    "sector": {"sector": 0, "arrearsReceived": 1, "amountReceived": 2},
    "receipt-monthly": {"month": 0, "noOfBills": 1, "arrearsReceived": 2, "currentAmountReceived": 3, "amountReceived": 4},
    "commercial-daily-income": {"sr": 0, "date": 1, "consumerName": 2, "connectionNo": 3, "sector": 4, "locality": 5, "arrearsReceived": 6, "amountReceived": 7},
    "connection-type-summary": {"name": 0, "noOfBills": 1, "arrearsReceived": 2, "currentAmountReceived": 3, "amountReceived": 4},
    "daily-staff-receive": {"sr": 0, "staffName": 1, "bills": 2, "arrears": 3, "amount": 4},
}

def _get_card_col_map(card, r):
    if card in CARD_COL_MAPS:
        return CARD_COL_MAPS[card]
    if card in ("commercial", "commercial-total"):
        if r.get("has_arrears"):
            return {"locality": 0, "count": 1, "arrearsReceived": 2, "amountReceived": 3}
        return {"locality": 0, "count": 1, "amountReceived": 2}
    if card == "commercial-monthly":
        if r.get("has_arrears"):
            return {"month": 0, "locality": 1, "count": 2, "arrearsReceived": 3, "amountReceived": 4}
        return {"month": 0, "locality": 1, "count": 2, "amountReceived": 3}
    if card == "commercial-month-wise":
        return {"month": 0, "noOfBills": 1, "arrearsReceived": 2, "currentAmountReceived": 3, "amountReceived": 4}
    return None

def _filter_card_export(cols_param, col_map, headers, rows, grand=None):
    if not cols_param or not col_map:
        return headers, rows, grand
    selected_keys = [k.strip() for k in cols_param.split(",") if k.strip()]
    selected_indices = []
    filtered_headers = []
    for key in selected_keys:
        if key in col_map:
            idx = col_map[key]
            selected_indices.append(idx)
            filtered_headers.append(headers[idx])
    if not selected_indices or selected_indices == list(range(len(headers))):
        return headers, rows, grand
    filtered_rows = [[row[i] for i in selected_indices] for row in rows]
    filtered_grand = [grand[i] for i in selected_indices] if grand else None
    return filtered_headers, filtered_rows, filtered_grand

# Column key maps for each export route
# Sector-wise report: Sr, Sector, Total Bills, Received Bills, Remaining Bills, Total Received Amount, Pending Amount
SECTOR_COL_MAP = {"sr": 0, "sector": 1, "totalBills": 2, "receivedBills": 3, "remainingBills": 4, "totalReceivedAmount": 5, "pendingAmount": 6}

# Zone report (CSV/Excel): Sr, Zone, Sector, Total Bills, Received Bills, Remaining Bills, Amount Received, Pending Amount
ZONE_COL_MAP = {"sr": 0, "zone": 1, "sector": 2, "totalBills": 3, "receivedBills": 4, "remainingBills": 5, "totalReceivedAmount": 6, "pendingAmount": 7}

# Zone PDF detail: Sr, Sector, Total Bills, Received Bills, Remaining Bills, Amount Received, Pending Amount
ZONE_PDF_COL_MAP = {"sr": 0, "sector": 1, "totalBills": 2, "receivedBills": 3, "remainingBills": 4, "totalReceivedAmount": 5, "pendingAmount": 6}

# Zone PDF summary: Sr, Zone, Total Bills, Received Bills, Total Amount, Arrears Received, Total Received, Pending Amount
ZONE_SUMMARY_COL_MAP = {"sr": 0, "zone": 1, "totalBills": 2, "receivedBills": 3, "totalAmount": 4, "arrearsReceived": 5, "totalReceivedAmount": 6, "pendingAmount": 7}

# Unpaid amount summary: Bills, Total Bill Amount, Total Arrears Amount, Current Bill Amount
UNPAID_SUMMARY_COL_MAP = {"bills": 0, "totalBillAmount": 1, "totalArrearsAmount": 2, "currentBillAmount": 3}

# Unpaid amount section (sector/zone/staff): Sr, Name, Bills, Total Bill Amount, Total Arrears Amount, Current Bill Amount
UNPAID_SECTION_COL_MAP = {"sr": 0, "name": 1, "bills": 2, "totalBillAmount": 3, "totalArrearsAmount": 4, "currentBillAmount": 5}

# Staff report detail: Sr, Staff, Zone, Sector, Locality, Total Bills, Received Bills, Remaining Bills, Amount Received, Pending Amount
STAFF_COL_MAP = {"sr": 0, "staff": 1, "zone": 2, "sector": 3, "locality": 4, "totalBills": 5, "receivedBills": 6, "remainingBills": 7, "totalReceivedAmount": 8, "pendingAmount": 9}
STAFF_SUMMARY_COL_MAP = {"sr": 0, "staffName": 1, "totalBills": 2, "receivedBills": 3, "remainingBills": 4, "totalAmount": 5, "amountReceived": 6, "pendingAmount": 7}
PITCH_COL_MAP = {"sr": 0, "staffName": 1, "totalBills": 2, "receivedBills": 3, "remainingBills": 4, "totalAmount": 5, "amountReceived": 6, "currentBillAmount": 7}

# Staff PDF detail: Sr, Sector, Locality, Total Bills, Received Bills, Remaining Bills, Amount Received, Pending Amount
STAFF_PDF_COL_MAP = {"sr": 0, "sector": 1, "locality": 2, "totalBills": 3, "receivedBills": 4, "remainingBills": 5, "totalReceivedAmount": 6, "pendingAmount": 7}

# Advanced filter CSV/Excel: Bill No, Reference No, Connection No, Sector, Locality, Zone, Total Bill, Arrears, Amount Received, Outstanding Amount, Status
ADV_COL_MAP = {"billNo": 0, "referenceNo": 1, "connectionNo": 2, "sector": 3, "locality": 4, "zone": 5, "totalBills": 6, "arrearsReceived": 7, "totalReceivedAmount": 8, "outstanding": 9, "status": 10, "mobileNo": 11}

# Advanced filter PDF: Sr, Bill No, Connection No, Sector, Locality, Zone, Total Bill, Amount Received, Outstanding, Status
ADV_PDF_COL_MAP = {"sr": 0, "billNo": 1, "connectionNo": 2, "sector": 3, "locality": 4, "zone": 5, "totalBills": 6, "totalReceivedAmount": 7, "outstanding": 8, "status": 9, "mobileNo": 10}

# Summary report: Name, Count, Total Amount
SUMMARY_COL_MAP = {"name": 0, "count": 1, "totalAmount": 2}

# Column map for connection type summary (shared between standalone export and month-wise attachment)
CONN_SUMMARY_COL_MAP = {"name": 0, "noOfBills": 1, "arrearsReceived": 2, "currentAmountReceived": 3, "amountReceived": 4}

# Daily Staff Receive summary: Sr, Staff Name, No. of Bills Received, Arrears Received, Total Amount Received
DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP = {"sr": 0, "staffName": 1, "bills": 2, "arrears": 3, "amount": 4}

# Daily Staff Receive detail: Staff Name, Zone, Sr, Sector, Locality, Bills, Arrears Received, Received Amount
DAILY_STAFF_RECEIVE_DETAIL_COL_MAP = {"staffName": 0, "zone": 1, "sr": 2, "sector": 3, "locality": 4, "bills": 5, "arrears": 6, "amount": 7}


def normalise_sector(name: str) -> str:
    """Normalise a sector name for grouping: trim, lowercase, collapse spaces."""
    return " ".join((name or "").strip().lower().split())


def merge_sector_rows(rows: list[dict], sector_key: str = "sector", locality_key: str = "locality",
                      bills_key: str = "bills", metric_key: str = "metric_total", amount_key: str = "amount_total") -> list[dict]:
    """Merge rows that share the same normalised sector name.

    For each unique sector:
    - Keeps the first locality as representative text
    - Sums bills, metric_total, amount_total
    - Returns one merged row per unique sector
    """
    merged: dict[str, dict] = {}
    for row in rows:
        norm = normalise_sector(row.get(sector_key) or "")
        if norm not in merged:
            merged[norm] = {
                "_sector": row.get(sector_key) or "",
                "_locality": row.get(locality_key) or "",
                "_bills": 0,
                "_metric": 0.0,
                "_amount": 0.0,
            }
        m = merged[norm]
        m["_bills"] += row.get(bills_key, 0) if isinstance(row.get(bills_key), (int, float)) else parse_number(str(row.get(bills_key, 0)).replace(",", ""))
        m["_metric"] += row.get(metric_key, 0) if isinstance(row.get(metric_key), (int, float)) else parse_number(str(row.get(metric_key, 0)).replace(",", ""))
        m["_amount"] += row.get(amount_key, 0) if isinstance(row.get(amount_key), (int, float)) else parse_number(str(row.get(amount_key, 0)).replace(",", ""))
    result = []
    for norm, m in merged.items():
        result.append({
            "sector": m["_sector"],
            "locality": m["_locality"],
            "bills": m["_bills"],
            "metric_total": m["_metric"],
            "amount_total": m["_amount"],
        })
    return result


def merge_sector_list_rows(rows: list[list], sector_idx: int, locality_idx: int,
                           numeric_indices: list[int]) -> list[list]:
    """Merge list-of-lists rows by normalised sector. Keeps first locality, sums numeric columns."""
    merged: dict[str, dict] = {}
    for row in rows:
        norm = normalise_sector(str(row[sector_idx]) if sector_idx < len(row) else "")
        if norm not in merged:
            merged[norm] = {"_row": list(row), "_seen": True}
        m = merged[norm]
        for ni in numeric_indices:
            if ni < len(row):
                m["_row"][ni] = parse_number(str(m["_row"][ni]).replace(",", "")) + parse_number(str(row[ni]).replace(",", ""))
    result = []
    for m in merged.values():
        out = list(m["_row"])
        for ni in numeric_indices:
            if ni < len(out):
                out[ni] = fmt(out[ni]) if isinstance(out[ni], (int, float)) else out[ni]
        result.append(out)
    return result


def build_connection_summary(r):
    """Compute connection type collection summary from results dict."""
    source_rows = r.get("receipt_monthly_rows") or r.get("monthly_rows") or []
    overall_count = sum(row.get("count", 0) for row in source_rows)
    overall_arrears = sum(row.get("arrears_total", 0) for row in source_rows)
    overall_amount = sum(row.get("amount_total", 0) for row in source_rows)
    overall_current = overall_amount - overall_arrears
    comm_count = sum(row.get("count", 0) for row in (r.get("commercial_total") or []))
    comm_arrears = sum(row.get("arrears_total", 0) for row in (r.get("commercial_total") or []))
    comm_amount = sum(row.get("amount_total", 0) for row in (r.get("commercial_total") or []))
    comm_current = comm_amount - comm_arrears
    res_count = overall_count - comm_count
    res_arrears = overall_arrears - comm_arrears
    res_amount = overall_amount - comm_amount
    res_current = res_amount - res_arrears
    headers = ["Name", "No. of Bills", "Arrears Received", "Current Amount Received", "Amount Received"]
    rows = [
        ["Normal / Residential Connections", fmt(res_count), fmt(res_arrears), fmt(res_current), fmt(res_amount)],
        ["Commercial Connections", fmt(comm_count), fmt(comm_arrears), fmt(comm_current), fmt(comm_amount)],
    ]
    grand = ["Grand Total", fmt(overall_count), fmt(overall_arrears), fmt(overall_current), fmt(overall_amount)]
    return headers, rows, grand


def generate_card_pdf(
    title,
    summary_lines,
    headers,
    rows,
    grand_total_row=None,
    pagesize=A4,
    col_widths=None,
    first_col_left=False,
    left_cols=None,
    header_font_size=12,
    body_font_size=11,
    cell_padding=8,
    extra_section=None,
    compact=False,
):
    buf = io.BytesIO()
    if compact:
        left_margin = right_margin = 12 * mm
        top_margin = 15 * mm
        bottom_margin = 15 * mm
        title_fs = 16
        title_sa = 5*mm
        summary_fs = 10
        summary_sa = 2*mm
        summary_leading = 14
        summary_spacer = 5*mm
        section_fs = 12
        section_sb = 4*mm
        section_sa = 2*mm
        hdr_fs = 10
        bdy_fs = 10
        cp = 8
        extra_spacer = 4*mm
    else:
        left_margin = right_margin = 15 * mm
        top_margin = 20 * mm
        bottom_margin = 15 * mm
        title_fs = 20
        title_sa = 6*mm
        summary_fs = 11
        summary_sa = 2*mm
        summary_leading = 16
        summary_spacer = 8*mm
        section_fs = 14
        section_sb = 6*mm
        section_sa = 3*mm
        hdr_fs = header_font_size
        bdy_fs = body_font_size
        cp = cell_padding
        extra_spacer = 2*mm

    doc = SimpleDocTemplate(
        buf,
        pagesize=pagesize,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
        leftMargin=left_margin,
        rightMargin=right_margin,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("PDFTitle", parent=styles["Heading1"], fontSize=title_fs,
                                  textColor=ACCENT, alignment=1, spaceAfter=title_sa,
                                  fontName="Helvetica-Bold")
    summary_style = ParagraphStyle("PDFSummary", parent=styles["Normal"], fontSize=summary_fs,
                                    alignment=0, spaceAfter=summary_sa,
                                    textColor=colors.HexColor("#333333"),
                                    leading=summary_leading)
    section_style = ParagraphStyle("PDFSection", parent=styles["Heading3"], fontSize=section_fs,
                                    textColor=ACCENT, alignment=0, spaceBefore=section_sb,
                                    spaceAfter=section_sa, fontName="Helvetica-Bold")

    elements = [Paragraph(title, title_style)]
    if summary_lines:
        for line in summary_lines:
            elements.append(Paragraph(line, summary_style))
        elements.append(Spacer(1, summary_spacer))

    header_cell_style = ParagraphStyle(
        "PDFHeaderCell",
        parent=styles["Normal"],
        fontSize=hdr_fs,
        leading=hdr_fs + 2,
        alignment=1,
        textColor=HEADER_FG,
        fontName="Helvetica-Bold",
    )
    wrapped_headers = [Paragraph(str(header), header_cell_style) for header in headers]

    data = [wrapped_headers] + rows
    if grand_total_row:
        data.append(grand_total_row)

    page_w = pagesize[0] - left_margin - right_margin
    n_cols = len(headers)
    if col_widths is None:
        col_w = page_w / n_cols
        col_widths = [col_w] * n_cols
    elif len(col_widths) != n_cols:
        col_w = page_w / n_cols
        col_widths = [col_w] * n_cols
    t = _make_pdf_table(
        data,
        col_widths=col_widths,
        first_col_left=first_col_left,
        left_cols=left_cols,
        header_font_size=hdr_fs,
        body_font_size=bdy_fs,
        cell_padding=cp,
    )
    elements.append(t)

    # Attach extra section (e.g. Connection Type Summary) below main table
    # Uses KeepTogether so title + table stay on same page; flows to next page only when needed
    if extra_section is not None:
        extra_headers = extra_section["headers"]
        extra_rows = extra_section["rows"]
        extra_grand = extra_section.get("grand")
        # Build wrapped header cells without CJK word-wrap to prevent mid-word breaks
        hdr_style = ParagraphStyle(
            "PDFExtraHeaderCell",
            parent=getSampleStyleSheet()["Normal"],
            fontSize=hdr_fs,
            leading=hdr_fs + 2,
            alignment=1,
            textColor=HEADER_FG,
            fontName="Helvetica-Bold",
        )
        extra_data = [[Paragraph(str(h), hdr_style) for h in extra_headers]]
        extra_data.extend(wrap_pdf_body_cells(extra_rows, font_size=bdy_fs,
                                              left_columns={0}))
        if extra_grand:
            extra_data.append(wrap_pdf_body_cells([extra_grand], font_size=bdy_fs,
                                                  left_columns={0})[0])
        n_extra = len(extra_headers)
        extra_page_w = pagesize[0] - left_margin - right_margin
        if n_extra == 5:
            extra_col_widths = [extra_page_w * 0.26, extra_page_w * 0.14,
                                extra_page_w * 0.20, extra_page_w * 0.20, extra_page_w * 0.20]
        elif n_extra == 4:
            extra_col_widths = [extra_page_w * 0.30, extra_page_w * 0.16,
                                extra_page_w * 0.27, extra_page_w * 0.27]
        elif n_extra == 3:
            extra_col_widths = [extra_page_w * 0.40, extra_page_w * 0.28,
                                extra_page_w * 0.32]
        elif n_extra == 2:
            extra_col_widths = [extra_page_w * 0.55, extra_page_w * 0.45]
        else:
            extra_col_widths = [extra_page_w / n_extra] * n_extra
        extra_table = _make_pdf_table(
            extra_data, col_widths=extra_col_widths,
            first_col_left=True, left_cols=[0],
            header_font_size=hdr_fs,
            body_font_size=bdy_fs,
            cell_padding=cp,
        )
        extra_section_elements = []
        extra_section_elements.append(Spacer(1, extra_spacer))
        extra_section_elements.append(Paragraph("Summary", section_style))
        extra_section_elements.append(extra_table)
        elements.append(KeepTogether(extra_section_elements))

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


def wrap_pdf_table_cells(rows: list[list], font_size: int = 7) -> list[list]:
    return wrap_pdf_body_cells(rows, font_size=font_size)


def wrap_pdf_header_cells(headers: list[str], font_size: int = 13) -> list:
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        "PDFWrappedHeaderCell",
        parent=styles["Normal"],
        fontSize=font_size,
        leading=font_size + 2,
        alignment=1,
        textColor=HEADER_FG,
        fontName="Helvetica-Bold",
        wordWrap="CJK",
    )
    return [Paragraph(str(header or ""), header_style) for header in headers]


def is_large_pdf_text(value, threshold: int = 22) -> bool:
    text = clean_cell(value)
    if not text:
        return False
    if parse_number(text.replace(",", "")) and re.fullmatch(r"[\d,.\-\s]+", text):
        return False
    return len(text) > threshold or "\n" in text


_BRACKET_RE = re.compile(r"(\([^()]+\))")


def _bracket_rich_text(text: str, base_style: ParagraphStyle, small_size: int) -> list:
    """Split text into normal and bracket segments for rich text rendering.

    Returns a list of (text, style) tuples.  Text inside (...) is rendered
    at *small_size* while the rest uses *base_style*.
    """
    parts = _BRACKET_RE.split(text)
    segments = []
    for part in parts:
        if not part:
            continue
        if part.startswith("(") and part.endswith(")"):
            small_style = ParagraphStyle(
                "BracketSmall",
                parent=base_style,
                fontSize=small_size,
                leading=small_size + 1,
            )
            segments.append((part, small_style))
        else:
            segments.append((part, base_style))
    return segments


def wrap_pdf_body_cells(
    rows: list[list],
    font_size: int = 8,
    large_text_threshold: int = 22,
    left_columns: set[int] | None = None,
    bold_rows: set[int] | None = None,
    bracket_cols: set[int] | None = None,
) -> list[list]:
    styles = getSampleStyleSheet()
    wrapped = []
    left_columns = left_columns or set()
    bold_rows = bold_rows or set()
    bracket_cols = bracket_cols or set()
    for ri, row in enumerate(rows):
        wrapped_row = []
        for col_idx, value in enumerate(row):
            if isinstance(value, Paragraph):
                wrapped_row.append(value)
                continue
            is_bold = ri in bold_rows
            align = 0 if col_idx in left_columns or is_large_pdf_text(value, large_text_threshold) else 1
            font_name = "Helvetica-Bold" if is_bold else "Helvetica"
            cell_style = ParagraphStyle(
                "PDFBodyCellLeft" if align == 0 else "PDFBodyCellCenter",
                parent=styles["Normal"],
                fontName=font_name,
                fontSize=font_size,
                leading=font_size + 1,
                alignment=align,
                wordWrap="CJK",
            )
            cell_text = str(value or "").replace("\n", "<br/>")
            # Render bracket text at a smaller font size for Sector/Locality columns
            if bracket_cols and col_idx in bracket_cols and "(" in cell_text:
                segments = _bracket_rich_text(cell_text, cell_style, max(7, font_size - 2))
                # Build a mini-ParagraphStyle that supports mixed font sizes
                # by concatenating with <font> tags
                rich_parts = []
                for seg_text, seg_style in segments:
                    if seg_style.fontName == "Helvetica-Bold":
                        rich_parts.append(f'<b><font size="{seg_style.fontSize}">{seg_text}</font></b>')
                    else:
                        rich_parts.append(f'<font size="{seg_style.fontSize}">{seg_text}</font>')
                wrapped_row.append(Paragraph("".join(rich_parts), cell_style))
            else:
                wrapped_row.append(Paragraph(cell_text, cell_style))
        wrapped.append(wrapped_row)
    return wrapped


def generate_commercial_pdf(
    summary_lines,
    total_headers,
    total_rows,
    total_grand,
    monthly_data,
    has_arrears,
):
    buf = io.BytesIO()
    left_margin = 10 * mm
    right_margin = 10 * mm
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=20 * mm,
        bottomMargin=15 * mm,
        leftMargin=left_margin,
        rightMargin=right_margin,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PDFTitle",
        parent=styles["Heading1"],
        fontSize=20,
        textColor=ACCENT,
        alignment=1,
        spaceAfter=6 * mm,
        fontName="Helvetica-Bold",
    )
    summary_style = ParagraphStyle(
        "PDFSummary",
        parent=styles["Normal"],
        fontSize=11,
        alignment=0,
        spaceAfter=2 * mm,
        textColor=colors.HexColor("#333333"),
        leading=16,
    )
    section_style = ParagraphStyle(
        "PDFSection",
        parent=styles["Heading3"],
        fontSize=13,
        textColor=ACCENT,
        alignment=0,
        spaceBefore=6 * mm,
        spaceAfter=3 * mm,
        fontName="Helvetica-Bold",
    )

    elements = [Paragraph("Commercial Sector — Locality Report", title_style)]
    for line in summary_lines:
        elements.append(Paragraph(line, summary_style))
    elements.append(Spacer(1, 6 * mm))

    data = [total_headers] + total_rows
    if total_grand:
        data.append(total_grand)

    page_w = A4[0] - left_margin - right_margin
    if len(total_headers) == 4:
        col_widths = [page_w * 0.42, page_w * 0.14, page_w * 0.22, page_w * 0.22]
    elif len(total_headers) == 3 and has_arrears:
        col_widths = [page_w * 0.48, page_w * 0.26, page_w * 0.26]
    elif len(total_headers) == 3:
        col_widths = [page_w * 0.55, page_w * 0.18, page_w * 0.27]
    else:
        col_widths = [page_w * 0.62, page_w * 0.38]

    elements.append(_make_pdf_table(data, col_widths=col_widths, first_col_left=True))

    if monthly_data:
        elements.append(Spacer(1, 4 * mm))
        for month_label, rows in monthly_data.items():
            elements.append(Paragraph(month_label, section_style))
            month_headers = total_headers
            month_data = [month_headers] + rows
            elements.append(
                _make_pdf_table(month_data, col_widths=col_widths, first_col_left=True)
            )

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


def generate_commercial_monthly_pdf(summary_lines, monthly_sections, overall_total, has_arrears,
                                     pdf_headers=None, col_widths=None):
    buf = io.BytesIO()
    left_margin = 15 * mm
    right_margin = 15 * mm
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=20 * mm,
        bottomMargin=15 * mm,
        leftMargin=left_margin,
        rightMargin=right_margin,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("PDFTitle", parent=styles["Heading1"], fontSize=20, textColor=ACCENT, alignment=1, spaceAfter=6 * mm, fontName="Helvetica-Bold")
    summary_style = ParagraphStyle("PDFSummary", parent=styles["Normal"], fontSize=11, alignment=0, spaceAfter=2 * mm, textColor=colors.HexColor("#333333"), leading=16)
    section_style = ParagraphStyle("PDFSection", parent=styles["Heading3"], fontSize=14, textColor=ACCENT, alignment=1, spaceBefore=5 * mm, spaceAfter=3 * mm, fontName="Helvetica-Bold")

    if pdf_headers is None:
        pdf_headers = ["Locality", "No. of Bills"] + (["Arrears Received"] if has_arrears else []) + ["Amount Received"]
    page_w = A4[0] - left_margin - right_margin
    if col_widths is None:
        n = len(pdf_headers)
        if n == 4:
            col_widths = [page_w * 0.38, page_w * 0.20, page_w * 0.21, page_w * 0.21]
        elif n == 3:
            col_widths = [page_w * 0.48, page_w * 0.26, page_w * 0.26]
        elif n == 2:
            col_widths = [page_w * 0.62, page_w * 0.38]
        else:
            col_widths = [page_w / n] * n

    elements = [Paragraph("Commercial Sector - Month-wise Locality Report", title_style)]
    for line in summary_lines:
        elements.append(Paragraph(line, summary_style))
    elements.append(Spacer(1, 5 * mm))

    for section in monthly_sections:
        elements.append(Paragraph(section["month"], section_style))
        data = [wrap_pdf_header_cells(pdf_headers, font_size=11)]
        data.extend(wrap_pdf_body_cells(section["rows"], font_size=9, left_columns={0}))
        data.append(wrap_pdf_body_cells([section["total"]], font_size=9, left_columns={0})[0])
        elements.append(_make_pdf_table(data, col_widths=col_widths, header_font_size=9, body_font_size=9, cell_padding=5, first_col_left=True, left_cols=[0]))

    if overall_total:
        elements.append(Spacer(1, 5 * mm))
        data = [wrap_pdf_header_cells(pdf_headers, font_size=11), wrap_pdf_body_cells([overall_total], font_size=9, left_columns={0})[0]]
        elements.append(_make_pdf_table(data, col_widths=col_widths, header_font_size=9, body_font_size=9, cell_padding=5, first_col_left=True, left_cols=[0]))

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


def generate_zone_grouped_pdf(title, summary_lines, grouped_sections, overall_total=None, zone_summary_data=None, pdf_detail_headers=None, zone_summary_sel=None):
    buf = io.BytesIO()
    left_margin = 15 * mm
    right_margin = 15 * mm
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        topMargin=20 * mm,
        bottomMargin=15 * mm,
        leftMargin=left_margin,
        rightMargin=right_margin,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("PDFTitle", parent=styles["Heading1"], fontSize=20, textColor=ACCENT, alignment=1, spaceAfter=6 * mm, fontName="Helvetica-Bold")
    summary_style = ParagraphStyle("PDFSummary", parent=styles["Normal"], fontSize=11, alignment=0, spaceAfter=2 * mm, textColor=colors.HexColor("#333333"), leading=16)
    section_style = ParagraphStyle("PDFSection", parent=styles["Heading3"], fontSize=14, textColor=ACCENT, alignment=1, spaceBefore=5 * mm, spaceAfter=3 * mm, fontName="Helvetica-Bold")
    card_header_style = ParagraphStyle("CardHeader", parent=styles["Normal"], fontSize=10, textColor=HEADER_FG, fontName="Helvetica-Bold", alignment=1)
    card_label_style = ParagraphStyle("CardLabel", parent=styles["Normal"], fontSize=10, alignment=1, fontName="Helvetica")
    card_value_style = ParagraphStyle("CardValue", parent=styles["Normal"], fontSize=10, alignment=1, fontName="Helvetica")
    headers = pdf_detail_headers or ["Sr", "Sector", "Total Bills", "Received Bills", "Remaining Bills", "Amount Received", "Pending Amount"]
    page_w = landscape(A4)[0] - left_margin - right_margin
    col_widths = [page_w * 0.05, page_w * 0.40, page_w * 0.09, page_w * 0.11, page_w * 0.11, page_w * 0.12, page_w * 0.12]

    elements = [Paragraph(title, title_style)]
    for line in summary_lines:
        elements.append(Paragraph(line, summary_style))

    if zone_summary_data:
        elements.append(Spacer(1, 5 * mm))

        _zs_all_h = ["Sr", "Zone", "Total Bills", "Received Bills", "Total Amount", "Arrears Received", "Total Received", "Pending Amount"]
        _zs_all_w = [page_w * 0.04, page_w * 0.10, page_w * 0.10, page_w * 0.10, page_w * 0.15, page_w * 0.16, page_w * 0.15, page_w * 0.20]
        if zone_summary_sel:
            summ_headers = [_zs_all_h[i] for i in zone_summary_sel]
            summ_col_widths = [_zs_all_w[i] for i in zone_summary_sel]
        else:
            summ_headers = _zs_all_h
            summ_col_widths = _zs_all_w

        summ_header_style = ParagraphStyle("SummHeader", fontSize=10, leading=13, alignment=1, textColor=HEADER_FG, fontName="Helvetica-Bold")
        summ_data = [[Paragraph(h, summ_header_style) for h in summ_headers]]
        gt_bills = gt_received = gt_amount = gt_arrears = gt_received_amt = gt_pending = 0
        for idx, row in enumerate(zone_summary_data, start=1):
            tb = int(row['total_bills'] or 0)
            rb = int(row['received_bills'] or 0)
            ta = int(row['total_amount'] or 0)
            ar = int(row['total_arrears_received'] or 0)
            rv = int(row['total_received_amount'] or 0)
            pd_amt = int(row['pending_amount'] or 0)
            gt_bills += tb
            gt_received += rb
            gt_amount += ta
            gt_arrears += ar
            gt_received_amt += rv
            gt_pending += pd_amt
            _full_zs_row = [f"{idx}", row["zone"], f"{tb:,}", f"{rb:,}", f"{ta:,}", f"{ar:,}", f"{rv:,}", f"{pd_amt:,}"]
            summ_data.append([_full_zs_row[i] for i in (zone_summary_sel or range(8))])
        grand_idx = len(summ_data)
        _full_zs_gt = ["", "Grand Total", f"{gt_bills:,}", f"{gt_received:,}", f"{gt_amount:,}", f"{gt_arrears:,}", f"{gt_received_amt:,}", f"{gt_pending:,}"]
        summ_data.append([_full_zs_gt[i] for i in (zone_summary_sel or range(8))])

        summary_table = Table(summ_data, colWidths=summ_col_widths, repeatRows=1)
        st = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
            ("TEXTCOLOR", (0, 0), (-1, 0), HEADER_FG),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_CLR),
            ("TOPPADDING", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("TOPPADDING", (0, 1), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 12),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
        ])
        for i in range(2, grand_idx):
            if i % 2 == 0:
                st.add("BACKGROUND", (0, i), (-1, i), ALT_ROW)
        st.add("BACKGROUND", (0, grand_idx), (-1, grand_idx), colors.HexColor("#e6d8c8"))
        st.add("FONTNAME", (0, grand_idx), (-1, grand_idx), "Helvetica-Bold")
        summary_table.setStyle(st)
        elements.append(summary_table)
        elements.append(Spacer(1, 4 * mm))

    for section in grouped_sections:
        elements.append(PageBreak())

        zone_name = section["zone"]
        total_row = section.get("total")

        summary_parts = []
        summary_parts.append(Spacer(1, 6 * mm))
        summary_parts.append(Paragraph(f"{zone_name} Summary", section_style))
        summary_parts.append(Spacer(1, 4 * mm))

        if total_row:
            card_data = [
                [Paragraph("Metric", card_header_style), Paragraph("Value", card_header_style)],
                [Paragraph("Total Bills", card_label_style), Paragraph(total_row[2], card_value_style)],
                [Paragraph("Received Bills", card_label_style), Paragraph(total_row[3], card_value_style)],
                [Paragraph("Remaining Bills", card_label_style), Paragraph(total_row[4], card_value_style)],
                [Paragraph("Amount Received", card_label_style), Paragraph(f"Rs. {total_row[5]}", card_value_style)],
                [Paragraph("Pending Amount", card_label_style), Paragraph(f"Rs. {total_row[6]}", card_value_style)],
            ]
            summary_cw = [page_w * 0.32, page_w * 0.32]
            card_table = Table(card_data, colWidths=summary_cw, hAlign="CENTER")
            card_style = TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
                ("TEXTCOLOR", (0, 0), (-1, 0), HEADER_FG),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, BORDER_CLR),
                ("TOPPADDING", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("TOPPADDING", (0, 1), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 12),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ])
            for i in range(2, len(card_data)):
                if i % 2 == 0:
                    card_style.add("BACKGROUND", (0, i), (-1, i), ALT_ROW)
            card_table.setStyle(card_style)
            summary_parts.append(card_table)
            summary_parts.append(Spacer(1, 3 * mm))

        elements.append(KeepTogether(summary_parts))
        elements.append(Spacer(1, 6 * mm))

        detail_header_style = ParagraphStyle("DetailHeader", fontSize=10, leading=13, alignment=1, textColor=HEADER_FG, fontName="Helvetica-Bold")
        detail_rows = [[Paragraph(h, detail_header_style) for h in headers]]
        has_dt = section.get("total")
        for row in section["rows"]:
            detail_rows.append(row)
        if has_dt:
            detail_rows.append(has_dt)
        detail_table = Table(detail_rows, colWidths=col_widths, repeatRows=1)
        dt_st = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
            ("TEXTCOLOR", (0, 0), (-1, 0), HEADER_FG),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (1, 1), (1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_CLR),
            ("TOPPADDING", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("TOPPADDING", (0, 1), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 12),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
        ])
        for i in range(2, len(detail_rows)):
            if i % 2 == 0:
                dt_st.add("BACKGROUND", (0, i), (-1, i), ALT_ROW)
        if has_dt:
            last = len(detail_rows) - 1
            dt_st.add("BACKGROUND", (0, last), (-1, last), colors.HexColor("#e6d8c8"))
            dt_st.add("FONTNAME", (0, last), (-1, last), "Helvetica-Bold")
        detail_table.setStyle(dt_st)
        elements.append(detail_table)
        elements.append(Spacer(1, 6 * mm))

    if overall_total:
        elements.append(Spacer(1, 5 * mm))
        gt_header_style = ParagraphStyle("GtHeader", fontSize=10, leading=13, alignment=1, textColor=HEADER_FG, fontName="Helvetica-Bold")
        gt_rows = [[Paragraph(h, gt_header_style) for h in headers], overall_total]
        gt_table = Table(gt_rows, colWidths=col_widths, repeatRows=1)
        gt_st = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
            ("TEXTCOLOR", (0, 0), (-1, 0), HEADER_FG),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_CLR),
            ("TOPPADDING", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("TOPPADDING", (0, 1), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 12),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#e6d8c8")),
        ])
        gt_table.setStyle(gt_st)
        elements.append(gt_table)

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Bill List database and reports
# ---------------------------------------------------------------------------

def get_db():
    conn = sqlite3.connect(BILL_LIST_DB)
    conn.row_factory = sqlite3.Row
    return conn


def init_bill_list_db() -> None:
    with get_db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS localities (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sector TEXT NOT NULL,
                locality TEXT NOT NULL,
                zone TEXT NOT NULL,
                UNIQUE(sector, locality)
            );

            CREATE TABLE IF NOT EXISTS zones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE
            );

            CREATE TABLE IF NOT EXISTS sectors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                zone TEXT NOT NULL,
                UNIQUE(name, zone)
            );

            CREATE TABLE IF NOT EXISTS bills (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bill_key TEXT NOT NULL UNIQUE,
                sector TEXT NOT NULL,
                locality TEXT NOT NULL,
                zone TEXT NOT NULL,
                bill_no TEXT,
                reference_no TEXT,
                connection_no TEXT,
                consumer_name TEXT,
                total_bill REAL NOT NULL DEFAULT 0,
                arrears REAL NOT NULL DEFAULT 0,
                amount_received REAL NOT NULL DEFAULT 0,
                status TEXT,
                raw_data TEXT NOT NULL,
                uploaded_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS staff (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS staff_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id INTEGER NOT NULL,
                zone TEXT NOT NULL,
                sector TEXT,
                locality TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(staff_id) REFERENCES staff(id),
                UNIQUE(staff_id, zone, sector, locality)
            );

            CREATE TABLE IF NOT EXISTS auto_assignment_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sector TEXT NOT NULL,
                locality TEXT NOT NULL,
                connection_min TEXT NOT NULL,
                connection_max TEXT,
                staff_name TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(sector, locality, connection_min, connection_max, staff_name)
            );
            """
        )
        columns = [row["name"] for row in conn.execute("PRAGMA table_info(staff_assignments)").fetchall()]
        if "locality" not in columns:
            conn.execute("ALTER TABLE staff_assignments ADD COLUMN locality TEXT")
        bill_columns = [row["name"] for row in conn.execute("PRAGMA table_info(bills)").fetchall()]
        if "arrears" not in bill_columns:
            conn.execute("ALTER TABLE bills ADD COLUMN arrears REAL NOT NULL DEFAULT 0")
            backfill_bill_arrears(conn)
        if "consumer_mobile" not in bill_columns:
            conn.execute("ALTER TABLE bills ADD COLUMN consumer_mobile TEXT")
        if "consumer_name" not in bill_columns:
            conn.execute("ALTER TABLE bills ADD COLUMN consumer_name TEXT")
        unique_indexes = conn.execute("PRAGMA index_list(staff_assignments)").fetchall()
        has_locality_unique = False
        for index in unique_indexes:
            if not index["unique"]:
                continue
            index_cols = [
                col["name"]
                for col in conn.execute(f"PRAGMA index_info({index['name']})").fetchall()
            ]
            if index_cols == ["staff_id", "zone", "sector", "locality"]:
                has_locality_unique = True
                break
        if not has_locality_unique:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS staff_assignments_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    staff_id INTEGER NOT NULL,
                    zone TEXT NOT NULL,
                    sector TEXT,
                    locality TEXT,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(staff_id) REFERENCES staff(id),
                    UNIQUE(staff_id, zone, sector, locality)
                );
                INSERT OR IGNORE INTO staff_assignments_new (id, staff_id, zone, sector, locality, created_at)
                SELECT id, staff_id, zone, sector, locality, created_at FROM staff_assignments;
                DROP TABLE staff_assignments;
                ALTER TABLE staff_assignments_new RENAME TO staff_assignments;
                """
            )
        for zone in ["A", "B", "C", "Commercial", "Unassigned"]:
            conn.execute("INSERT OR IGNORE INTO zones (name) VALUES (?)", (zone,))
        conn.execute(
            """
            INSERT OR IGNORE INTO sectors (name, zone)
            SELECT DISTINCT sector, zone
            FROM localities
            WHERE sector IS NOT NULL AND TRIM(sector) != ''
            """
        )
        conn.execute(
            """
            INSERT OR IGNORE INTO sectors (name, zone)
            SELECT DISTINCT sector, zone
            FROM bills
            WHERE sector IS NOT NULL AND TRIM(sector) != ''
            """
        )
        apply_manual_zone_overrides(conn)
        seed_auto_assignment_rules(conn)
        seed_staff_assignments_from_file(conn)


def seed_auto_assignment_rules(conn) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    conn.execute("DROP INDEX IF EXISTS idx_auto_unique")
    conn.execute("DELETE FROM auto_assignment_rules")
    for rule in AUTO_ASSIGNMENT_RULES:
        conn.execute(
            """
            INSERT OR IGNORE INTO auto_assignment_rules (sector, locality, connection_min, connection_max, staff_name, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (rule["sector"], rule["locality"], rule["connection_min"], rule["connection_max"], rule["staff_name"], now),
        )
    conn.execute(
        """CREATE UNIQUE INDEX IF NOT EXISTS idx_auto_unique
        ON auto_assignment_rules(sector, locality, connection_min, connection_max, staff_name)"""
    )


_ALIAS_RULES: list[dict] = []
_ALIAS_RULES_LOADED = False


def load_alias_rules(force_reload: bool = False) -> list[dict]:
    global _ALIAS_RULES, _ALIAS_RULES_LOADED
    if _ALIAS_RULES_LOADED and not force_reload:
        return list(_ALIAS_RULES)
    rules: list[dict] = []
    if os.path.exists(SEED_ASSIGNMENTS_JSON):
        try:
            with open(SEED_ASSIGNMENTS_JSON, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            for item in data.get("aliases", []) or []:
                match = str(item.get("match") or "").strip()
                if not match:
                    continue
                rules.append({
                    "match": match,
                    "match_key": match_key(match),
                    "staff_name": str(item.get("staff_name") or "").strip(),
                    "zone": str(item.get("zone") or "").strip() or "Unassigned",
                })
        except (OSError, json.JSONDecodeError, TypeError, ValueError):
            rules = []
    _ALIAS_RULES = rules
    _ALIAS_RULES_LOADED = True
    return list(_ALIAS_RULES)


def match_by_alias(sector: str, locality: str) -> dict | None:
    rules = load_alias_rules()
    if not rules:
        return None
    sector_k = match_key(sector)
    locality_k = match_key(locality) if locality else ""
    haystack = f"{sector_k} {locality_k}".strip()
    if not haystack:
        return None
    for rule in rules:
        needle = rule["match_key"]
        if not needle:
            continue
        if needle in haystack:
            with get_db() as conn:
                row = conn.execute(
                    "SELECT id, name FROM staff WHERE name = ?",
                    (rule["staff_name"],),
                ).fetchone()
            if row:
                return {
                    "staff_id": int(row["id"]),
                    "staff_name": row["name"],
                    "zone": rule["zone"],
                    "sector": sector,
                    "locality": locality,
                }
            return {
                "staff_id": 0,
                "staff_name": rule["staff_name"],
                "zone": rule["zone"],
                "sector": sector,
                "locality": locality,
            }
    return None


def seed_staff_assignments_from_file(conn) -> None:
    if os.environ.get("VERCEL") != "1":
        load_alias_rules(force_reload=True)
        return
    existing_assignments = conn.execute("SELECT COUNT(*) AS cnt FROM staff_assignments").fetchone()
    existing_sectors = conn.execute("SELECT COUNT(*) AS cnt FROM sectors").fetchone()
    existing_localities = conn.execute("SELECT COUNT(*) AS cnt FROM localities").fetchone()
    seed_full = (not existing_sectors["cnt"]) or (not existing_localities["cnt"])

    seed_staff: list[str] = []
    seed_assignments: list[dict] = []
    seed_sectors: list[dict] = []
    seed_localities: list[dict] = []

    if os.path.exists(SEED_ASSIGNMENTS_JSON):
        try:
            with open(SEED_ASSIGNMENTS_JSON, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            seed_staff = [str(s).strip() for s in data.get("staff", []) if str(s).strip()]
            seed_assignments = data.get("assignments", []) or []
            seed_sectors = data.get("sectors", []) or []
            seed_localities = data.get("localities", []) or []
        except (OSError, json.JSONDecodeError, TypeError, ValueError):
            return
    elif os.path.exists(SEED_ASSIGNMENTS_CSV):
        try:
            with open(SEED_ASSIGNMENTS_CSV, "r", encoding="utf-8", newline="") as handle:
                reader = csv.DictReader(handle)
                for row in reader:
                    seed_assignments.append(row)
                    staff_name = (row.get("staff_name") or "").strip()
                    if staff_name:
                        seed_staff.append(staff_name)
        except OSError:
            return
    else:
        return

    now = datetime.now().isoformat(timespec="seconds")

    if seed_full:
        seen_zones: set[str] = set()
        for item in seed_sectors:
            zone = str(item.get("zone") or "").strip() or "Unassigned"
            if zone and zone not in seen_zones:
                conn.execute("INSERT OR IGNORE INTO zones (name) VALUES (?)", (zone,))
                seen_zones.add(zone)
            name = str(item.get("name") or "").strip()
            if name:
                conn.execute(
                    "INSERT OR IGNORE INTO sectors (name, zone) VALUES (?, ?)",
                    (name, zone),
                )
        for item in seed_localities:
            sector_name = str(item.get("sector") or "").strip()
            locality_name = str(item.get("locality") or "").strip()
            zone = str(item.get("zone") or "").strip() or "Unassigned"
            if not sector_name or not locality_name:
                continue
            if zone and zone not in seen_zones:
                conn.execute("INSERT OR IGNORE INTO zones (name) VALUES (?)", (zone,))
                seen_zones.add(zone)
            conn.execute(
                "INSERT OR IGNORE INTO localities (sector, locality, zone) VALUES (?, ?, ?)",
                (sector_name, locality_name, zone),
            )
            conn.execute(
                "INSERT OR IGNORE INTO sectors (name, zone) VALUES (?, ?)",
                (sector_name, zone),
            )

    if existing_assignments["cnt"] > 0:
        load_alias_rules(force_reload=True)
        return

    staff_names = sorted({name for name in seed_staff if name})
    for name in staff_names:
        conn.execute(
            "INSERT OR IGNORE INTO staff (name, created_at) VALUES (?, ?)",
            (name, now),
        )

    staff_map = {row["name"]: row["id"] for row in conn.execute("SELECT id, name FROM staff").fetchall()}

    for item in seed_assignments:
        staff_name = str(item.get("staff_name") or "").strip()
        if not staff_name:
            continue
        staff_id = staff_map.get(staff_name)
        if staff_id is None:
            conn.execute(
                "INSERT OR IGNORE INTO staff (name, created_at) VALUES (?, ?)",
                (staff_name, now),
            )
            row = conn.execute("SELECT id FROM staff WHERE name = ?", (staff_name,)).fetchone()
            if not row:
                continue
            staff_id = row["id"]
            staff_map[staff_name] = staff_id

        zone = str(item.get("zone") or "").strip() or "Unassigned"
        sector = str(item.get("sector") or "").strip() or None
        locality = str(item.get("locality") or "").strip() or None
        conn.execute(
            """
            INSERT OR IGNORE INTO staff_assignments (staff_id, zone, sector, locality, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (staff_id, zone, sector, locality, now),
        )

    load_alias_rules(force_reload=True)


def backfill_bill_arrears(conn) -> None:
    rows = conn.execute("SELECT id, raw_data FROM bills").fetchall()
    for row in rows:
        try:
            raw = json.loads(row["raw_data"] or "{}")
        except (TypeError, json.JSONDecodeError):
            raw = {}
        conn.execute(
            "UPDATE bills SET arrears = ? WHERE id = ?",
            (parse_number(raw.get("arrears")), row["id"]),
        )


def apply_manual_zone_overrides(conn) -> None:
    overrides = [
        (
            "Bahu Chowk to Sabz mandi Road",
            "67-Bahu Chowk, Sabzi  Mandi Road,Faisal Colony Zon",
            "A",
        )
    ]
    for sector, locality_prefix, zone in overrides:
        conn.execute(
            """
            UPDATE localities
            SET zone = ?
            WHERE sector = ? AND locality LIKE ? AND zone = 'Unassigned'
            """,
            (zone, sector, f"{locality_prefix}%"),
        )
        conn.execute(
            """
            UPDATE bills
            SET zone = ?
            WHERE sector = ? AND locality LIKE ? AND zone = 'Unassigned'
            """,
            (zone, sector, f"{locality_prefix}%"),
        )
        conn.execute("INSERT OR IGNORE INTO zones (name) VALUES (?)", (zone,))
        conn.execute(
            """
            INSERT OR IGNORE INTO sectors (name, zone)
            SELECT DISTINCT sector, zone
            FROM localities
            WHERE sector = ?
            """,
            (sector,),
        )
    conn.execute(
        """
        DELETE FROM sectors
        WHERE NOT EXISTS (
            SELECT 1 FROM localities WHERE localities.sector = sectors.name AND localities.zone = sectors.zone
        )
        AND NOT EXISTS (
            SELECT 1 FROM bills WHERE bills.sector = sectors.name AND bills.zone = sectors.zone
        )
        """
    )


def pick_saved_zone(conn, sector: str, locality: str, inferred_zone: str) -> str:
    row = conn.execute(
        "SELECT zone FROM localities WHERE sector = ? AND locality = ?",
        (sector, locality),
    ).fetchone()
    if row:
        return row["zone"]

    row = conn.execute(
        """
        SELECT zone
        FROM sectors
        WHERE name = ?
        ORDER BY CASE zone WHEN 'A' THEN 1 WHEN 'B' THEN 2 WHEN 'C' THEN 3 WHEN 'Commercial' THEN 4 ELSE 5 END
        LIMIT 1
        """,
        (sector,),
    ).fetchone()
    if row:
        return row["zone"]

    return inferred_zone


def update_sector_zone(conn, sector: str, zone: str) -> None:
    old_zones = [
        row["zone"]
        for row in conn.execute(
            "SELECT DISTINCT zone FROM sectors WHERE name = ?",
            (sector,),
        ).fetchall()
    ]
    conn.execute("INSERT OR IGNORE INTO zones (name) VALUES (?)", (zone,))
    conn.execute("UPDATE localities SET zone = ? WHERE sector = ?", (zone, sector))
    conn.execute("UPDATE bills SET zone = ? WHERE sector = ?", (zone, sector))
    conn.execute("DELETE FROM sectors WHERE name = ?", (sector,))
    conn.execute("INSERT OR IGNORE INTO sectors (name, zone) VALUES (?, ?)", (sector, zone))
    for old_zone in old_zones:
        conn.execute(
            """
            UPDATE staff_assignments
            SET zone = ?
            WHERE sector = ? AND zone = ?
            """,
            (zone, sector, old_zone),
        )


def clear_bill_list_data() -> None:
    init_bill_list_db()
    with get_db() as conn:
        conn.execute("DELETE FROM bills")
        conn.execute(
            """
            DELETE FROM sqlite_sequence
            WHERE name = 'bills'
            """
        )

    upload_root = os.path.abspath(app.config["UPLOAD_FOLDER"])
    for filename in os.listdir(upload_root):
        if not filename.startswith("bill_list_"):
            continue
        path = os.path.abspath(os.path.join(upload_root, filename))
        if path.startswith(upload_root + os.sep) and os.path.isfile(path):
            try:
                os.remove(path)
            except OSError:
                pass


def get_assignment_conflicts(conn, staff_id: str, zone: str, sectors: list[str], localities: list[tuple[str, str]] | None = None) -> list[str]:
    localities = localities or []
    if localities:
        conflicts = []
        for sector, locality in localities:
            row = conn.execute(
                """
                SELECT s.name
                FROM staff_assignments sa
                JOIN staff s ON s.id = sa.staff_id
                WHERE sa.zone = ?
                  AND (
                    sa.sector IS NULL
                    OR (sa.sector = ? AND sa.locality IS NULL)
                    OR (sa.sector = ? AND sa.locality = ?)
                  )
                  AND sa.staff_id != ?
                LIMIT 1
                """,
                (zone, sector, sector, locality, staff_id),
            ).fetchone()
            if row:
                conflicts.append(f"{sector} / {locality} is already assigned to {row['name']}")
        return conflicts

    if sectors:
        conflicts = []
        for sector in sectors:
            row = conn.execute(
                """
                SELECT s.name
                FROM staff_assignments sa
                JOIN staff s ON s.id = sa.staff_id
                WHERE sa.zone = ?
                  AND (sa.sector = ? OR sa.sector IS NULL)
                  AND sa.staff_id != ?
                LIMIT 1
                """,
                (zone, sector, staff_id),
            ).fetchone()
            if row:
                conflicts.append(f"{sector} is already assigned to {row['name']}")
        return conflicts

    rows = conn.execute(
        """
        SELECT DISTINCT COALESCE(sa.sector, 'All sectors') AS sector, s.name
        FROM staff_assignments sa
        JOIN staff s ON s.id = sa.staff_id
        WHERE sa.zone = ? AND sa.staff_id != ?
        ORDER BY sector
        """,
        (zone, staff_id),
    ).fetchall()
    return [f"{row['sector']} is already assigned to {row['name']}" for row in rows]


def infer_zone(sector: str, locality: str, row: dict) -> str:
    text = " ".join(str(value) for value in [sector, locality, row.get("connection type", ""), row.get("bill type", "")]).lower()
    if "commercial" in text:
        return "Commercial"
    match = re.search(r"\bzone\s*[-:]?\s*([abc])\b", text)
    if match:
        return match.group(1).upper()
    match = re.search(r"\b([abc])\s*zone\b", text)
    if match:
        return match.group(1).upper()
    return "Unassigned"


def build_bill_key(row: dict) -> str:
    key_sets = [
        ["bill no"],
        ["reference no", "due date", "total bill"],
        ["reference no", "amount received", "total bill"],
        ["connection no", "due date", "total bill"],
        ["connection no", "amount received", "total bill"],
    ]
    for columns in key_sets:
        values = [_dedupe_value(row.get(col)) for col in columns]
        if all(values):
            return "|".join(values)
    return "|".join(_dedupe_value(row.get(col)) for col in sorted(row.keys()))


def fast_bill_no_key(value) -> str:
    """Fast duplicate key for Bill Reports uploads when Bill No is present."""
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    compact = text.replace(",", "")
    if re.fullmatch(r"-?\d+(?:\.0+)?", compact):
        return str(int(float(compact)))
    return " ".join(text.lower().split())


def fast_upload_number(value) -> float:
    """Fast numeric parser for Bill Reports upload amount columns."""
    if value is None or pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def fast_upload_text(value) -> str | None:
    """Fast text normalizer for Bill Reports upload identifier/name fields."""
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return str(int(value)) if float(value).is_integer() else f"{float(value):.6f}".rstrip("0").rstrip(".")
    text = str(value).strip()
    if not text:
        return None
    compact = text.replace(",", "")
    if re.fullmatch(r"-?\d+(?:\.\d+)?", compact):
        number = float(compact)
        return str(int(number)) if number.is_integer() else f"{number:.6f}".rstrip("0").rstrip(".")
    return " ".join(text.lower().split())


def sql_text(value) -> str | None:
    text = _dedupe_value(value)
    return text or None


def import_bill_list_dataframe(df: pd.DataFrame) -> tuple[int, int]:
    df = normalize_dataframe(df)
    required = {"sector", "locality"}
    missing = sorted(required - set(df.columns))
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")

    df, duplicate_rows_removed, _ = drop_duplicate_bills(df)
    now = datetime.now().isoformat(timespec="seconds")
    inserted_or_updated = 0

    with get_db() as conn:
        # Bill Reports upload can receive thousands of rows on Vercel.  Cache
        # saved zone lookups once and bulk-write rows so the upload finishes
        # before the serverless request times out.  This keeps the same fields
        # and duplicate rules; it only removes per-row database chatter.
        locality_zone_cache = {
            (row["sector"], row["locality"]): row["zone"]
            for row in conn.execute("SELECT sector, locality, zone FROM localities").fetchall()
        }
        sector_zone_cache = {}
        for row in conn.execute(
            """
            SELECT name, zone
            FROM sectors
            ORDER BY CASE zone WHEN 'A' THEN 1 WHEN 'B' THEN 2 WHEN 'C' THEN 3 WHEN 'Commercial' THEN 4 ELSE 5 END
            """
        ).fetchall():
            sector_zone_cache.setdefault(row["name"], row["zone"])

        zone_rows: set[tuple[str]] = set()
        sector_rows: set[tuple[str, str]] = set()
        locality_rows: set[tuple[str, str, str]] = set()
        bill_rows: list[tuple] = []

        for raw_row in df.to_dict(orient="records"):
            row = {str(key): (None if pd.isna(value) else value) for key, value in raw_row.items()}
            sector = str(row.get("sector") or "Unknown").strip() or "Unknown"
            locality = str(row.get("locality") or "Unknown").strip() or "Unknown"
            zone = (
                locality_zone_cache.get((sector, locality))
                or sector_zone_cache.get(sector)
                or infer_zone(sector, locality, row)
            )
            locality_zone_cache[(sector, locality)] = zone
            sector_zone_cache.setdefault(sector, zone)
            # Bill No is the first duplicate-key rule.  Use a small fast path
            # here because Vercel uploads large bill files and the generic
            # normalizer is intentionally more expensive for mixed date fields.
            bill_key = fast_bill_no_key(row.get("bill no")) or build_bill_key(row)
            total_bill = fast_upload_number(row.get("total bill") or row.get("after due date"))
            consumer_name_value = row.get("consumer name / f/h name")
            if consumer_name_value is None:
                consumer_name_value = row.get("consumer name")
            if consumer_name_value is None:
                consumer_name_value = row.get("f/h name")
            arrears_value = row.get("arrears")
            if arrears_value is None:
                arrears_value = row.get("outstanding amount")
            if arrears_value is None:
                arrears_value = row.get("outstanding")
            arrears = fast_upload_number(arrears_value)
            amount_received = fast_upload_number(row.get("amount received"))
            mobile_value = row.get("consumer mobile")
            if mobile_value is None:
                mobile_value = row.get("mobile no")
            if mobile_value is None:
                mobile_value = row.get("mobile")
            if mobile_value is None:
                mobile_value = row.get("consumer phone")
            if mobile_value is None:
                mobile_value = row.get("phone")
            consumer_mobile = str(mobile_value).strip() if mobile_value is not None else ""

            zone_rows.add((zone,))
            locality_rows.add((sector, locality, zone))
            sector_rows.add((sector, zone))
            bill_rows.append(
                (
                    bill_key,
                    sector,
                    locality,
                    zone,
                    fast_upload_text(row.get("bill no")),
                    fast_upload_text(row.get("reference no")),
                    fast_upload_text(row.get("connection no")),
                    fast_upload_text(consumer_name_value),
                    total_bill,
                    arrears,
                    amount_received,
                    row.get("status"),
                    consumer_mobile or None,
                    json.dumps(row, ensure_ascii=True, default=str),
                    now,
                )
            )
            inserted_or_updated += 1

        conn.executemany("INSERT OR IGNORE INTO zones (name) VALUES (?)", sorted(zone_rows))
        conn.executemany(
            """
            INSERT INTO localities (sector, locality, zone)
            VALUES (?, ?, ?)
            ON CONFLICT(sector, locality) DO NOTHING
            """,
            sorted(locality_rows),
        )
        conn.executemany(
            """
            INSERT INTO sectors (name, zone)
            VALUES (?, ?)
            ON CONFLICT(name, zone) DO NOTHING
            """,
            sorted(sector_rows),
        )
        conn.executemany(
            """
            INSERT INTO bills (
                bill_key, sector, locality, zone, bill_no, reference_no, connection_no, consumer_name,
                total_bill, arrears, amount_received, status, consumer_mobile, raw_data, uploaded_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(bill_key) DO UPDATE SET
                sector = excluded.sector,
                locality = excluded.locality,
                zone = excluded.zone,
                bill_no = excluded.bill_no,
                reference_no = excluded.reference_no,
                connection_no = excluded.connection_no,
                consumer_name = excluded.consumer_name,
                total_bill = excluded.total_bill,
                arrears = excluded.arrears,
                amount_received = excluded.amount_received,
                status = excluded.status,
                consumer_mobile = excluded.consumer_mobile,
                raw_data = excluded.raw_data,
                uploaded_at = excluded.uploaded_at
            """,
            bill_rows,
        )

    return inserted_or_updated, duplicate_rows_removed


UNPAID_STATUS_SQL = "LOWER(REPLACE(REPLACE(TRIM(COALESCE(status, '')), '-', ''), ' ', '')) IN ('unpaid', 'expired')"


# ---------------------------------------------------------------------------
# Advanced Bill Filtering and Export
# ---------------------------------------------------------------------------

def get_filtered_bills(
    outstanding_amount: float | None = None,
    outstanding_operator: str = "gt",
    bill_status: str = "",
    sector: str | None = None,
    zone: str | None = None,
    staff_id: int | None = None,
):
    init_bill_list_db()
    with get_db() as conn:
        base_query = """
            SELECT
                b.id,
                b.bill_no,
                b.reference_no,
                b.connection_no,
                b.consumer_name,
                b.sector,
                b.locality,
                b.zone,
                b.total_bill,
                b.arrears,
                b.amount_received,
                b.status,
                b.consumer_mobile,
                b.raw_data,
                (b.total_bill - COALESCE(b.amount_received, 0)) AS outstanding_amount
            FROM bills b
            WHERE 1=1
        """
        params = []

        if bill_status == "paid":
            base_query += " AND b.amount_received >= b.total_bill"
        elif bill_status == "unpaid":
            base_query += " AND b.amount_received < b.total_bill"

        if outstanding_amount is not None and outstanding_amount > 0:
            if outstanding_operator == "lt":
                base_query += " AND (b.total_bill - COALESCE(b.amount_received, 0)) < ?"
            else:
                base_query += " AND (b.total_bill - COALESCE(b.amount_received, 0)) > ?"
            params.append(outstanding_amount)

        if sector:
            base_query += " AND b.sector = ?"
            params.append(sector)

        if zone:
            base_query += " AND b.zone = ?"
            params.append(zone)

        if staff_id is not None:
            base_query += """
                AND (
                    EXISTS (
                        SELECT 1 FROM staff_assignments sa
                        WHERE sa.staff_id = ?
                        AND sa.zone = b.zone
                        AND (sa.sector IS NULL OR b.sector = sa.sector)
                        AND (sa.locality IS NULL OR b.locality = sa.locality)
                    )
                    OR EXISTS (
                        SELECT 1 FROM auto_assignment_rules aar
                        JOIN staff s ON UPPER(TRIM(s.name)) = UPPER(TRIM(aar.staff_name))
                        WHERE s.id = ?
                        AND aar.sector = b.sector
                        AND aar.locality = b.locality
                        AND CAST(b.connection_no AS TEXT) >= aar.connection_min
                        AND (aar.connection_max IS NULL OR CAST(b.connection_no AS TEXT) <= aar.connection_max)
                    )
                )
            """
            params.append(staff_id)
            params.append(staff_id)

        base_query += " ORDER BY b.zone, b.sector, b.locality, b.bill_no"

        rows = conn.execute(base_query, params).fetchall()

    bills = []
    for row in rows:
        consumer_mobile = row["consumer_mobile"] or ""
        consumer_name = row["consumer_name"] or ""
        if not consumer_mobile:
            try:
                raw = json.loads(row["raw_data"] or "{}")
                if not consumer_name:
                    for nk in ("consumer name / f/h name", "consumer name", "f/h name"):
                        nv = raw.get(nk)
                        if nv is not None and str(nv).strip():
                            consumer_name = str(nv).strip()
                            break
                for mk in ("consumer mobile", "mobile no", "mobile", "consumer phone", "phone"):
                    mv = raw.get(mk)
                    if mv is not None and str(mv).strip():
                        consumer_mobile = str(mv).strip()
                        break
            except (json.JSONDecodeError, TypeError, ValueError):
                pass
        consumer_mobile = format_mobile(consumer_mobile)
        bills.append({
            "bill_no": row["bill_no"] or "",
            "reference_no": row["reference_no"] or "",
            "connection_no": row["connection_no"] or "",
            "consumer_name": consumer_name,
            "sector": row["sector"],
            "locality": row["locality"],
            "zone": row["zone"],
            "total_bill": float(row["total_bill"] or 0),
            "arrears": float(row["arrears"] or 0),
            "amount_received": float(row["amount_received"] or 0),
            "outstanding_amount": float(row["outstanding_amount"] or 0),
            "status": row["status"] or "",
            "consumer_mobile": consumer_mobile,
        })
    return bills


# ---------------------------------------------------------------------------
# Advanced Bill Checking — Grouping helpers
# ---------------------------------------------------------------------------

def map_bills_to_staff(bills: list[dict]) -> list[dict]:
    """Add a 'staff_name' key to each bill based on staff_assignments (locality > sector > zone) then auto_assignment_rules."""
    if not bills:
        return bills
    init_bill_list_db()
    with get_db() as conn:
        assignments = conn.execute("""
            SELECT sa.zone, sa.sector, sa.locality, s.name
            FROM staff_assignments sa
            JOIN staff s ON s.id = sa.staff_id
        """).fetchall()
        auto_rules = conn.execute("""
            SELECT aar.sector, aar.locality, aar.connection_min, aar.connection_max, s.name
            FROM auto_assignment_rules aar
            JOIN staff s ON UPPER(TRIM(s.name)) = UPPER(TRIM(aar.staff_name))
        """).fetchall()

    loc_map = {}
    sec_map = {}
    zone_map = {}
    for a in assignments:
        z, sec, loc, name_ = a["zone"], a["sector"], a["locality"], a["name"]
        if loc:
            loc_map[(z, sec, loc)] = name_
        elif sec:
            sec_map[(z, sec)] = name_
        else:
            zone_map[z] = name_

    for bill in bills:
        k_loc = (bill["zone"], bill["sector"], bill["locality"])
        k_sec = (bill["zone"], bill["sector"])
        k_zone = bill["zone"]
        if k_loc in loc_map:
            bill["staff_name"] = loc_map[k_loc]
        elif k_sec in sec_map:
            bill["staff_name"] = sec_map[k_sec]
        elif k_zone in zone_map:
            bill["staff_name"] = zone_map[k_zone]
        else:
            assigned = "Unassigned"
            for r in auto_rules:
                if (r["sector"] == bill["sector"] and r["locality"] == bill["locality"]
                        and str(bill.get("connection_no") or "") >= str(r["connection_min"])
                        and (r["connection_max"] is None or str(bill.get("connection_no") or "") <= str(r["connection_max"]))):
                    assigned = r["name"]
                    break
            bill["staff_name"] = assigned
    return bills


def _zone_sort_key(zone_name: str) -> tuple:
    """Return sort key for zone ordering: A=1, B=2, C=3, Commercial=4, unknown=99."""
    ZONE_RANK = {"a": 1, "b": 2, "c": 3, "commercial": 4}
    if not zone_name:
        return (99, "")
    n = zone_name.strip().lower().replace("zone ", "").replace("zone", "").strip()
    rank = ZONE_RANK.get(n, 99)
    return (rank, n)


def group_bills(bills: list[dict], group_by: str) -> list[tuple]:
    """Group bills by sector/zone/staff.

    Returns:
      - sector/zone: list of (group_label, [bill_dicts])
      - staff:      list of (zone, staff_name, [bill_dicts])
    """
    if not bills:
        return []
    if group_by == "sector":
        groups = {}
        for bill in bills:
            k = bill.get("sector") or "Unknown Sector"
            groups.setdefault(k, []).append(bill)
        keys = sorted(groups.keys(), key=lambda k: (k == "Unknown Sector", k.lower()))
        return [(k, groups[k]) for k in keys]
    if group_by == "zone":
        groups = {}
        for bill in bills:
            k = bill.get("zone") or "Unknown Zone"
            groups.setdefault(k, []).append(bill)
        keys = sorted(groups.keys(), key=lambda k: (_zone_sort_key(k)[0] if k != "Unknown Zone" else 99, k.lower()))
        return [(k, groups[k]) for k in keys]
    if group_by == "staff":
        bills = map_bills_to_staff(bills)
        zone_groups: dict[str, dict[str, list]] = {}
        for bill in bills:
            z = bill.get("zone") or "Unknown Zone"
            sn = bill.get("staff_name") or "Unassigned"
            zone_groups.setdefault(z, {}).setdefault(sn, []).append(bill)
        z_keys = sorted(zone_groups.keys(), key=_zone_sort_key)
        out = []
        for z in z_keys:
            for sn in sorted(zone_groups[z].keys(), key=str.lower):
                out.append((z, sn, zone_groups[z][sn]))
        return out
    return [("All Bills", bills)]


def generate_grouped_advanced_pdf(
    group_type: str,
    groups: list[tuple],
    filters_applied: str,
    cols_param: str = None,
) -> bytes:
    """Generate a landscape PDF with one section per group showing detailed bill rows."""
    _all_headers = ["Sr", "Bill No", "Reference No", "Connection No", "Consumer Name", "Sector", "Locality", "Zone", "Total Bill", "Arrears", "Amount Received", "Outstanding", "Status", "Mobile No"]
    _all_col_keys = ["sr", "billNo", "referenceNo", "connectionNo", "consumerName", "sector", "locality", "zone", "totalBills", "arrearsReceived", "totalReceivedAmount", "outstanding", "status", "mobileNo"]
    _all_key_map = {k: i for i, k in enumerate(_all_col_keys)}
    if cols_param:
        _sel_keys = [k.strip() for k in cols_param.split(",") if k.strip() in _all_col_keys]
        _col_indices = [_all_key_map[k] for k in _sel_keys]
    else:
        _sel_keys = DEFAULT_ADV_KEYS
        _col_indices = [_all_key_map[k] for k in _sel_keys]
    headers = [_all_headers[i] for i in _col_indices]

    total_outstanding_all = 0
    total_bills_all = 0
    total_amount_all = 0
    for item in groups:
        item_bills = item[2] if len(item) == 3 else item[1]
        for b in item_bills:
            total_bills_all += 1
            total_amount_all += b["total_bill"]
            total_outstanding_all += b["outstanding_amount"]

    group_label_singular = {"sector": "Sector", "zone": "Zone", "staff": "Staff"}.get(group_type, "")
    group_label_plural = {"sector": "Sectors", "zone": "Zones", "staff": "Staff"}.get(group_type, "")

    buf = io.BytesIO()
    margin = 4 * mm
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        topMargin=8 * mm,
        bottomMargin=6 * mm,
        leftMargin=margin,
        rightMargin=margin,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("PDFTitle", parent=styles["Heading1"], fontSize=16, textColor=ACCENT, alignment=1, spaceAfter=3 * mm, fontName="Helvetica-Bold")
    summary_style = ParagraphStyle("PDFSummary", parent=styles["Normal"], fontSize=9, alignment=0, spaceAfter=1 * mm, textColor=colors.HexColor("#333333"), leading=13)
    group_heading_style = ParagraphStyle("GroupHeading", parent=styles["Heading2"], fontSize=12, textColor=ACCENT, spaceBefore=4 * mm, spaceAfter=1 * mm, fontName="Helvetica-Bold", alignment=0)
    group_sub_style = ParagraphStyle("GroupSub", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#555555"), spaceAfter=1.5 * mm, alignment=0, leading=11)

    elements = [Paragraph(f"Advanced Bill Filter Report — {group_label_plural}", title_style)]
    elements.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}", summary_style))
    elements.append(Paragraph(f"<b>Filters:</b> {filters_applied}", summary_style))
    elements.append(Paragraph(f"<b>Total Bills:</b> {total_bills_all:,} &nbsp;&nbsp; <b>Total Outstanding:</b> Rs. {fmt(total_outstanding_all)}", summary_style))
    elements.append(Spacer(1, 3 * mm))

    n = len(headers)
    page_w = landscape(A4)[0] - margin - margin
    col_widths = _calc_col_widths(headers, page_w, n)

    left_cols = {i for i, h in enumerate(headers) if h == "Consumer Name"}

    page_h = landscape(A4)[1] - 8 * mm - 6 * mm
    wrapper = _GroupedPdfWrapper(doc, elements, headers, col_widths, group_heading_style, group_sub_style, group_label_singular, _col_indices, group_type, left_cols, page_h)

    for item in groups:
        if len(item) == 3:
            wrapper.add_staff_group(item[0], item[1], item[2])
        else:
            wrapper.add_group(item[0], item[1])

    wrapper.finish()
    buf.seek(0)
    return buf.getvalue()


class _GroupedPdfWrapper:
    """Helper to write grouped PDF sections with smart pagination and staff zone support."""

    _ROW_H = 6.5        # mm – estimated body row height (with padding)
    _TOTAL_ROW_H = 7.5  # mm – estimated grand total row height

    def __init__(self, doc, elements, headers, col_widths, group_heading_style, group_sub_style, group_label, col_indices, group_type, left_cols, page_h):
        self.doc = doc
        self.elements = elements
        self.headers = headers
        self.col_widths = col_widths
        self.group_heading_style = group_heading_style
        self.group_sub_style = group_sub_style
        self.group_label = group_label
        self.col_indices = col_indices
        self.group_type = group_type
        self.left_cols = left_cols
        self._page_h = page_h       # usable height per page
        self._used = 30 * mm        # title block: title + 3 summary lines + spacer
        self._zone_label = ""       # current zone being rendered (staff-wise)

    # ------------------------------------------------------------------
    # Height estimation helpers
    # ------------------------------------------------------------------
    def _est(self, n_rows: int, sub_count: int = 0) -> float:
        """Estimated mm needed for a group: heading + sub-lines + table header + body + total."""
        heading = 9 * mm          # spaceBefore 4mm + text ~4mm + spaceAfter 1mm
        sub = sub_count * 5 * mm
        summary = 5 * mm          # leading 11pt + spaceAfter 1.5mm
        spacer = 2 * mm
        table_h = 7 * mm + n_rows * self._ROW_H + self._TOTAL_ROW_H
        return heading + sub + summary + spacer + table_h

    _MIN_GROUP = 45 * mm  # minimum mm needed: heading + summary + table header + 2 rows + total

    def _maybe_new_page(self, needed_mm: float):
        """Insert PageBreak before a group if remaining space is too small."""
        if self._used > 0 and self._used + self._MIN_GROUP > self._page_h:
            self.elements.append(PageBreak())
            self._used = 0

    def _track(self, added_mm: float):
        self._used += added_mm
        if self._used >= self._page_h:
            self._used = self._used % self._page_h

    # ------------------------------------------------------------------
    # Group rendering methods
    # ------------------------------------------------------------------
    def add_group(self, group_key, group_bills):
        needed = self._est(len(group_bills))
        self._maybe_new_page(needed)
        self._write_heading(f"{self.group_label}: {group_key}")
        self._write_summary(group_bills)
        self._add_detail_table(group_bills)

    def add_staff_group(self, zone, staff_name, group_bills):
        if zone != self._zone_label:
            needed = 9 * mm + self._est(len(group_bills), sub_count=1)
            self._maybe_new_page(needed)
            self._write_heading(f"Zone: {zone}")
            self._zone_label = zone
        else:
            needed = self._est(len(group_bills), sub_count=1)
            self._maybe_new_page(needed)

        self._write_staff_heading(staff_name, group_bills)
        self._write_summary(group_bills)
        self._add_detail_table(group_bills)

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------
    def _write_heading(self, text):
        self.elements.append(Paragraph(text, self.group_heading_style))
        self._track(9 * mm)

    def _write_staff_heading(self, staff_name, bills):
        display = fmt_staff_name(staff_name).replace("\n", " / ")
        self.elements.append(Paragraph(f"Staff: {display}", self.group_heading_style))

        zones = sorted(set(b["zone"] for b in bills if b.get("zone")), key=lambda z: (_zone_sort_key(z)[0], z.lower()))
        if not zones:
            zones = ["Unknown"]
        zone_prefix = "Zones" if len(zones) > 1 else "Zone"
        self.elements.append(Paragraph(f"<b>{zone_prefix}:</b> {', '.join(zones)}", self.group_sub_style))
        self._track(5 * mm)

        sectors = sorted(set(b["sector"] for b in bills if b.get("sector")))
        ctx_parts = []
        if sectors:
            ctx_parts.append(f"<b>Sectors:</b> {', '.join(sectors)}")
        if ctx_parts:
            self.elements.append(Paragraph(" &nbsp;|&nbsp; ".join(ctx_parts), self.group_sub_style))
            self._track(5 * mm)
        self._track(9 * mm)

    def _write_summary(self, bills):
        total_b = len(bills)
        total_amt = sum(b["total_bill"] for b in bills)
        total_rec = sum(b["amount_received"] for b in bills)
        total_out = sum(b["outstanding_amount"] for b in bills)
        total_arr = sum(b["arrears"] for b in bills)
        self.elements.append(Paragraph(
            f"<b>Bills:</b> {total_b:,} &nbsp;|&nbsp; <b>Total Amount:</b> Rs. {fmt(total_amt)} &nbsp;|&nbsp; "
            f"<b>Received:</b> Rs. {fmt(total_rec)} &nbsp;|&nbsp; <b>Outstanding:</b> Rs. {fmt(total_out)} &nbsp;|&nbsp; "
            f"<b>Arrears:</b> Rs. {fmt(total_arr)}",
            self.group_sub_style,
        ))
        self.elements.append(Spacer(1, 2 * mm))
        self._track(5 * mm + 2 * mm)

    def _add_detail_table(self, group_bills):
        hdr_style = ParagraphStyle("GpHdr", fontSize=8, leading=10, alignment=1, textColor=HEADER_FG, fontName="Helvetica-Bold")
        wrapped_headers = [Paragraph(str(h), hdr_style) for h in self.headers]
        data = [wrapped_headers]

        for idx, bill in enumerate(group_bills, start=1):
            _full = [
                idx,
                bill["bill_no"] or "",
                bill["reference_no"] or "",
                bill["connection_no"] or "",
                bill.get("consumer_name") or "",
                bill["sector"],
                bill["locality"],
                bill["zone"],
                fmt(bill["total_bill"]),
                fmt(bill["arrears"]),
                fmt(bill["amount_received"]),
                fmt(bill["outstanding_amount"]),
                bill["status"] or "",
                bill.get("consumer_mobile") or "",
            ]
            row = [_full[i] for i in self.col_indices]
            data.append(row)

        g_total_b = sum(b["total_bill"] for b in group_bills)
        g_total_r = sum(b["amount_received"] for b in group_bills)
        g_total_o = sum(b["outstanding_amount"] for b in group_bills)
        g_total_a = sum(b["arrears"] for b in group_bills)
        _full_g = ["", "", "", "", "", "", "", "Group Total", fmt(g_total_b), fmt(g_total_a), fmt(g_total_r), fmt(g_total_o), "", ""]
        grand_row = [_full_g[i] for i in self.col_indices]
        data.append(grand_row)

        body_rows = wrap_pdf_body_cells(data[1:], font_size=7, left_columns=self.left_cols)
        all_rows = [data[0]] + body_rows

        t = _make_pdf_table(
            all_rows,
            col_widths=self.col_widths,
            header_font_size=8,
            body_font_size=7,
            cell_padding=5,
            left_cols=self.left_cols,
        )
        self.elements.append(t)
        self.elements.append(Spacer(1, 3 * mm))
        table_est = 7 * mm + len(group_bills) * self._ROW_H + self._TOTAL_ROW_H + 3 * mm
        self._track(table_est)

    def finish(self):
        self.doc.build(self.elements)


def _calc_col_widths(headers, page_w, n):
    """Calculate column widths proportional to content, scaled to fill page_w."""
    prop = {
        "Sr": 3.5,
        "Bill No": 9,
        "Reference No": 9,
        "Connection No": 11,
        "Consumer Name": 24,
        "Sector": 9,
        "Locality": 14,
        "Zone": 7,
        "Total Bill": 10,
        "Arrears": 10,
        "Amount Received": 10,
        "Outstanding": 10,
        "Status": 7,
        "Mobile No": 12,
    }
    widths = []
    total_p = 0
    for h in headers:
        p = prop.get(h, 10)
        widths.append(p)
        total_p += p
    if total_p > 0:
        widths = [page_w * w / total_p for w in widths]
    return widths


def sanitize_filename(name: str) -> str:
    safe = re.sub(r'[<>:"/\\|?*]', '_', name)
    safe = safe.replace(' ', '_')
    safe = re.sub(r'_+', '_', safe)
    safe = safe.strip('_.')
    return safe or "Unknown"


def generate_single_group_pdf(
    group_type: str,
    group_key: str,
    group_bills: list,
    filters_applied: str,
    cols_param: str = None,
    staff_zone: str = None,
) -> bytes:
    _all_headers = ["Sr", "Bill No", "Reference No", "Connection No", "Consumer Name", "Sector", "Locality", "Zone", "Total Bill", "Arrears", "Amount Received", "Outstanding", "Status", "Mobile No"]
    _all_col_keys = ["sr", "billNo", "referenceNo", "connectionNo", "consumerName", "sector", "locality", "zone", "totalBills", "arrearsReceived", "totalReceivedAmount", "outstanding", "status", "mobileNo"]
    _all_key_map = {k: i for i, k in enumerate(_all_col_keys)}
    if cols_param:
        _sel_keys = [k.strip() for k in cols_param.split(",") if k.strip() in _all_col_keys]
        _col_indices = [_all_key_map[k] for k in _sel_keys]
    else:
        _sel_keys = DEFAULT_ADV_KEYS
        _col_indices = [_all_key_map[k] for k in _sel_keys]
    headers = [_all_headers[i] for i in _col_indices]

    total_bills = len(group_bills)
    total_amount = sum(b["total_bill"] for b in group_bills)
    total_outstanding = sum(b["outstanding_amount"] for b in group_bills)

    group_label = {"sector": "Sector", "zone": "Zone", "staff": "Staff"}.get(group_type, "")

    buf = io.BytesIO()
    margin = 4 * mm
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        topMargin=8 * mm,
        bottomMargin=6 * mm,
        leftMargin=margin,
        rightMargin=margin,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("PDFTitle", parent=styles["Heading1"], fontSize=16, textColor=ACCENT, alignment=1, spaceAfter=3 * mm, fontName="Helvetica-Bold")
    summary_style = ParagraphStyle("PDFSummary", parent=styles["Normal"], fontSize=9, alignment=0, spaceAfter=1 * mm, textColor=colors.HexColor("#333333"), leading=13)
    group_heading_style = ParagraphStyle("GroupHeading", parent=styles["Heading2"], fontSize=12, textColor=ACCENT, spaceBefore=4 * mm, spaceAfter=1 * mm, fontName="Helvetica-Bold", alignment=0)
    group_sub_style = ParagraphStyle("GroupSub", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#555555"), spaceAfter=1.5 * mm, alignment=0, leading=11)

    elements = [Paragraph(f"Advanced Bill — {group_label}: {group_key}", title_style)]
    elements.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}", summary_style))
    elements.append(Paragraph(f"<b>Filters:</b> {filters_applied}", summary_style))
    elements.append(Paragraph(f"<b>Total Bills:</b> {total_bills:,} &nbsp;&nbsp; <b>Total Outstanding:</b> Rs. {fmt(total_outstanding)}", summary_style))
    elements.append(Spacer(1, 3 * mm))

    n = len(headers)
    page_w = landscape(A4)[0] - margin - margin
    col_widths = _calc_col_widths(headers, page_w, n)

    left_cols = {i for i, h in enumerate(headers) if h == "Consumer Name"}

    page_h = landscape(A4)[1] - 8 * mm - 6 * mm
    wrapper = _GroupedPdfWrapper(doc, elements, headers, col_widths, group_heading_style, group_sub_style, group_label, _col_indices, group_type, left_cols, page_h)

    if group_type == "staff" and staff_zone:
        wrapper.add_staff_group(staff_zone, group_key, group_bills)
    else:
        wrapper.add_group(group_key, group_bills)

    wrapper.finish()
    buf.seek(0)
    return buf.getvalue()


def generate_zip_of_group_pdfs(
    group_type: str,
    groups: list[tuple],
    filters_applied: str,
    cols_param: str = None,
) -> io.BytesIO:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        filenames_used = set()
        for item in groups:
            if group_type == "staff":
                zone, staff_name, group_bills = item
                base_name = sanitize_filename(f"Staff_{staff_name}")
                pdf_bytes = generate_single_group_pdf(group_type, staff_name, group_bills, filters_applied, cols_param, staff_zone=zone)
            elif group_type == "zone":
                zone_name, group_bills = item
                base_name = sanitize_filename(f"Zone_{zone_name}")
                pdf_bytes = generate_single_group_pdf(group_type, zone_name, group_bills, filters_applied, cols_param)
            elif group_type == "sector":
                sector_name, group_bills = item
                base_name = sanitize_filename(f"Sector_{sector_name}")
                pdf_bytes = generate_single_group_pdf(group_type, sector_name, group_bills, filters_applied, cols_param)
            else:
                continue
            filename = f"{base_name}.pdf"
            if filename in filenames_used:
                counter = 2
                while f"{base_name}_{counter}.pdf" in filenames_used:
                    counter += 1
                filename = f"{base_name}_{counter}.pdf"
            filenames_used.add(filename)
            zf.writestr(filename, pdf_bytes)
    buf.seek(0)
    return buf


# Default selected column keys for Advanced Bill Checking in display order
DEFAULT_ADV_KEYS = ["sr", "connectionNo", "consumerName", "mobileNo", "locality", "totalBills", "arrearsReceived", "outstanding"]


def generate_advanced_filtered_pdf(bills: list[dict], filters_applied: str, show_summary: bool = True, cols_param: str = None) -> bytes:
    if not bills:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=18, textColor=ACCENT, alignment=1)
        elements = [Paragraph("Advanced Bill Filter Report", title_style), Paragraph("No bills match the selected filters.", styles["Normal"])]
        doc.build(elements)
        buf.seek(0)
        return buf.getvalue()

    _all_headers = ["Sr", "Bill No", "Reference No", "Connection No", "Consumer Name", "Sector", "Locality", "Zone", "Total Bill", "Arrears", "Amount Received", "Outstanding", "Status", "Mobile No"]
    _all_col_keys = ["sr", "billNo", "referenceNo", "connectionNo", "consumerName", "sector", "locality", "zone", "totalBills", "arrearsReceived", "totalReceivedAmount", "outstanding", "status", "mobileNo"]
    _all_key_map = {k: i for i, k in enumerate(_all_col_keys)}
    if cols_param:
        _sel_keys = [k.strip() for k in cols_param.split(",") if k.strip() in _all_col_keys]
        _pdf_col_indices = [_all_key_map[k] for k in _sel_keys]
    else:
        _sel_keys = DEFAULT_ADV_KEYS
        _pdf_col_indices = [_all_key_map[k] for k in _sel_keys]
    headers = [_all_headers[i] for i in _pdf_col_indices]

    rows = []
    for idx, bill in enumerate(bills, start=1):
        _full_row = [
            idx,
            bill["bill_no"] or "",
            bill["reference_no"] or "",
            bill["connection_no"] or "",
            bill.get("consumer_name") or "",
            bill["sector"],
            bill["locality"],
            bill["zone"],
            fmt(bill["total_bill"]),
            fmt(bill["arrears"]),
            fmt(bill["amount_received"]),
            fmt(bill["outstanding_amount"]),
            bill["status"] or "",
            bill.get("consumer_mobile") or "",
        ]
        rows.append([_full_row[i] for i in _pdf_col_indices])

    total_bill = sum(b["total_bill"] for b in bills)
    amount_received = sum(b["amount_received"] for b in bills)
    outstanding = sum(b["outstanding_amount"] for b in bills)
    _full_grand = ["", "", "", "", "", "", "", "Grand Total", fmt(total_bill), fmt(amount_received), fmt(outstanding), "", "", ""]
    grand_total = [_full_grand[i] for i in _pdf_col_indices]

    page_w = landscape(A4)[0] - 8 * mm
    n = len(headers)
    col_widths = _calc_col_widths(headers, page_w, n)

    left_cols = {i for i, h in enumerate(headers) if h == "Consumer Name"}

    summary_lines = [
        f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}",
        f"<b>Filters:</b> {filters_applied}",
        f"<b>Total Bills:</b> {len(bills)}",
        f"<b>Total Outstanding:</b> Rs. {fmt(outstanding)}",
    ] if show_summary else []

    return generate_card_pdf(
        "Advanced Bill Filter Report",
        summary_lines,
        headers,
        rows,
        grand_total,
        pagesize=landscape(A4),
        col_widths=col_widths,
        left_cols=left_cols,
        header_font_size=8,
        body_font_size=8,
        cell_padding=6,
    )


def export_advanced_bills_response(fmt_type: str, bills: list[dict], filters_applied: str, show_summary: bool = True, cols_param: str = None, group_by: str = "normal"):
    if not bills:
        flash("No bills match the selected filters.")
        return redirect(url_for("bill_list"))

    # Unified 14-column set matching PDF format
    _all_headers_adv = ["Sr", "Bill No", "Reference No", "Connection No", "Consumer Name", "Sector", "Locality", "Zone", "Total Bill", "Arrears", "Amount Received", "Outstanding", "Status", "Mobile No"]
    _all_adv_keys = ["sr", "billNo", "referenceNo", "connectionNo", "consumerName", "sector", "locality", "zone", "totalBills", "arrearsReceived", "totalReceivedAmount", "outstanding", "status", "mobileNo"]
    _adv_key_map = {k: i for i, k in enumerate(_all_adv_keys)}
    if cols_param:
        _sel = [k.strip() for k in cols_param.split(",") if k.strip() in _all_adv_keys]
        _adv_cols = [_adv_key_map[k] for k in _sel]
    else:
        _sel = DEFAULT_ADV_KEYS
        _adv_cols = [_adv_key_map[k] for k in _sel]
    headers = [_all_headers_adv[i] for i in _adv_cols]
    rows = [
        [
            idx,
            bill["bill_no"] or "",
            bill["reference_no"] or "",
            bill["connection_no"] or "",
            bill.get("consumer_name") or "",
            bill["sector"],
            bill["locality"],
            bill["zone"],
            bill["total_bill"],
            bill["arrears"],
            bill["amount_received"],
            bill["outstanding_amount"],
            bill["status"] or "",
            bill.get("consumer_mobile") or "",
        ]
        for idx, bill in enumerate(bills, start=1)
    ]
    rows = [[r[i] for i in _adv_cols] for r in rows]
    # Find mobile column index in filtered output
    mobile_col_idx = None
    for ci, h in enumerate(headers):
        if h == "Mobile No":
            mobile_col_idx = ci
            break

    filename = f"advanced_bills_{datetime.now().strftime('%Y%m%d%H%M%S')}"

    if fmt_type == "pdf":
        if group_by in ("sector", "zone", "staff"):
            groups = group_bills(bills, group_by)
            pdf_bytes = generate_grouped_advanced_pdf(group_by, groups, filters_applied, cols_param=cols_param)
        else:
            pdf_bytes = generate_advanced_filtered_pdf(bills, filters_applied, show_summary=show_summary, cols_param=cols_param)
        return Response(pdf_bytes, mimetype="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}.pdf"})

    if fmt_type == "csv":
        if mobile_col_idx is not None:
            for r in rows:
                mv = r[mobile_col_idx]
                if mv and mv != "-" and mv[0].isdigit():
                    r[mobile_col_idx] = '="' + mv + '"'
        df = pd.DataFrame(rows, columns=headers)
        csv_data = df.to_csv(index=False)
        return Response(csv_data, mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}.csv"})

    if fmt_type == "xlsx":
        buf = io.BytesIO()
        df = pd.DataFrame(rows, columns=headers)
        if mobile_col_idx is not None:
            from openpyxl.utils import get_column_letter
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                ws = writer.sheets['Sheet1']
                col_letter = get_column_letter(mobile_col_idx + 1)
                for ri in range(2, len(rows) + 2):
                    ws[f'{col_letter}{ri}'].number_format = '@'
            buf.seek(0)
        else:
            df.to_excel(buf, index=False)
            buf.seek(0)
        return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"})

    flash("Unknown export format.")
    return redirect(url_for("bill_list"))


def build_unpaid_amount_summary(conn, bill_ids: set[int] | None = None) -> dict:
    _bid_sql = ""
    _bid_sql_b = ""
    if bill_ids is not None:
        if not bill_ids:
            _bid_sql = "AND 1=0"
            _bid_sql_b = "AND 1=0"
        else:
            id_list = ",".join(str(i) for i in bill_ids)
            _bid_sql = f"AND id IN ({id_list})"
            _bid_sql_b = f"AND b.id IN ({id_list})"

    metric_select = """
        COUNT(*) AS bill_count,
        SUM(COALESCE(total_bill, 0)) AS total_bill_amount,
        SUM(COALESCE(arrears, 0)) AS total_arrears_amount,
        SUM(COALESCE(total_bill, 0) - COALESCE(arrears, 0)) AS current_bill_amount
    """
    total = conn.execute(
        f"""
        SELECT {metric_select}
        FROM bills
        WHERE {UNPAID_STATUS_SQL}
        {_bid_sql}
        """
    ).fetchone()

    sector_rows = conn.execute(
        f"""
        SELECT sector AS name, {metric_select}
        FROM bills
        WHERE {UNPAID_STATUS_SQL}
        {_bid_sql}
        GROUP BY sector
        ORDER BY LOWER(sector)
        """
    ).fetchall()

    zone_rows = conn.execute(
        f"""
        SELECT zone AS name, {metric_select}
        FROM bills
        WHERE {UNPAID_STATUS_SQL}
        {_bid_sql}
        GROUP BY zone
        ORDER BY {zone_sort_expr('zone')}, LOWER(zone)
        """
    ).fetchall()

    has_auto_rules = conn.execute("SELECT COUNT(*) AS cnt FROM auto_assignment_rules").fetchone()["cnt"] > 0
    unpaid_where = UNPAID_STATUS_SQL.replace("status", "b.status")

    if has_auto_rules:
        auto_exclude_clause = "AND NOT EXISTS (SELECT 1 FROM auto_assignment_rules aar WHERE aar.sector = b.sector AND aar.locality = b.locality)"
        auto_assign_sql = f"""
        UNION ALL
        SELECT
            s.name AS name,
            COUNT(b.id) AS bill_count,
            SUM(COALESCE(b.total_bill, 0)) AS total_bill_amount,
            SUM(COALESCE(b.arrears, 0)) AS total_arrears_amount,
            SUM(COALESCE(b.total_bill, 0) - COALESCE(b.arrears, 0)) AS current_bill_amount
        FROM bills b
        JOIN auto_assignment_rules aar
            ON aar.sector = b.sector AND aar.locality = b.locality
            AND CAST(b.connection_no AS TEXT) >= aar.connection_min
            AND (aar.connection_max IS NULL OR CAST(b.connection_no AS TEXT) <= aar.connection_max)
        JOIN staff s ON UPPER(TRIM(s.name)) = UPPER(TRIM(aar.staff_name))
        WHERE {unpaid_where}
        {_bid_sql_b}
        GROUP BY s.id, s.name
        """
    else:
        auto_exclude_clause = ""
        auto_assign_sql = ""

    staff_rows = conn.execute(
        f"""
        SELECT
            s.name AS name,
            COUNT(b.id) AS bill_count,
            SUM(COALESCE(b.total_bill, 0)) AS total_bill_amount,
            SUM(COALESCE(b.arrears, 0)) AS total_arrears_amount,
            SUM(COALESCE(b.total_bill, 0) - COALESCE(b.arrears, 0)) AS current_bill_amount
        FROM staff_assignments sa
        JOIN staff s ON s.id = sa.staff_id
        JOIN bills b
            ON b.zone = sa.zone
            AND (sa.sector IS NULL OR b.sector = sa.sector)
            AND (sa.locality IS NULL OR b.locality = sa.locality)
            {auto_exclude_clause}
        WHERE {unpaid_where}
        {_bid_sql_b}
        GROUP BY s.id, s.name

        {auto_assign_sql}

        UNION ALL
        SELECT
            'Unassigned' AS name,
            COUNT(b.id) AS bill_count,
            SUM(COALESCE(b.total_bill, 0)) AS total_bill_amount,
            SUM(COALESCE(b.arrears, 0)) AS total_arrears_amount,
            SUM(COALESCE(b.total_bill, 0) - COALESCE(b.arrears, 0)) AS current_bill_amount
        FROM bills b
        WHERE {unpaid_where}
        {_bid_sql_b}
          AND NOT EXISTS (
              SELECT 1
              FROM staff_assignments sa
              WHERE sa.zone = b.zone
                AND (sa.sector IS NULL OR sa.sector = b.sector)
                AND (sa.locality IS NULL OR sa.locality = b.locality)
          )
          {('AND NOT EXISTS (SELECT 1 FROM auto_assignment_rules aar WHERE aar.sector = b.sector AND aar.locality = b.locality)' if has_auto_rules else '')}
        HAVING COUNT(b.id) > 0
        ORDER BY name
        """
    ).fetchall()

    def to_amount_row(row, idx: int | None = None) -> dict:
        item = {
            "name": row["name"] if row["name"] else "Unknown",
            "bill_count": int(row["bill_count"] or 0),
            "total_bill_amount": float(row["total_bill_amount"] or 0),
            "total_arrears_amount": float(row["total_arrears_amount"] or 0),
            "current_bill_amount": float(row["current_bill_amount"] or 0),
        }
        if idx is not None:
            item["sr"] = idx
        return item

    return {
        "total": {
            "bill_count": int(total["bill_count"] or 0),
            "total_bill_amount": float(total["total_bill_amount"] or 0),
            "total_arrears_amount": float(total["total_arrears_amount"] or 0),
            "current_bill_amount": float(total["current_bill_amount"] or 0),
        },
        "sector_rows": [to_amount_row(row, idx) for idx, row in enumerate(sector_rows, start=1)],
        "zone_rows": [to_amount_row(row, idx) for idx, row in enumerate(zone_rows, start=1)],
        "staff_rows": [to_amount_row(row, idx) for idx, row in enumerate(staff_rows, start=1)],
    }


def get_bill_list_context():
    init_bill_list_db()
    with get_db() as conn:
        unpaid_amount_summary = build_unpaid_amount_summary(conn)
        for row in unpaid_amount_summary["staff_rows"]:
            row["name"] = fmt_staff_name(row["name"])
        rows = conn.execute(
            """
            SELECT
                sector,
                COUNT(*) AS total_bills,
                SUM(CASE WHEN amount_received > 0 THEN 1 ELSE 0 END) AS received_bills,
                SUM(amount_received) AS total_received_amount,
                SUM(CASE WHEN total_bill > amount_received THEN total_bill - amount_received ELSE 0 END) AS remaining_amount
            FROM bills
            GROUP BY sector
            ORDER BY sector
            """
        ).fetchall()
        report_rows = []
        for idx, row in enumerate(rows, start=1):
            total_bills = int(row["total_bills"] or 0)
            received_bills = int(row["received_bills"] or 0)
            report_rows.append(
                {
                    "sr": idx,
                    "sector": row["sector"],
                    "total_bills": total_bills,
                    "received_bills": received_bills,
                    "remaining_bills": total_bills - received_bills,
                    "total_received_amount": float(row["total_received_amount"] or 0),
                    "remaining_amount": float(row["remaining_amount"] or 0),
                }
            )

        summary = conn.execute(
            """
            SELECT
                COUNT(*) AS total_bills,
                COUNT(DISTINCT NULLIF(connection_no, '')) AS total_connections,
                SUM(CASE WHEN amount_received > 0 THEN 1 ELSE 0 END) AS received_bills,
                SUM(amount_received) AS total_received_amount,
                SUM(CASE WHEN total_bill > amount_received THEN total_bill - amount_received ELSE 0 END) AS remaining_amount
            FROM bills
            """
        ).fetchone()
        total_connections = int(summary["total_connections"] or 0)
        total_bills = int(summary["total_bills"] or 0)
        if total_connections == 0:
            total_connections = total_bills
        received_bills = int(summary["received_bills"] or 0)

        zones = conn.execute(
            """
            SELECT zone, sector, COUNT(DISTINCT locality) AS locality_count, COUNT(*) AS bill_count
            FROM bills
            GROUP BY zone, sector
            ORDER BY CASE zone WHEN 'A' THEN 1 WHEN 'B' THEN 2 WHEN 'C' THEN 3 WHEN 'Commercial' THEN 4 ELSE 5 END, sector
            """
        ).fetchall()
        localities = conn.execute(
            """
            SELECT zone, sector, locality
            FROM localities
            ORDER BY CASE zone WHEN 'A' THEN 1 WHEN 'B' THEN 2 WHEN 'C' THEN 3 WHEN 'Commercial' THEN 4 ELSE 5 END, sector, locality
            """
        ).fetchall()
        staff = conn.execute("SELECT id, name FROM staff ORDER BY name").fetchall()
        assignments = conn.execute(
            """
            SELECT
                sa.id,
                sa.staff_id,
                s.name,
                sa.zone,
                COALESCE(sa.sector, 'All sectors') AS sector,
                COALESCE(sa.locality, '') AS locality
            FROM staff_assignments sa
            JOIN staff s ON s.id = sa.staff_id
            ORDER BY sa.zone, sa.sector, sa.locality, s.name
            """
        ).fetchall()
        sectors = conn.execute(
            f"""
            SELECT name AS sector, zone
            FROM sectors
            ORDER BY {zone_sort_expr('zone')}, LOWER(name)
            """
        ).fetchall()
        assigned_options = conn.execute(
            """
            SELECT
                sa.staff_id,
                sa.zone,
                COALESCE(sa.sector, '*') AS sector,
                COALESCE(sa.locality, '*') AS locality,
                s.name AS staff_name
            FROM staff_assignments sa
            JOIN staff s ON s.id = sa.staff_id
            """
        ).fetchall()

    assigned_exact = {
        f"{row['zone']}||{row['sector']}": {"name": row["staff_name"], "id": str(row["staff_id"])}
        for row in assigned_options
        if row["sector"] != "*" and row["locality"] == "*"
    }
    assigned_localities = {
        f"{row['zone']}||{row['sector']}||{row['locality']}": {"name": row["staff_name"], "id": str(row["staff_id"])}
        for row in assigned_options
        if row["sector"] != "*" and row["locality"] != "*"
    }
    assigned_zones = {
        row["zone"]: {"name": row["staff_name"], "id": str(row["staff_id"])}
        for row in assigned_options
        if row["sector"] == "*"
    }
    sector_rows_by_key = {}
    for row in sectors:
        sector = dict(row)
        key = f"{sector['zone']}||{sector['sector']}"
        assigned = assigned_zones.get(sector["zone"]) or assigned_exact.get(key)
        sector["assigned_to"] = assigned["name"] if assigned else None
        sector["assigned_staff_id"] = assigned["id"] if assigned else ""
        sector["assigned"] = assigned is not None
        sector["sector_values"] = [sector["sector"]]
        display_key = normalize_sector_key(sector["sector"])
        existing = sector_rows_by_key.get(display_key)
        if not existing:
            sector_rows_by_key[display_key] = sector
        else:
            existing["sector_values"].append(sector["sector"])
            if len(sector["sector"]) < len(existing["sector"]):
                existing["sector"] = sector["sector"]
            if existing["zone"] == "Unassigned" and sector["zone"] != "Unassigned":
                existing["zone"] = sector["zone"]
            if not existing["assigned"] and sector["assigned"]:
                existing["assigned_to"] = sector["assigned_to"]
                existing["assigned_staff_id"] = sector["assigned_staff_id"]
                existing["assigned"] = True
    sector_rows = sorted(
        sector_rows_by_key.values(),
        key=lambda row: (
            {"A": 1, "B": 2, "C": 3, "Commercial": 4, "Unassigned": 5}.get(row["zone"], 6),
            row["sector"].lower(),
        ),
    )
    locality_rows = []
    for row in localities:
        locality = dict(row)
        locality_key = f"{locality['zone']}||{locality['sector']}||{locality['locality']}"
        sector_key = f"{locality['zone']}||{locality['sector']}"
        assigned = assigned_zones.get(locality["zone"]) or assigned_exact.get(sector_key) or assigned_localities.get(locality_key)
        locality["assigned_to"] = assigned["name"] if assigned else None
        locality["assigned_staff_id"] = assigned["id"] if assigned else ""
        locality["assigned"] = assigned is not None
        locality_rows.append(locality)

    return {
        "summary": {
            "total_connections": total_connections,
            "total_bills": total_bills,
            "received_bills": received_bills,
            "remaining_bills": total_bills - received_bills,
            "total_received_amount": float(summary["total_received_amount"] or 0),
            "remaining_amount": float(summary["remaining_amount"] or 0),
        },
        "report_rows": report_rows,
        "zones": [dict(row) for row in zones],
        "localities": locality_rows,
        "staff": [dict(row) for row in staff],
        "assignments": [dict(row) for row in assignments],
        "sectors": sector_rows,
        "zone_options": ["A", "B", "C", "Commercial", "Unassigned"],
        "unpaid_amount_summary": unpaid_amount_summary,
        "summary_report_zones": get_zones_summary(),
        "summary_report_sectors": get_sectors_summary(),
        "summary_report_staff": get_staff_summary(),
        "staff_report_rows": bill_list_staff_export_rows()[1],
    }


def bill_list_export_rows():
    context = get_bill_list_context()
    headers = [
        "Sr",
        "Sector",
        "Total Bills",
        "Received Bills",
        "Remaining Bills",
        "Total Received Amount",
        "Pending Amount",
    ]
    rows = [
        [
            row["sr"],
            row["sector"],
            fmt(row["total_bills"]),
            fmt(row["received_bills"]),
            fmt(row["remaining_bills"]),
            fmt(row["total_received_amount"]),
            fmt(row["remaining_amount"]),
        ]
        for row in context["report_rows"]
    ]
    return headers, rows, context["summary"]


def _bill_list_summary_from_rows(rows: list[list]) -> dict:
    """Recalculate sector export totals from checked rows only."""
    return {
        "total_bills": sum(parse_number(row[2]) for row in rows),
        "received_bills": sum(parse_number(row[3]) for row in rows),
        "remaining_bills": sum(parse_number(row[4]) for row in rows),
        "total_received_amount": sum(parse_number(row[5]) for row in rows),
        "remaining_amount": sum(parse_number(row[6]) for row in rows),
    }


def unpaid_amount_export_data():
    init_bill_list_db()
    with get_db() as conn:
        summary = build_unpaid_amount_summary(conn)

    amount_headers = [
        "Sr",
        "Name",
        "Bills",
        "Total Bill Amount",
        "Total Arrears Amount",
        "Current Bill Amount",
    ]
    total = summary["total"]
    summary_headers = ["Bills", "Total Bill Amount", "Total Arrears Amount", "Current Bill Amount"]
    summary_rows = [[
        fmt(total["bill_count"]),
        fmt(total["total_bill_amount"]),
        fmt(total["total_arrears_amount"]),
        fmt(total["current_bill_amount"]),
    ]]

    def make_rows(source_rows: list[dict], *, is_staff: bool = False) -> list[list]:
        return [
            [
                row["sr"],
                fmt_staff_name(row["name"]) if is_staff else row["name"],
                fmt(row["bill_count"]),
                fmt(row["total_bill_amount"]),
                fmt(row["total_arrears_amount"]),
                fmt(row["current_bill_amount"]),
            ]
            for row in source_rows
        ]

    sections = [
        ("summary", "Summary Totals", summary_headers, summary_rows),
        ("sector", "Sector-wise", amount_headers, make_rows(summary["sector_rows"])),
        ("zone", "Zone-wise", amount_headers, make_rows(summary["zone_rows"])),
        ("staff", "Staff-wise", amount_headers, make_rows(summary["staff_rows"], is_staff=True)),
    ]
    return sections, total


def unpaid_amount_section(sections: list[tuple[str, str, list[str], list[list]]], section_key: str):
    for key, title, headers, rows in sections:
        if key == section_key:
            return key, title, headers, rows
    return None


def _unpaid_total_from_section_rows(rows: list[list]) -> dict:
    """Recalculate unpaid amount totals from checked section rows only."""
    return {
        "bill_count": sum(parse_number(row[2]) for row in rows),
        "total_bill_amount": sum(parse_number(row[3]) for row in rows),
        "total_arrears_amount": sum(parse_number(row[4]) for row in rows),
        "current_bill_amount": sum(parse_number(row[5]) for row in rows),
    }


def generate_unpaid_amount_pdf(sections: list[tuple[str, str, list[str], list[list]]], total: dict, show_summary: bool = True) -> bytes:
    buf = io.BytesIO()
    pagesize = landscape(A4)
    left_margin = right_margin = 15 * mm
    doc = SimpleDocTemplate(
        buf,
        pagesize=pagesize,
        topMargin=18 * mm,
        bottomMargin=14 * mm,
        leftMargin=left_margin,
        rightMargin=right_margin,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "UnpaidAmountTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=ACCENT,
        alignment=1,
        spaceAfter=5 * mm,
        fontName="Helvetica-Bold",
    )
    section_style = ParagraphStyle(
        "UnpaidAmountSection",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=colors.HexColor("#333333"),
        spaceBefore=4 * mm,
        spaceAfter=2 * mm,
        fontName="Helvetica-Bold",
    )
    summary_style = ParagraphStyle(
        "UnpaidAmountSummary",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#333333"),
        spaceAfter=1 * mm,
    )
    elements = [
        Paragraph("Unpaid & Expired Bill Amounts", title_style),
    ]
    if show_summary:
        elements.extend([
            Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}", summary_style),
            Paragraph(f"<b>Bills:</b> {fmt(total['bill_count'])}", summary_style),
            Paragraph(f"<b>Total Bill Amount:</b> Rs. {fmt(total['total_bill_amount'])}", summary_style),
            Paragraph(f"<b>Total Arrears Amount:</b> Rs. {fmt(total['total_arrears_amount'])}", summary_style),
            Paragraph(f"<b>Current Bill Amount:</b> Rs. {fmt(total['current_bill_amount'])}", summary_style),
            Spacer(1, 4 * mm),
        ])
    page_w = pagesize[0] - left_margin - right_margin
    for _, title, headers, rows in sections:
        section_rows = rows or [["", "No unpaid or expired bills found.", "", "", "", ""][:len(headers)]]
        if len(headers) == 4:
            col_widths = [page_w * 0.18, page_w * 0.27, page_w * 0.27, page_w * 0.28]
            left_cols = []
        else:
            col_widths = [page_w * 0.05, page_w * 0.35, page_w * 0.10, page_w * 0.17, page_w * 0.17, page_w * 0.16]
            left_cols = [1]
        elements.append(Paragraph(title, section_style))
        elements.append(
            _make_pdf_table(
                [headers] + section_rows,
                col_widths=col_widths,
                left_cols=left_cols,
                header_font_size=8,
                body_font_size=8,
                cell_padding=5,
            )
        )
        elements.append(Spacer(1, 4 * mm))
    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


def zone_sort_expr(column_name: str = "zone") -> str:
    return f"CASE {column_name} WHEN 'A' THEN 1 WHEN 'B' THEN 2 WHEN 'C' THEN 3 WHEN 'Commercial' THEN 4 ELSE 5 END"


def bill_list_zone_export_rows(selected_zone: str = "All"):
    init_bill_list_db()
    headers = [
        "Sr",
        "Zone",
        "Sector",
        "Total Bills",
        "Received Bills",
        "Remaining Bills",
        "Amount Received",
        "Pending Amount",
    ]
    zone_filter = "" if selected_zone == "All" else "WHERE zone = ?"
    params = () if selected_zone == "All" else (selected_zone,)
    with get_db() as conn:
        data = conn.execute(
            f"""
            SELECT
                zone,
                sector,
                COUNT(*) AS total_bills,
                SUM(CASE WHEN amount_received > 0 THEN 1 ELSE 0 END) AS received_bills,
                SUM(amount_received) AS total_received_amount,
                SUM(CASE WHEN total_bill > amount_received THEN total_bill - amount_received ELSE 0 END) AS remaining_amount
            FROM bills
            {zone_filter}
            GROUP BY zone, sector
            ORDER BY {zone_sort_expr('zone')}, sector
            """,
            params,
        ).fetchall()
    def make_total_row(label: str, source_rows: list[list], zone: str = "") -> list:
        return [
            "",
            zone,
            label,
            fmt(sum(parse_number(str(row[3]).replace(",", "")) for row in source_rows)),
            fmt(sum(parse_number(str(row[4]).replace(",", "")) for row in source_rows)),
            fmt(sum(parse_number(str(row[5]).replace(",", "")) for row in source_rows)),
            fmt(sum(parse_number(str(row[6]).replace(",", "")) for row in source_rows)),
            fmt(sum(parse_number(str(row[7]).replace(",", "")) for row in source_rows)),
        ]

    base_rows = []
    for idx, row in enumerate(data, start=1):
        total_bills = int(row["total_bills"] or 0)
        received_bills = int(row["received_bills"] or 0)
        base_rows.append(
            [
                idx,
                row["zone"],
                row["sector"],
                fmt(total_bills),
                fmt(received_bills),
                fmt(total_bills - received_bills),
                fmt(row["total_received_amount"] or 0),
                fmt(row["remaining_amount"] or 0),
            ]
        )
    base_rows = _filter_rows_by_selection(base_rows, lambda row: f"{row[1]}|||{row[2]}")

    rows = []
    if selected_zone == "All":
        current_zone = None
        zone_rows = []
        for row in base_rows:
            if current_zone is not None and row[1] != current_zone:
                rows.extend(zone_rows)
                rows.append(make_total_row(f"{current_zone} Total", zone_rows, current_zone))
                zone_rows = []
            current_zone = row[1]
            zone_rows.append(row)
        if zone_rows:
            rows.extend(zone_rows)
            rows.append(make_total_row(f"{current_zone} Total", zone_rows, current_zone))
        if base_rows:
            rows.append(make_total_row("Grand Total", base_rows, "All"))
    else:
        rows = base_rows
        if base_rows:
            rows.append(make_total_row(f"{selected_zone} Total", base_rows, selected_zone))
    return headers, rows


def get_zone_summary_data():
    init_bill_list_db()
    with get_db() as conn:
        rows = conn.execute(
            f"""
            SELECT
                zone,
                COUNT(*) AS total_bills,
                SUM(CASE WHEN amount_received > 0 THEN 1 ELSE 0 END) AS received_bills,
                SUM(COALESCE(total_bill, 0)) AS total_amount,
                SUM(COALESCE(arrears, 0)) AS total_arrears_received,
                SUM(COALESCE(amount_received, 0)) AS total_received_amount,
                SUM(CASE WHEN COALESCE(total_bill, 0) > COALESCE(amount_received, 0)
                    THEN COALESCE(total_bill, 0) - COALESCE(amount_received, 0) ELSE 0 END) AS pending_amount
            FROM bills
            GROUP BY zone
            ORDER BY {zone_sort_expr('zone')}
            """
        ).fetchall()
    return rows


def bill_list_staff_export_rows(staff_id=None, bill_ids: set[int] | None = None):
    init_bill_list_db()
    headers = [
        "Sr",
        "Staff",
        "Zone",
        "Sector",
        "Locality",
        "Total Bills",
        "Received Bills",
        "Remaining Bills",
        "Amount Received",
        "Pending Amount",
    ]
    staff_filter = ""
    params: list = []
    if staff_id:
        staff_filter = "WHERE sa.staff_id = ?"
        params.append(staff_id)

    bill_ids_clause = ""
    if bill_ids is not None:
        if not bill_ids:
            bill_ids_clause = "AND 1=0"
        else:
            id_list = ",".join(str(i) for i in bill_ids)
            bill_ids_clause = f"AND b.id IN ({id_list})"

    with get_db() as conn:
        has_auto_rules = conn.execute("SELECT COUNT(*) AS cnt FROM auto_assignment_rules").fetchone()["cnt"] > 0

        auto_exclude = ""
        if has_auto_rules:
            auto_exclude = "AND NOT EXISTS (SELECT 1 FROM auto_assignment_rules aar WHERE aar.sector = b.sector AND aar.locality = b.locality)"

        data = conn.execute(
            f"""
            SELECT
                sa.staff_id,
                s.name AS staff_name,
                sa.zone,
                sa.sector AS assigned_sector,
                sa.locality AS assigned_locality,
                COALESCE(sa.sector, 'All sectors') AS sector,
                COALESCE(sa.locality, '') AS locality,
                COUNT(DISTINCT b.id) AS total_bills,
                COUNT(DISTINCT CASE WHEN b.amount_received > 0 THEN b.id END) AS received_bills,
                SUM(COALESCE(b.amount_received, 0)) AS total_received_amount,
                SUM(CASE WHEN b.total_bill > b.amount_received THEN b.total_bill - b.amount_received ELSE 0 END) AS remaining_amount
            FROM staff_assignments sa
            JOIN staff s ON s.id = sa.staff_id
            LEFT JOIN bills b
                ON b.zone = sa.zone
                AND (sa.sector IS NULL OR b.sector = sa.sector)
                AND (sa.locality IS NULL OR b.locality = sa.locality)
                {auto_exclude}
                {bill_ids_clause}
            {staff_filter}
            GROUP BY sa.id, s.name, sa.zone, sa.sector, sa.locality
            ORDER BY s.name, sa.zone, sa.sector, sa.locality
            """,
            params,
        ).fetchall()
        data = list(data)
        expanded_data = []
        for row in data:
            if row["zone"] == "Commercial" and not row["assigned_locality"]:
                commercial_bill_ids_clause = ""
                if bill_ids is not None:
                    if not bill_ids:
                        commercial_bill_ids_clause = "AND 1=0"
                    else:
                        id_list = ",".join(str(i) for i in bill_ids)
                        commercial_bill_ids_clause = f"AND b.id IN ({id_list})"
                commercial_rows = conn.execute(
                    f"""
                    SELECT
                        ? AS staff_id,
                        ? AS staff_name,
                        b.zone,
                        b.sector,
                        b.locality,
                        COUNT(DISTINCT b.id) AS total_bills,
                        COUNT(DISTINCT CASE WHEN b.amount_received > 0 THEN b.id END) AS received_bills,
                        SUM(COALESCE(b.amount_received, 0)) AS total_received_amount,
                        SUM(CASE WHEN b.total_bill > b.amount_received THEN b.total_bill - b.amount_received ELSE 0 END) AS remaining_amount
                    FROM bills b
                    WHERE b.zone = ?
                    {auto_exclude if auto_exclude else ''}
                    {commercial_bill_ids_clause}
                    GROUP BY b.zone, b.sector, b.locality
                    ORDER BY b.sector, b.locality
                    """,
                    (
                        row["staff_id"],
                        row["staff_name"],
                        row["zone"],
                    ),
                ).fetchall()
                expanded_data.extend(commercial_rows or [row])
            else:
                expanded_data.append(row)
        data = expanded_data

        if has_auto_rules:
            auto_params: list = []
            auto_filter_sql = ""
            if staff_id:
                auto_filter_sql = "AND s.id = ?"
                auto_params.append(staff_id)
            auto_bill_ids_clause = ""
            if bill_ids is not None:
                if not bill_ids:
                    auto_bill_ids_clause = "AND 1=0"
                else:
                    id_list = ",".join(str(i) for i in bill_ids)
                    auto_bill_ids_clause = f"AND b.id IN ({id_list})"
            auto_rows = conn.execute(
                f"""
                SELECT
                    s.id AS staff_id,
                    s.name AS staff_name,
                    b.zone,
                    b.sector AS assigned_sector,
                    b.locality AS assigned_locality,
                    b.sector,
                    b.locality,
                    COUNT(DISTINCT b.id) AS total_bills,
                    COUNT(DISTINCT CASE WHEN b.amount_received > 0 THEN b.id END) AS received_bills,
                    SUM(COALESCE(b.amount_received, 0)) AS total_received_amount,
                    SUM(CASE WHEN b.total_bill > b.amount_received THEN b.total_bill - b.amount_received ELSE 0 END) AS remaining_amount
                FROM bills b
                JOIN auto_assignment_rules aar
                    ON aar.sector = b.sector AND aar.locality = b.locality
                    AND CAST(b.connection_no AS TEXT) >= aar.connection_min
                    AND (aar.connection_max IS NULL OR CAST(b.connection_no AS TEXT) <= aar.connection_max)
                JOIN staff s ON UPPER(TRIM(s.name)) = UPPER(TRIM(aar.staff_name))
                {auto_filter_sql}
                {auto_bill_ids_clause}
                GROUP BY s.id, b.zone, b.sector, b.locality
                ORDER BY s.name, b.zone, b.sector, b.locality
                """,
                auto_params,
            ).fetchall()
            data.extend(auto_rows)

        if not staff_id:
            unassigned_bill_ids_clause = ""
            if bill_ids is not None:
                if not bill_ids:
                    unassigned_bill_ids_clause = "AND 1=0"
                else:
                    id_list = ",".join(str(i) for i in bill_ids)
                    unassigned_bill_ids_clause = f"AND b.id IN ({id_list})"
            unassigned_data = conn.execute(
                f"""
                SELECT
                    'Unassigned' AS staff_name,
                    b.zone,
                    b.sector,
                    b.locality,
                    COUNT(b.id) AS total_bills,
                    SUM(CASE WHEN b.amount_received > 0 THEN 1 ELSE 0 END) AS received_bills,
                    SUM(COALESCE(b.amount_received, 0)) AS total_received_amount,
                    SUM(CASE WHEN b.total_bill > b.amount_received THEN b.total_bill - b.amount_received ELSE 0 END) AS remaining_amount
                FROM bills b
                WHERE NOT EXISTS (
                    SELECT 1
                    FROM staff_assignments sa
                    WHERE sa.zone = b.zone
                      AND (sa.sector IS NULL OR sa.sector = b.sector)
                      AND (sa.locality IS NULL OR sa.locality = b.locality)
                )
                {('AND NOT EXISTS (SELECT 1 FROM auto_assignment_rules aar WHERE aar.sector = b.sector AND aar.locality = b.locality)' if has_auto_rules else '')}
                {unassigned_bill_ids_clause}
                GROUP BY b.zone, b.sector, b.locality
                ORDER BY b.zone, b.sector, b.locality
                """
            ).fetchall()
            data = list(data) + list(unassigned_data)
    rows = []
    for idx, row in enumerate(data, start=1):
        total_bills = int(row["total_bills"] or 0)
        received_bills = int(row["received_bills"] or 0)
        total_received_amount = float(row["total_received_amount"] or 0)
        remaining_amount = float(row["remaining_amount"] or 0)
        if total_bills == 0 and received_bills == 0 and total_received_amount == 0 and remaining_amount == 0:
            continue
        rows.append(
            [
                idx,
                row["staff_name"],
                row["zone"],
                row["sector"],
                row["locality"],
                fmt(total_bills),
                fmt(received_bills),
                fmt(total_bills - received_bills),
                fmt(total_received_amount),
                fmt(remaining_amount),
            ]
        )
    return headers, rows


def export_table_response(fmt_type: str, title: str, headers: list[str], rows: list[list], filename: str):
    if fmt_type == "pdf":
        page_w = landscape(A4)[0] - 30 * mm
        col_widths = [page_w / len(headers)] * len(headers)
        if len(headers) == 8:
            col_widths = [
                page_w * 0.04,
                page_w * 0.09,
                page_w * 0.31,
                page_w * 0.08,
                page_w * 0.10,
                page_w * 0.10,
                page_w * 0.14,
                page_w * 0.14,
            ]
        elif len(headers) == 9:
            col_widths = [
                page_w * 0.04,
                page_w * 0.15,
                page_w * 0.08,
                page_w * 0.25,
                page_w * 0.08,
                page_w * 0.09,
                page_w * 0.09,
                page_w * 0.11,
                page_w * 0.11,
            ]
        pdf_bytes = generate_card_pdf(
            title,
            [f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}"],
            headers,
            rows,
            pagesize=landscape(A4),
            col_widths=col_widths,
            left_cols=[2] if len(headers) == 8 else [1, 3],
            header_font_size=8,
            body_font_size=8,
        )
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
        )
    if fmt_type == "csv":
        csv_data = pd.DataFrame(rows, columns=headers).to_csv(index=False)
        return Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )
    if fmt_type == "xlsx":
        buf = io.BytesIO()
        pd.DataFrame(rows, columns=headers).to_excel(buf, index=False)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )
    flash("Unknown export format.")
    return redirect(url_for("bill_list"))


def generate_staff_report_pdf(rows: list[list], show_summary: bool = True, cols_param: str = None, summary_cols_param: str = None) -> bytes:
    def numeric(value) -> float:
        return parse_number(str(value).replace(",", ""))

    # 10-col input: sr(0), staff(1), zone(2), sector(3), locality(4), totalBills(5), receivedBills(6), remainingBills(7), totalReceivedAmount(8), pendingAmount(9)
    # PDF 8-col: sr(0), sector(1), locality(2), totalBills(3), receivedBills(4), remainingBills(5), totalReceivedAmount(6), pendingAmount(7)
    # 10-col → PDF: {0:0, 3:1, 4:2, 5:3, 6:4, 7:5, 8:6, 9:7}
    _10TOPDF = {0: 0, 3: 1, 4: 2, 5: 3, 6: 4, 7: 5, 8: 6, 9: 7}
    _pdf_cols = list(range(8))
    _summary_cols = list(range(8))
    if cols_param:
        _sel = [k.strip() for k in cols_param.split(",") if k.strip() in STAFF_COL_MAP]
        _sel_set = set(_sel)
        _sel10 = set(STAFF_COL_MAP[k] for k in _sel)
        _pdf_cols = sorted([_10TOPDF[i] for i in _sel10 if i in _10TOPDF])
        # Summary cols: Sr, Staff Name, Zone, Total Assigned Bills, Total Received Bills, Remaining Bills, Amount Received, Pending Amount
        _sum_keys = ["sr", "staff", "zone", "totalBills", "receivedBills", "remainingBills", "totalReceivedAmount", "pendingAmount"]
        _summary_cols = [i for i, k in enumerate(_sum_keys) if k in _sel_set]

    def total_row(label: str, source_rows: list[list], staff_pdf_row: bool = True) -> list:
        total_bills = sum(numeric(row[5]) for row in source_rows)
        received_bills = sum(numeric(row[6]) for row in source_rows)
        remaining_bills = sum(numeric(row[7]) for row in source_rows)
        amount_received = sum(numeric(row[8]) for row in source_rows)
        pending_amount = sum(numeric(row[9]) for row in source_rows)
        if staff_pdf_row:
            return [
                "",
                label,
                "",
                fmt(total_bills),
                fmt(received_bills),
                fmt(remaining_bills),
                fmt(amount_received),
                fmt(pending_amount),
            ]
        return [
            label,
            fmt(total_bills),
            fmt(received_bills),
            fmt(remaining_bills),
            fmt(amount_received),
            fmt(pending_amount),
        ]

    buf = io.BytesIO()
    left_margin = 8 * mm
    right_margin = 8 * mm
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        topMargin=6 * mm,
        bottomMargin=6 * mm,
        leftMargin=6 * mm,
        rightMargin=6 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PDFTitle",
        parent=styles["Heading1"],
        fontSize=20,
        textColor=ACCENT,
        alignment=1,
        spaceAfter=6 * mm,
        fontName="Helvetica-Bold",
    )
    summary_style = ParagraphStyle(
        "PDFSummary",
        parent=styles["Normal"],
        fontSize=14,
        alignment=0,
        spaceAfter=2 * mm,
        textColor=colors.HexColor("#333333"),
        leading=18,
    )
    group_style = ParagraphStyle(
        "PDFGroup",
        parent=styles["Heading3"],
        fontSize=10,
        textColor=colors.HexColor("#222222"),
        alignment=0,
        spaceBefore=1 * mm,
        spaceAfter=0.5 * mm,
        fontName="Helvetica-Bold",
    )

    page_w = landscape(A4)[0] - left_margin - right_margin

    # ── Parse summary column selection ──
    _sum_sel = []
    if summary_cols_param:
        _sum_sel = [k.strip() for k in summary_cols_param.split(",") if k.strip() in STAFF_SUMMARY_COL_MAP]

    _sum_all_headers = [
        "Sr",
        "Staff Name",
        "Total Bills",
        "Received Bills",
        "Remaining Bills",
        "Total Amount",
        "Amount Received",
        "Pending Amount",
    ]
    _sum_all_widths = [
        page_w * 0.04,
        page_w * 0.20,
        page_w * 0.09,
        page_w * 0.09,
        page_w * 0.09,
        page_w * 0.13,
        page_w * 0.13,
        page_w * 0.13,
    ]
    _sum_pdf_cols = sorted([STAFF_SUMMARY_COL_MAP[k] for k in _sum_sel]) if _sum_sel else list(range(8))
    sum_headers = [_sum_all_headers[i] for i in _sum_pdf_cols]
    sum_widths = [_sum_all_widths[i] for i in _sum_pdf_cols]
    # left-align staff name column (index 1 in full 8-col)
    _sum_left_cols = {_sum_pdf_cols.index(1)} if 1 in _sum_pdf_cols else set()

    def wrap_sum_text(value):
        return wrap_pdf_body_cells([[value]], font_size=9)[0][0]

    def wrap_sum_left(value):
        return wrap_pdf_body_cells([[value]], font_size=9, left_columns={0})[0][0]

    grouped = {}
    for row in rows:
        grouped.setdefault((row[1], row[2]), []).append(row)

    zone_order = {"A": 1, "B": 2, "C": 3, "Commercial": 4, "Unassigned": 5}
    grouped_items = sorted(
        grouped.items(),
        key=lambda item: (
            zone_order.get(item[0][1], 99),
            -len(item[1]),
            item[0][0].lower(),
        ),
    )

    summary_data = []
    grand_total_all = {"total_bills": 0, "received_bills": 0, "remaining_bills": 0, "amount_received": 0, "pending_amount": 0}

    for (staff_name, zone), group_rows in grouped_items:
        total_bills = sum(numeric(row[5]) for row in group_rows)
        received_bills = sum(numeric(row[6]) for row in group_rows)
        remaining_bills = sum(numeric(row[7]) for row in group_rows)
        amount_received = sum(numeric(row[8]) for row in group_rows)
        pending_amount = sum(numeric(row[9]) for row in group_rows)

        summary_data.append({
            "staff_name": staff_name,
            "zone": zone,
            "total_bills": total_bills,
            "received_bills": received_bills,
            "remaining_bills": remaining_bills,
            "amount_received": amount_received,
            "pending_amount": pending_amount,
        })

        grand_total_all["total_bills"] += total_bills
        grand_total_all["received_bills"] += received_bills
        grand_total_all["remaining_bills"] += remaining_bills
        grand_total_all["amount_received"] += amount_received
        grand_total_all["pending_amount"] += pending_amount

    # ── Build elements list ──
    elements = []

    # Page 1: Staff-wise Summary Report
    if summary_data:
        _sum_full_headers = _sum_all_headers
        _sum_full_widths = _sum_all_widths
        _sum_sr_headers = [_sum_full_headers[i] for i in _sum_pdf_cols]
        _sum_sr_widths = [_sum_full_widths[i] for i in _sum_pdf_cols]
        _sum_body_rows = []
        _sum_total_vals = [0] * 8
        for idx, sd in enumerate(summary_data, start=1):
            total_amount = sd["amount_received"] + sd["pending_amount"]
            _sr = [
                idx,
                wrap_sum_left(fmt_staff_name(sd["staff_name"])),
                fmt(sd["total_bills"]),
                fmt(sd["received_bills"]),
                fmt(sd["remaining_bills"]),
                fmt(total_amount),
                fmt(sd["amount_received"]),
                fmt(sd["pending_amount"]),
            ]
            _sum_body_rows.append([_sr[i] for i in _sum_pdf_cols])
            _sum_total_vals[2] += sd["total_bills"]
            _sum_total_vals[3] += sd["received_bills"]
            _sum_total_vals[4] += sd["remaining_bills"]
            _sum_total_vals[5] += sd["amount_received"] + sd["pending_amount"]
            _sum_total_vals[6] += sd["amount_received"]
            _sum_total_vals[7] += sd["pending_amount"]
        # Grand total row
        gt_sum = [
            "",
            "Grand Total",
            fmt(_sum_total_vals[2]),
            fmt(_sum_total_vals[3]),
            fmt(_sum_total_vals[4]),
            fmt(_sum_total_vals[5]),
            fmt(_sum_total_vals[6]),
            fmt(_sum_total_vals[7]),
        ]
        _sum_body_rows.append([gt_sum[i] for i in _sum_pdf_cols])
        _sum_table_data = [_sum_sr_headers] + _sum_body_rows
        elements.append(Paragraph("Staff-wise Summary Report", title_style))
        if show_summary:
            elements.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}", summary_style))
        elements.append(Spacer(1, 4 * mm))
        elements.append(
            _make_pdf_table(
                _sum_table_data,
                col_widths=_sum_sr_widths,
                left_cols=_sum_left_cols,
                header_font_size=10,
                body_font_size=9,
                cell_padding=6,
            )
        )
        elements.append(PageBreak())

    # Page 2+: Detail pages
    elements.append(Paragraph("Bill List - Staff Report", title_style))
    if show_summary:
        elements.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}", summary_style))

    _all_table_headers = [
        "Sr",
        "Sector",
        "Locality",
        "Total Bills",
        "Received Bills",
        "Remaining Bills",
        "Amount Received",
        "Pending Amount",
    ]
    table_headers = [_all_table_headers[i] for i in _pdf_cols]
    _all_col_widths = [
        page_w * 0.04,
        page_w * 0.24,
        page_w * 0.22,
        page_w * 0.08,
        page_w * 0.10,
        page_w * 0.10,
        page_w * 0.11,
        page_w * 0.11,
    ]
    col_widths = [_all_col_widths[i] for i in _pdf_cols]

    def wrap_text(value):
        return wrap_pdf_body_cells([[value]], font_size=8)[0][0]

    def wrap_text_left(value):
        return wrap_pdf_body_cells([[value]], font_size=8, left_columns={0})[0][0]

    for (staff_name, zone), group_rows in grouped_items:
        staff_elements = [
            Paragraph(f"Staff: {fmt_staff_name(staff_name).replace(chr(10), '<br/>')}", group_style),
            Paragraph(f"Zone: {zone}", group_style),
        ]
        table_rows = []
        for idx, row in enumerate(group_rows, start=1):
            _full_row = [idx, wrap_text_left(row[3]), wrap_text(row[4]), row[5], row[6], row[7], row[8], row[9]]
            table_rows.append([_full_row[i] for i in _pdf_cols])
        data_rows = [table_headers] + table_rows + [total_row("Grand Total", group_rows)]
        staff_elements.append(
            _make_pdf_table(
                data_rows,
                col_widths=col_widths,
                left_cols=[],
                header_font_size=8,
                body_font_size=8,
                cell_padding=5,
            )
        )
        staff_elements.append(Spacer(1, 5 * mm))
        elements.append(KeepTogether(staff_elements))

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


def export_zone_report_response(fmt_type: str, selected_zone: str, show_summary: bool = False):
    allowed_zones = {"All", "A", "B", "C", "Commercial", "Unassigned"}
    if selected_zone not in allowed_zones:
        selected_zone = "All"
    headers, rows = bill_list_zone_export_rows(selected_zone)
    cols_param = request.args.get("cols")
    title = "Zone-wise Report" if selected_zone == "All" else f"Zone {selected_zone} Report"
    filename_zone = selected_zone.lower().replace(" ", "_")

    def without_zone(row: list) -> list:
        return [row[0], row[2], row[3], row[4], row[5], row[6], row[7]]

    if fmt_type != "pdf":
        # CSV/Excel: filter full 8-column data (with Zone)
        csv_headers, csv_rows = parse_export_cols(cols_param, ZONE_COL_MAP, headers, rows)
        return export_table_response(fmt_type, title, csv_headers, csv_rows, f"bill_list_zone_{filename_zone}_report")

    # PDF: filter using PDF column map (7 cols, no Zone)
    pdf_col_headers = ["Sr", "Sector", "Total Bills", "Received Bills", "Remaining Bills", "Amount Received", "Pending Amount"]
    # Build a PDF-specific column map based on which of the original 8 cols are selected
    zone_selected = True
    if cols_param:
        selected_keys = [k.strip() for k in cols_param.split(",") if k.strip()]
        zone_selected = "zone" in selected_keys
    # If full 8-col set is used, strip zone; otherwise use filtered subset
    if cols_param:
        # Map original 8-col indices to 7-col PDF indices (removing zone at index 1)
        orig_sel = [k.strip() for k in cols_param.split(",") if k.strip() in ZONE_COL_MAP]
        pdf_indices = []
        for key in orig_sel:
            orig_idx = ZONE_COL_MAP[key]
            if orig_idx == 1:
                continue  # skip Zone
            pdf_idx = orig_idx if orig_idx < 1 else orig_idx - 1
            pdf_indices.append(pdf_idx)
        pdf_headers = [pdf_col_headers[i] for i in pdf_indices]
    else:
        pdf_headers = pdf_col_headers
        pdf_indices = list(range(7))

    def transform_row(row):
        r7 = without_zone(row)
        return [r7[i] for i in pdf_indices]

    grouped_sections = []
    overall_total = None
    if selected_zone == "All":
        current_zone = None
        current_rows = []
        current_total = None
        for row in rows:
            if row[2] == "Grand Total":
                overall_total = transform_row(row)
                continue
            if str(row[2]).endswith(" Total"):
                current_total = transform_row(row)
                if current_zone is not None:
                    grouped_sections.append({"zone": f"Zone {current_zone}", "rows": current_rows, "total": current_total if show_summary else None})
                current_zone = None
                current_rows = []
                current_total = None
                continue
            if current_zone is None:
                current_zone = row[1]
            current_rows.append(transform_row(row))
        if current_zone is not None:
            grouped_sections.append({"zone": f"Zone {current_zone}", "rows": current_rows, "total": current_total if show_summary else None})
    else:
        section_rows = []
        section_total = None
        for row in rows:
            if str(row[2]).endswith(" Total"):
                section_total = transform_row(row)
            else:
                section_rows.append(transform_row(row))
        grouped_sections.append({"zone": f"Zone {selected_zone}", "rows": section_rows, "total": section_total if show_summary else None})

    zone_summary_data = get_zone_summary_data() if show_summary and selected_zone == "All" else None
    # Zone summary table column mapping (table headers include Sr auto-generated)
    # 8 header cols: Sr, Zone, Total Bills, Received Bills, Total Amount, Arrears Received, Total Received, Pending Amount
    _zs_key_order = ["sr", "zone", "totalBills", "receivedBills", "totalAmount", "arrearsReceived", "totalReceivedAmount", "pendingAmount"]
    _zs_sel_indices = list(range(len(_zs_key_order)))  # all by default
    if zone_summary_data and cols_param:
        _zs_sel = [k.strip() for k in cols_param.split(",") if k.strip() in _zs_key_order]
        _zs_sel_indices = [i for i, k in enumerate(_zs_key_order) if k in _zs_sel]
    # We store the selection indices for use in generate_zone_grouped_pdf
    _zs_sel_str = ",".join(str(i) for i in _zs_sel_indices) if _zs_sel_indices != list(range(len(_zs_key_order))) else ""

    summary_lines = [
        f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}",
        f"<b>Zone:</b> {selected_zone}",
    ] if show_summary else []

    pdf_bytes = generate_zone_grouped_pdf(
        title,
        summary_lines,
        grouped_sections,
        overall_total,
        zone_summary_data,
        pdf_detail_headers=pdf_headers,
        zone_summary_sel=_zs_sel_indices if _zs_sel_indices != list(range(8)) else None,
    )
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=bill_list_zone_{filename_zone}_report.pdf"},
    )


def _save_results_cache(results: dict) -> None:
    try:
        with open(RESULTS_CACHE, "w", encoding="utf-8") as handle:
            json.dump(results, handle, ensure_ascii=True, default=str)
    except OSError:
        pass


def _load_results_cache() -> dict:
    if not os.path.exists(RESULTS_CACHE):
        return {}
    try:
        with open(RESULTS_CACHE, "r", encoding="utf-8") as handle:
            return json.load(handle)
    except (OSError, json.JSONDecodeError):
        return {}


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

# We store dashboard uploads in memory only. Bill List keeps its own SQLite
# data; this first-page data intentionally disappears when the process restarts.
_last_results = {}
_last_merged_df = None
_last_uploaded_names = None
_last_daily_staff_results = {}
_last_daily_staff_uploaded_names = []


def build_dashboard_results(merged: pd.DataFrame) -> dict:
    """Build the All Received Bills dashboard once and reuse it for the rendered page and exports."""
    results = summarize_dataframe(merged)
    results["raw_row_count"] = int(len(merged))
    results["duplicate_rows_removed"] = 0
    results["duplicate_key_columns"] = []
    results["merged_csv_name"] = "merged_latest.csv"
    results["merged_xlsx_name"] = "merged_latest.xlsx"
    return results


def read_and_merge_uploaded_files(files) -> tuple[pd.DataFrame | None, list[str]]:
    dataframes = []
    base_columns = None
    uploaded_names = []

    for file in files:
        if not file or not file.filename:
            continue
        if not allowed_file(file.filename):
            flash(f"Unsupported file type: {file.filename}")
            continue

        filename = secure_filename(file.filename)
        try:
            df = read_uploaded_dataframe(file)
            if base_columns is None:
                base_columns = list(df.columns)
            else:
                df = df.reindex(columns=base_columns)
            dataframes.append(df)
            uploaded_names.append(filename)
        except Exception as exc:
            flash(f"Failed to read {filename}: {exc}")

    if not dataframes:
        return None, uploaded_names
    return pd.concat(dataframes, ignore_index=True, sort=False), uploaded_names


@app.route("/", methods=["GET", "POST"])
def index():
    global _last_results, _last_merged_df, _last_uploaded_names
    if request.method == "POST":
        action = request.form.get("action", "")

        if action == "save_data":
            files = request.files.getlist("files")
            if not files or all(not f.filename for f in files):
                if is_ajax():
                    return ajax_error("Please choose a CSV or Excel file first.")
                flash("Please choose a CSV or Excel file first.")
                return redirect(url_for("index"))

            merged, uploaded_names = read_and_merge_uploaded_files(files)
            if merged is None:
                if is_ajax():
                    return ajax_error("No valid files were processed.")
                flash("No valid files were processed.")
                return redirect(url_for("index"))

            saved_meta = {
                "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "file_names": uploaded_names,
                "row_count": len(merged),
            }

            # Persist to disk so data survives page refresh.
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            merged.to_csv(SAVED_DASHBOARD_CSV, index=False)
            with open(SAVED_DASHBOARD_META, "w") as f:
                json.dump(saved_meta, f)

            # Heavy uploads must not stop at the saved-data banner. Generate the
            # dashboard from the saved dataframe immediately so the page shows
            # report data after upload instead of the confusing empty state.
            results = build_dashboard_results(merged)

            _last_merged_df = merged.copy()
            _last_uploaded_names = uploaded_names
            _last_results = results
            _save_results_cache(results)
            if is_ajax():
                return ajax_ok(
                    message=f"Reports generated successfully. {len(merged):,} rows from {len(uploaded_names)} file(s).",
                    redirect_url=url_for("index"),
                )
            flash("Reports generated successfully.")
            return render_template("index.html", results=results, uploaded_names=uploaded_names,
                                   saved_meta=saved_meta)

        if action == "generate_reports":
            merged = _last_merged_df
            if merged is None and os.path.exists(SAVED_DASHBOARD_CSV):
                merged = pd.read_csv(SAVED_DASHBOARD_CSV)
                if merged is not None:
                    _last_merged_df = merged.copy()
                    if os.path.exists(SAVED_DASHBOARD_META):
                        with open(SAVED_DASHBOARD_META) as f:
                            meta = json.load(f)
                        _last_uploaded_names = meta.get("file_names")
            if merged is None:
                flash("Please save data first before generating reports.")
                return redirect(url_for("index"))

            results = build_dashboard_results(merged)

            _last_results = results
            _save_results_cache(results)
            saved_meta = None
            if os.path.exists(SAVED_DASHBOARD_META):
                with open(SAVED_DASHBOARD_META) as f:
                    saved_meta = json.load(f)
            flash("Reports generated successfully.")
            return render_template("index.html", results=results,
                                   uploaded_names=_last_uploaded_names,
                                   saved_meta=saved_meta)

        if action == "clear_saved_data":
            for path in [SAVED_DASHBOARD_CSV, SAVED_DASHBOARD_META, RESULTS_CACHE]:
                if os.path.exists(path):
                    os.remove(path)
            _last_merged_df = None
            _last_results = {}
            _last_uploaded_names = None
            flash("Saved data cleared.")
            return redirect(url_for("index"))

        # Fallback — treat as old-style process-files for backward compat
        files = request.files.getlist("files")
        if not files or all(not f.filename for f in files):
            flash("Please choose at least one CSV or Excel file.")
            return redirect(url_for("index"))

        merged, uploaded_names = read_and_merge_uploaded_files(files)
        if merged is None:
            flash("No valid files were processed.")
            return redirect(url_for("index"))

        results = build_dashboard_results(merged)

        _last_results = results
        _last_merged_df = merged.copy()
        _last_uploaded_names = uploaded_names
        _save_results_cache(results)
        return render_template("index.html", results=results, uploaded_names=uploaded_names)

    # GET — restore metadata from persistent storage and pass to template
    # A saved heavy upload can already have generated results cached; keep those
    # visible on refresh instead of showing a blank "no data" dashboard.
    _last_results = _last_results or _load_results_cache()
    _last_merged_df = None
    _last_uploaded_names = None
    saved_meta = None
    if os.path.exists(SAVED_DASHBOARD_META):
        with open(SAVED_DASHBOARD_META) as f:
            saved_meta = json.load(f)
        _last_uploaded_names = saved_meta.get("file_names")
    return render_template("index.html", results=(_last_results or None), uploaded_names=_last_uploaded_names,
                           saved_meta=saved_meta)


@app.route("/download/<path:filename>")
def download_file(filename: str):
    if filename == "merged_latest.csv" and _last_merged_df is not None:
        return Response(
            _last_merged_df.to_csv(index=False),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=merged_latest.csv"},
        )
    if filename == "merged_latest.xlsx" and _last_merged_df is not None:
        buf = io.BytesIO()
        _last_merged_df.to_excel(buf, index=False)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=merged_latest.xlsx"},
        )
    flash("No uploaded dashboard data is available. Please upload the file again.")
    return redirect(url_for("index"))


@app.route("/daily-staff-receive", methods=["GET", "POST"])
def daily_staff_receive():
    global _last_daily_staff_results, _last_daily_staff_uploaded_names
    if request.method == "POST":
        files = request.files.getlist("files")
        if not files or all(not f.filename for f in files):
            if is_ajax():
                return ajax_error("Please choose at least one CSV or Excel file.")
            flash("Please choose at least one CSV or Excel file.")
            return redirect(url_for("daily_staff_receive"))

        merged, uploaded_names = read_and_merge_uploaded_files(files)
        if merged is None:
            if is_ajax():
                return ajax_error("No valid files were processed.")
            flash("No valid files were processed.")
            return redirect(url_for("daily_staff_receive"))

        results = summarize_dataframe(merged)
        _last_daily_staff_results = {"daily_staff_receive": results.get("daily_staff_receive") or {}}
        _last_daily_staff_uploaded_names = uploaded_names
        if is_ajax():
            staff_count = len((_last_daily_staff_results["daily_staff_receive"] or {}).get("summary_rows") or [])
            return ajax_ok(
                message=f"Report generated. {staff_count} staff member(s) found.",
                redirect_url=url_for("daily_staff_receive"),
            )
        return render_template(
            "daily_staff_receive.html",
            report=_last_daily_staff_results["daily_staff_receive"],
            uploaded_names=uploaded_names,
        )

    sort_order = request.args.get("sort", "default")
    if sort_order not in ("default", "asc", "desc"):
        sort_order = "default"
    return render_template(
        "daily_staff_receive.html",
        report=(_last_daily_staff_results or {}).get("daily_staff_receive"),
        uploaded_names=_last_daily_staff_uploaded_names,
        sort_order=sort_order,
    )


@app.route("/bill-list", methods=["GET", "POST"])
def bill_list():
    init_bill_list_db()
    if request.method == "POST":
        action = request.form.get("action", "")
        if not action and request.files.get("bill_file"):
            action = "save_data"
        if action == "clear_bill_list_data":
            clear_bill_list_data()
            flash("Bill rows and saved upload copies have been removed. Staff, assignments, sectors, localities, and zones were kept.")
            return redirect(url_for("bill_list"))

        if action == "save_data":
            file = request.files.get("bill_file")
            if not file or not file.filename:
                if is_ajax():
                    return ajax_error("Please choose a bill file first.")
                flash("Please choose a bill file first.")
                return redirect(url_for("bill_list"))
            if not allowed_file(file.filename):
                if is_ajax():
                    return ajax_error(f"Unsupported file type: {file.filename}")
                flash(f"Unsupported file type: {file.filename}")
                return redirect(url_for("bill_list"))

            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
            save_path = os.path.join(app.config["UPLOAD_FOLDER"], f"bill_list_{timestamp}_{filename}")
            file.save(save_path)
            try:
                df = read_dataframe(save_path)
                imported, duplicates = import_bill_list_dataframe(df)
                msg = f"Bill data saved. Imported {imported:,} row(s)."
                if duplicates:
                    msg += f" Skipped {duplicates:,} duplicate(s)."
                if is_ajax():
                    return ajax_ok(message=msg, redirect_url=url_for("bill_list"))
                flash(msg)
            except Exception as exc:
                if is_ajax():
                    return ajax_error(f"Failed to import file: {exc}")
                flash(f"Failed to import Bill List file: {exc}")
            return redirect(url_for("bill_list"))

        if action == "generate_reports":
            with get_db() as conn:
                count = conn.execute("SELECT COUNT(*) as cnt FROM bills").fetchone()["cnt"]
                if count == 0:
                    flash("Please save or upload bill data first.")
                    return redirect(url_for("bill_list"))
            flash("Reports generated successfully.")
            return redirect(url_for("bill_list"))

        if action == "add_staff":
            name = (request.form.get("staff_name") or "").strip()
            if not name:
                flash("Enter a staff name.")
                return redirect(url_for("bill_list"))
            try:
                with get_db() as conn:
                    conn.execute(
                        "INSERT INTO staff (name, created_at) VALUES (?, ?)",
                        (name, datetime.now().isoformat(timespec="seconds")),
                    )
                flash(f"Added staff member: {name}")
            except sqlite3.IntegrityError:
                flash(f"Staff member already exists: {name}")
            return redirect(url_for("bill_list"))

        if action == "update_staff":
            staff_id = request.form.get("staff_id")
            name = (request.form.get("staff_name") or "").strip()
            if not staff_id or not name:
                flash("Choose a staff member and enter a new name.")
                return redirect(url_for("bill_list"))
            try:
                with get_db() as conn:
                    conn.execute("UPDATE staff SET name = ? WHERE id = ?", (name, staff_id))
                flash(f"Updated staff member: {name}")
            except sqlite3.IntegrityError:
                flash(f"Staff member already exists: {name}")
            return redirect(url_for("bill_list"))

        if action == "delete_staff":
            staff_id = request.form.get("staff_id")
            if staff_id:
                with get_db() as conn:
                    row = conn.execute("SELECT name FROM staff WHERE id = ?", (staff_id,)).fetchone()
                    conn.execute("DELETE FROM staff_assignments WHERE staff_id = ?", (staff_id,))
                    conn.execute("DELETE FROM staff WHERE id = ?", (staff_id,))
                flash(f"Deleted staff member: {row['name'] if row else staff_id}")
            return redirect(url_for("bill_list"))

        if action == "assign_staff":
            staff_id = request.form.get("staff_id")
            zone = (request.form.get("zone") or "").strip()
            selected_sectors = split_sector_values(request.form.getlist("sectors"))
            selected_localities = []
            for value in request.form.getlist("localities"):
                if "||" not in value:
                    continue
                sector_value, locality_value = value.split("||", 1)
                sector_value = sector_value.strip()
                locality_value = locality_value.strip()
                if sector_value and locality_value:
                    selected_localities.append((sector_value, locality_value))
            if not staff_id or not zone:
                flash("Choose a staff member and zone.")
                return redirect(url_for("bill_list"))
            effective_localities = [] if selected_sectors else selected_localities
            try:
                with get_db() as conn:
                    conflicts = get_assignment_conflicts(conn, staff_id, zone, selected_sectors, effective_localities)
                    if conflicts:
                        flash("Cannot save assignment. " + "; ".join(conflicts[:5]))
                        return redirect(url_for("bill_list"))
                    if effective_localities:
                        conn.execute(
                            "DELETE FROM staff_assignments WHERE staff_id = ? AND zone = ?",
                            (staff_id, zone),
                        )
                        for sector, locality in effective_localities:
                            conn.execute(
                                """
                                INSERT INTO staff_assignments (staff_id, zone, sector, locality, created_at)
                                VALUES (?, ?, ?, ?, ?)
                                """,
                                (staff_id, zone, sector, locality, datetime.now().isoformat(timespec="seconds")),
                            )
                    elif selected_sectors:
                        conn.execute(
                            "DELETE FROM staff_assignments WHERE staff_id = ? AND zone = ?",
                            (staff_id, zone),
                        )
                        for sector in selected_sectors:
                            conn.execute(
                                """
                                INSERT INTO staff_assignments (staff_id, zone, sector, locality, created_at)
                                VALUES (?, ?, ?, ?, ?)
                                """,
                                (staff_id, zone, sector, None, datetime.now().isoformat(timespec="seconds")),
                            )
                    else:
                        available_sectors = conn.execute(
                            """
                            SELECT name
                            FROM sectors
                            WHERE zone = ?
                              AND NOT EXISTS (
                                  SELECT 1
                                  FROM staff_assignments sa
                                  WHERE sa.zone = sectors.zone
                                    AND (sa.sector = sectors.name OR sa.sector IS NULL)
                              )
                            ORDER BY name
                            """,
                            (zone,),
                        ).fetchall()
                        if not available_sectors:
                            flash("No available sectors remain in that zone.")
                            return redirect(url_for("bill_list"))
                        for row in available_sectors:
                            conn.execute(
                                """
                                INSERT INTO staff_assignments (staff_id, zone, sector, locality, created_at)
                                VALUES (?, ?, ?, ?, ?)
                                """,
                                (staff_id, zone, row["name"], None, datetime.now().isoformat(timespec="seconds")),
                            )
                flash("Staff assignment saved.")
            except sqlite3.IntegrityError:
                flash("One or more selected staff assignments already exist.")
            return redirect(url_for("bill_list"))

        if action == "delete_assignment":
            assignment_id = request.form.get("assignment_id")
            if assignment_id:
                with get_db() as conn:
                    conn.execute("DELETE FROM staff_assignments WHERE id = ?", (assignment_id,))
                flash("Staff assignment removed.")
            return redirect(url_for("bill_list"))

        if action == "update_sector_zone":
            selected_sectors = split_sector_values(request.form.getlist("sectors"))
            fallback_sector = (request.form.get("sector") or "").strip()
            if fallback_sector:
                selected_sectors.append(fallback_sector)
            zone = (request.form.get("zone") or "").strip()
            selected_sectors = sorted(set(selected_sectors))
            if not selected_sectors or not zone:
                flash("Choose one or more sectors and a zone to update.")
                return redirect(url_for("bill_list"))
            with get_db() as conn:
                for sector in selected_sectors:
                    update_sector_zone(conn, sector, zone)
            flash(f"Updated {len(selected_sectors):,} sector(s) to Zone {zone}.")
            return redirect(url_for("bill_list"))

    return render_template("bill_list.html", **get_bill_list_context())


@app.route("/bill-list/export/<fmt_type>")
def export_bill_list(fmt_type: str):
    headers, rows, summary = bill_list_export_rows()
    rows = _filter_rows_by_selection(rows, lambda row: str(row[1]))
    if _selection_has_filter():
        summary = _bill_list_summary_from_rows(rows)
    cols_param = request.args.get("cols")
    headers, rows = parse_export_cols(cols_param, SECTOR_COL_MAP, headers, rows)
    if fmt_type == "pdf":
        show_summary = request.args.get("show_summary", "0") == "1"
        n = len(headers)
        page_w = landscape(A4)[0] - 30 * mm
        col_widths = [page_w / n] * n
        summary_lines = [
            f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}",
            f"<b>Total Bills:</b> {fmt(summary['total_bills'])}",
            f"<b>Received Bills:</b> {fmt(summary['received_bills'])}",
            f"<b>Remaining Bills:</b> {fmt(summary['remaining_bills'])}",
            f"<b>Amount Received:</b> Rs. {fmt(summary['total_received_amount'])}",
            f"<b>Pending Amount:</b> Rs. {fmt(summary['remaining_amount'])}",
        ] if show_summary else []
        pdf_bytes = generate_card_pdf(
            "Bill List - Sector-wise Report",
            summary_lines,
            headers,
            rows,
            pagesize=landscape(A4),
            col_widths=col_widths,
            left_cols=[i for i, h in enumerate(headers) if h.lower() == "sector"],
            header_font_size=9,
            body_font_size=9,
        )
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": "attachment; filename=bill_list_sector_report.pdf"},
        )
    if fmt_type == "csv":
        csv_data = pd.DataFrame(rows, columns=headers).to_csv(index=False)
        return Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=bill_list_sector_report.csv"},
        )
    if fmt_type == "xlsx":
        buf = io.BytesIO()
        pd.DataFrame(rows, columns=headers).to_excel(buf, index=False)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=bill_list_sector_report.xlsx"},
        )
    flash("Unknown export format.")
    return redirect(url_for("bill_list"))


@app.route("/bill-list/export/unpaid-amounts/<fmt_type>")
def export_unpaid_amount_summary(fmt_type: str):
    sections, total = unpaid_amount_export_data()
    cols_param = request.args.get("cols")
    filename = "unpaid_expired_bill_amounts"

    # Filter columns for each section
    if cols_param:
        filtered_sections = []
        for key, title, headers, rows in sections:
            if key == "summary":
                col_map = UNPAID_SUMMARY_COL_MAP
            else:
                col_map = UNPAID_SECTION_COL_MAP
            fh, fr = parse_export_cols(cols_param, col_map, headers, rows)
            filtered_sections.append((key, title, fh, fr))
        sections = filtered_sections

    if fmt_type == "pdf":
        show_summary = request.args.get("show_summary", "0") == "1"
        pdf_bytes = generate_unpaid_amount_pdf(sections, total, show_summary=show_summary)
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
        )

    if fmt_type == "csv":
        out = io.StringIO()
        writer = csv.writer(out)
        for _, section_title, headers, rows in sections:
            writer.writerow([section_title])
            writer.writerow(headers)
            writer.writerows(rows)
            writer.writerow([])
        return Response(
            out.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    if fmt_type == "xlsx":
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            for _, section_title, headers, rows in sections:
                pd.DataFrame(rows, columns=headers).to_excel(
                    writer,
                    sheet_name=section_title[:31],
                    index=False,
                )
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )

    flash("Unknown export format.")
    return redirect(url_for("bill_list"))


@app.route("/bill-list/export/unpaid-amounts/<section_key>/<fmt_type>")
def export_unpaid_amount_section(section_key: str, fmt_type: str):
    sections, total = unpaid_amount_export_data()
    section = unpaid_amount_section(sections, section_key)
    if section is None:
        flash("Unknown unpaid amount report section.")
        return redirect(url_for("bill_list"))

    key, title, headers, rows = section
    rows = _filter_rows_by_selection(rows, lambda row: str(row[1]))
    if key != "summary" and _selection_has_filter():
        total = _unpaid_total_from_section_rows(rows)
    cols_param = request.args.get("cols")
    if key == "summary":
        col_map = UNPAID_SUMMARY_COL_MAP
    else:
        col_map = UNPAID_SECTION_COL_MAP
    headers, rows = parse_export_cols(cols_param, col_map, headers, rows)
    section = (key, title, headers, rows)
    filename = f"unpaid_expired_{key}_amounts"
    selected_sections = [section]

    if fmt_type == "pdf":
        show_summary = request.args.get("show_summary", "0") == "1"
        pdf_bytes = generate_unpaid_amount_pdf(selected_sections, total, show_summary=show_summary)
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
        )

    if fmt_type == "csv":
        csv_data = pd.DataFrame(rows, columns=headers).to_csv(index=False)
        return Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    if fmt_type == "xlsx":
        buf = io.BytesIO()
        pd.DataFrame(rows, columns=headers).to_excel(buf, sheet_name=title[:31], index=False)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )

    flash("Unknown export format.")
    return redirect(url_for("bill_list"))


@app.route("/bill-list/export/zone/<fmt_type>")
def export_bill_list_zone(fmt_type: str):
    show_summary = request.args.get("show_summary", "0") == "1"
    return export_zone_report_response(fmt_type, request.args.get("zone", "All"), show_summary=show_summary)


@app.route("/bill-list/export/staff/<fmt_type>")
def export_bill_list_staff(fmt_type: str):
    staff_id = request.args.get("staff_id") or None
    show_summary = request.args.get("show_summary", "0") == "1"
    cols_param = request.args.get("cols")
    summary_cols_param = request.args.get("summary_cols")
    headers, rows = bill_list_staff_export_rows(staff_id)
    # Staff-wise exports must use only the rows currently checked in the page table.
    rows = _filter_rows_by_selection(rows, _staff_row_key)
    if fmt_type == "pdf":
        pdf_bytes = generate_staff_report_pdf(rows, show_summary=show_summary, cols_param=cols_param, summary_cols_param=summary_cols_param)
        suffix = f"_{staff_id}" if staff_id else ""
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=bill_list_staff_report{suffix}.pdf"},
        )

    if staff_id:
        fh, fr = parse_export_cols(cols_param, STAFF_COL_MAP, headers, rows)
        return export_table_response(fmt_type, "Bill List - Staff Report", fh, fr, f"bill_list_staff_report_{staff_id}")

    def parse_num(val):
        return parse_number(str(val).replace(",", ""))

    grouped = {}
    for row in rows:
        grouped.setdefault((row[1], row[2]), []).append(row)

    summary_rows = []
    grand_total = {"total_bills": 0, "received_bills": 0, "remaining_bills": 0, "amount_received": 0, "pending_amount": 0}

    for idx, ((staff_name, zone), group_rows) in enumerate(sorted(grouped.items()), start=1):
        total_bills = sum(parse_num(r[5]) for r in group_rows)
        received_bills = sum(parse_num(r[6]) for r in group_rows)
        remaining_bills = sum(parse_num(r[7]) for r in group_rows)
        amount_received = sum(parse_num(r[8]) for r in group_rows)
        pending_amount = sum(parse_num(r[9]) for r in group_rows)
        total_amount = amount_received + pending_amount

        summary_rows.append([
            idx,
            fmt_staff_name(staff_name),
            total_bills,
            received_bills,
            remaining_bills,
            total_amount,
            amount_received,
            pending_amount,
        ])

        grand_total["total_bills"] += total_bills
        grand_total["received_bills"] += received_bills
        grand_total["remaining_bills"] += remaining_bills
        grand_total["amount_received"] += amount_received
        grand_total["pending_amount"] += pending_amount

    gt_total_amount = grand_total["amount_received"] + grand_total["pending_amount"]
    summary_rows.append(["", "Grand Total", grand_total["total_bills"], grand_total["received_bills"], grand_total["remaining_bills"], gt_total_amount, grand_total["amount_received"], grand_total["pending_amount"]])

    summary_headers = ["Sr", "Staff Name", "Total Bills", "Received Bills", "Remaining Bills", "Total Amount", "Amount Received", "Pending Amount"]

    # Apply column filtering for CSV/Excel
    # Detail cols filtering
    if cols_param:
        _sel = [k.strip() for k in cols_param.split(",") if k.strip() in STAFF_COL_MAP]
        _csv_detail_cols = sorted([STAFF_COL_MAP[k] for k in _sel if k in STAFF_COL_MAP])
        if _csv_detail_cols:
            headers = [headers[i] for i in _csv_detail_cols]
            rows = [[r[i] for i in _csv_detail_cols] for r in rows]
    # Summary cols filtering
    _sum_keys = ["sr", "staffName", "totalBills", "receivedBills", "remainingBills", "totalAmount", "amountReceived", "pendingAmount"]
    _csv_summary_cols = list(range(8))
    if summary_cols_param:
        _sum_sel = [k.strip() for k in summary_cols_param.split(",") if k.strip() in STAFF_SUMMARY_COL_MAP]
        _csv_summary_cols = sorted([STAFF_SUMMARY_COL_MAP[k] for k in _sum_sel]) if _sum_sel else list(range(8))
    if _csv_summary_cols != list(range(8)):
        summary_headers = [summary_headers[i] for i in _csv_summary_cols]
        summary_rows = [[r[i] for i in _csv_summary_cols] for r in summary_rows]

    if fmt_type == "csv":
        csv_lines = []
        csv_lines.append(",".join(summary_headers))
        for sr in summary_rows:
            csv_lines.append(",".join(str(x) for x in sr))
        csv_lines.append("")
        csv_lines.append("Detailed Report")
        csv_lines.append(",".join(headers))
        for r in rows:
            csv_lines.append(",".join(str(x) for x in r))
        csv_data = "\n".join(csv_lines)
        return Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=bill_list_staff_report.csv"},
        )

    if fmt_type == "xlsx":
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            pd.DataFrame(summary_rows, columns=summary_headers).to_excel(writer, sheet_name="Summary", index=False)
            pd.DataFrame(rows, columns=headers).to_excel(writer, sheet_name="Detailed", index=False)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=bill_list_staff_report.xlsx"},
        )

    flash("Unknown export format.")
    return redirect(url_for("bill_list"))


def _get_season_bill_ids(season: str) -> set[int]:
    """Return set of bill IDs whose due date falls in the given season.
    season: 'jan-jun' or 'jul-dec'
    """
    init_bill_list_db()
    target_months = set(range(1, 7)) if season == "jan-jun" else set(range(7, 13))
    ids: set[int] = set()
    with get_db() as conn:
        rows = conn.execute("SELECT id, raw_data FROM bills").fetchall()
    for row in rows:
        try:
            data = json.loads(row["raw_data"])
            due = data.get("due date")
            if due:
                dt = pd.to_datetime(str(due), dayfirst=True, errors="coerce")
                if pd.notna(dt) and dt.month in target_months:
                    ids.add(row["id"])
        except Exception:
            continue
    return ids


def _get_season_bill_ids(year: int, season: str) -> set[int]:
    """Return set of bill IDs whose due date falls in the given season and year.
    season: 'jan-jun' or 'jul-dec'
    year: the calendar year
    """
    init_bill_list_db()
    target_months = set(range(1, 7)) if season == "jan-jun" else set(range(7, 13))
    ids: set[int] = set()
    with get_db() as conn:
        rows = conn.execute("SELECT id, raw_data FROM bills").fetchall()
    for row in rows:
        try:
            data = json.loads(row["raw_data"])
            due = data.get("due date")
            if due:
                dt = pd.to_datetime(str(due), dayfirst=True, errors="coerce")
                if pd.notna(dt) and dt.year == year and dt.month in target_months:
                    ids.add(row["id"])
        except Exception:
            continue
    return ids


def bill_list_sector_seasonly_export_rows(year: int, season: str):
    """Build sector-wise six-month season report rows.

    Returns (headers, detail_rows, grand_total_row) where each detail row is:
    [sr, sector_name, total_bills, received_bills, remaining_bills, amount_received, pending_amount]
    """
    init_bill_list_db()
    season_bill_ids = _get_season_bill_ids(year, season)

    headers = [
        "Sr",
        "Sector Name",
        "Total Bills",
        "Received Bills",
        "Remaining Bills",
        "Amount Received",
        "Pending Amount",
    ]

    if not season_bill_ids:
        return headers, [], ["", "Grand Total", "0", "0", "0", "0", "0"]

    id_list = ",".join(str(i) for i in season_bill_ids)

    with get_db() as conn:
        # First, get bill-level data to calculate Water Fee from raw_data
        bill_rows = conn.execute(
            f"""
            SELECT
                id,
                COALESCE(NULLIF(TRIM(sector), ''), 'Unassigned Sector') AS sector,
                amount_received,
                raw_data
            FROM bills
            WHERE id IN ({id_list})
            """
        ).fetchall()

    # Process bill by bill to extract Water Fee from raw_data
    sector_stats: dict[str, dict] = {}
    for row in bill_rows:
        sector = row["sector"]
        amount_received = float(row["amount_received"] or 0)

        # Extract Water Fee from raw_data JSON
        try:
            data = json.loads(row["raw_data"])
            water_fee = 0.0
            wf = data.get("water fee")
            if wf is not None:
                water_fee = parse_number(str(wf).replace(",", ""))
        except (json.JSONDecodeError, ValueError):
            water_fee = 0.0

        if sector not in sector_stats:
            sector_stats[sector] = {
                "total_bills": 0,
                "received_bills": 0,
                "remaining_bills": 0,
                "amount_received": 0.0,
                "pending_amount": 0.0,
            }

        stats = sector_stats[sector]
        stats["total_bills"] += 1

        if amount_received > 0:
            stats["received_bills"] += 1
            stats["amount_received"] += amount_received
        else:
            stats["remaining_bills"] += 1
            stats["pending_amount"] += water_fee

    rows = []
    grand = {"total_bills": 0, "received_bills": 0, "remaining_bills": 0,
             "amount_received": 0, "pending_amount": 0}

    for idx, (sector_name, stats) in enumerate(sorted(sector_stats.items(), key=lambda x: x[0].lower()), start=1):
        rows.append([
            idx,
            sector_name,
            fmt(stats["total_bills"]),
            fmt(stats["received_bills"]),
            fmt(stats["remaining_bills"]),
            fmt(stats["amount_received"]),
            fmt(stats["pending_amount"]),
        ])

        grand["total_bills"] += stats["total_bills"]
        grand["received_bills"] += stats["received_bills"]
        grand["remaining_bills"] += stats["remaining_bills"]
        grand["amount_received"] += stats["amount_received"]
        grand["pending_amount"] += stats["pending_amount"]

    grand_row = [
        "",
        "Grand Total",
        fmt(grand["total_bills"]),
        fmt(grand["received_bills"]),
        fmt(grand["remaining_bills"]),
        fmt(grand["amount_received"]),
        fmt(grand["pending_amount"]),
    ]

    return headers, rows, grand_row


@app.route("/bill-list/export/six-month-pitch/<fmt_type>")
def export_six_month_pitch(fmt_type: str):
    cols_param = request.args.get("cols")
    season = request.args.get("season", "").strip().lower()

    # Auto-detect default season from current month
    if season not in ("jan-jun", "jul-dec"):
        current_month = datetime.now().month
        season = "jan-jun" if current_month <= 6 else "jul-dec"

    season_label = "January to June" if season == "jan-jun" else "July to December"
    report_title = f"Six Month {season_label} Report"
    file_slug = f"Six_Month_{season_label.replace(' ', '_')}_Report"

    # Get bill IDs for the selected season
    season_bill_ids = _get_season_bill_ids(season)

    headers, detail_rows = bill_list_staff_export_rows(bill_ids=season_bill_ids)
    init_bill_list_db()
    with get_db() as conn:
        unpaid_summary = build_unpaid_amount_summary(conn, bill_ids=season_bill_ids)
    unpaid_staff_rows = unpaid_summary.get("staff_rows", [])

    # Build unpaid lookup: normalized base name → current_bill_amount
    unpaid_lookup = {}
    for row in unpaid_staff_rows:
        norm = _normalize_staff_name(row["name"])
        closest = _closest_staff_key(norm)
        key = _normalize_staff_name(closest) if closest and closest != norm else norm
        unpaid_lookup[key] = float(row["current_bill_amount"])

    # Group staff summary by staff name (10-col: sr(0), staff(1), zone(2), sector(3), locality(4), totalBills(5), receivedBills(6), remainingBills(7), totalReceivedAmount(8), pendingAmount(9))
    def pn(v):
        return parse_number(str(v).replace(",", ""))
    staff_groups = {}
    for row in detail_rows:
        name = row[1]
        staff_groups.setdefault(name, []).append(row)

    pitch_rows = []
    pitch_grand = {"totalBills": 0, "receivedBills": 0, "remainingBills": 0, "totalAmount": 0, "amountReceived": 0, "currentBillAmount": 0}
    for idx, (staff_name, group_rows) in enumerate(sorted(staff_groups.items()), start=1):
        total_bills = sum(pn(r[5]) for r in group_rows)
        received_bills = sum(pn(r[6]) for r in group_rows)
        remaining_bills = sum(pn(r[7]) for r in group_rows)
        amount_received = sum(pn(r[8]) for r in group_rows)
        pending_amount = sum(pn(r[9]) for r in group_rows)
        total_amount = amount_received + pending_amount

        norm = _normalize_staff_name(staff_name)
        closest = _closest_staff_key(norm)
        lookup_key = _normalize_staff_name(closest) if closest and closest != norm else norm if norm in unpaid_lookup else _normalize_staff_name(staff_name)
        current_bill = unpaid_lookup.get(lookup_key, 0)

        pitch_rows.append([
            idx,
            fmt_staff_name(staff_name),
            fmt(total_bills),
            fmt(received_bills),
            fmt(remaining_bills),
            fmt(total_amount),
            fmt(amount_received),
            fmt(current_bill),
        ])
        pitch_grand["totalBills"] += total_bills
        pitch_grand["receivedBills"] += received_bills
        pitch_grand["remainingBills"] += remaining_bills
        pitch_grand["totalAmount"] += total_amount
        pitch_grand["amountReceived"] += amount_received
        pitch_grand["currentBillAmount"] += current_bill

    # The six-month season report follows the same checkbox selection rule as the visible staff rows.
    pitch_rows = _filter_rows_by_selection(pitch_rows, lambda row: str(row[1]))
    if _selection_has_filter():
        pitch_grand = {
            "totalBills": sum(parse_number(row[2]) for row in pitch_rows),
            "receivedBills": sum(parse_number(row[3]) for row in pitch_rows),
            "remainingBills": sum(parse_number(row[4]) for row in pitch_rows),
            "totalAmount": sum(parse_number(row[5]) for row in pitch_rows),
            "amountReceived": sum(parse_number(row[6]) for row in pitch_rows),
            "currentBillAmount": sum(parse_number(row[7]) for row in pitch_rows),
        }

    pitch_grand_row = ["", "Grand Total", fmt(pitch_grand["totalBills"]), fmt(pitch_grand["receivedBills"]), fmt(pitch_grand["remainingBills"]), fmt(pitch_grand["totalAmount"]), fmt(pitch_grand["amountReceived"]), fmt(pitch_grand["currentBillAmount"])]

    pitch_headers = ["Sr", "Staff Name", "Total Bills", "Received Bills", "Remaining Bills", "Total Amount", "Amount Received", "Pending Amount"]

    # Column filtering
    _pitch_sel = []
    if cols_param:
        _pitch_sel = [k.strip() for k in cols_param.split(",") if k.strip() in PITCH_COL_MAP]
    _pitch_pdf_cols = sorted([PITCH_COL_MAP[k] for k in _pitch_sel]) if _pitch_sel else list(range(8))
    _pitch_headers = [pitch_headers[i] for i in _pitch_pdf_cols]
    _pitch_left_cols = {_pitch_pdf_cols.index(1)} if 1 in _pitch_pdf_cols else set()

    # PDF
    if fmt_type == "pdf":
        buf = io.BytesIO()
        margins = 8 * mm
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=6*mm, bottomMargin=6*mm, leftMargin=6*mm, rightMargin=6*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("PitchTitle", parent=styles["Heading1"], fontSize=20, textColor=ACCENT, alignment=1, spaceAfter=6*mm, fontName="Helvetica-Bold")
        elements = [Paragraph(report_title, title_style)]
        page_w = landscape(A4)[0] - 2 * margins
        _all_pw = [page_w*0.04, page_w*0.15, page_w*0.09, page_w*0.13, page_w*0.13, page_w*0.14, page_w*0.14, page_w*0.18]
        _pw = [_all_pw[i] for i in _pitch_pdf_cols]

        def wrap_left(v):
            return wrap_pdf_body_cells([[v]], font_size=9, left_columns={0})[0][0]

        body_rows = []
        for r in pitch_rows:
            full = [r[0], wrap_left(r[1]), r[2], r[3], r[4], r[5], r[6], r[7]]
            body_rows.append([full[i] for i in _pitch_pdf_cols])
        gt_full = [pitch_grand_row[0], pitch_grand_row[1], pitch_grand_row[2], pitch_grand_row[3], pitch_grand_row[4], pitch_grand_row[5], pitch_grand_row[6], pitch_grand_row[7]]
        body_rows.append([gt_full[i] for i in _pitch_pdf_cols])
        table_data = [_pitch_headers] + body_rows
        elements.append(_make_pdf_table(table_data, col_widths=_pw, left_cols=_pitch_left_cols, header_font_size=10, body_font_size=9, cell_padding=6))
        doc.build(elements)
        buf.seek(0)
        return Response(buf.getvalue(), mimetype="application/pdf", headers={"Content-Disposition": f"attachment; filename={file_slug}.pdf"})

    # CSV
    if fmt_type == "csv":
        _csv_rows = [[r[i] for i in _pitch_pdf_cols] for r in pitch_rows]
        _csv_gt = [pitch_grand_row[i] for i in _pitch_pdf_cols]
        _csv_rows.append(_csv_gt)
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(_pitch_headers)
        writer.writerows(_csv_rows)
        return Response(out.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={file_slug}.csv"})

    # Excel
    if fmt_type == "xlsx":
        _ex_rows = [[r[i] for i in _pitch_pdf_cols] for r in pitch_rows]
        _ex_gt = [pitch_grand_row[i] for i in _pitch_pdf_cols]
        _ex_rows.append(_ex_gt)
        df = pd.DataFrame(_ex_rows, columns=_pitch_headers)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="SixMonthPitch", index=False)
        buf.seek(0)
        return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={file_slug}.xlsx"})

    flash("Unknown export format.")
    return redirect(url_for("bill_list"))


# Column key map for six-month sector-wise report: Sr, Sector Name, Total Bills, Received Bills, Remaining Bills, Amount Received, Pending Amount
SEASON_SECTOR_COL_MAP = {"sr": 0, "sector": 1, "totalBills": 2, "receivedBills": 3, "remainingBills": 4, "amountReceived": 5, "pendingAmount": 6}


@app.route("/bill-list/export/season-sector/<fmt_type>")
def export_season_sector_pitch(fmt_type: str):
    cols_param = request.args.get("cols")
    season = request.args.get("season", "").strip().lower()
    year_param = request.args.get("year", "").strip()

    # Validate season
    if season not in ("jan-jun", "jul-dec"):
        season = "jan-jun"

    # Validate year
    try:
        year = int(year_param)
    except (TypeError, ValueError):
        year = datetime.now().year

    season_label = "January to June" if season == "jan-jun" else "July to December"
    report_title = season_label
    file_slug = f"Sector_Wise_{season_label.replace(' ', '_')}_{year}"

    headers, detail_rows, grand_row = bill_list_sector_seasonly_export_rows(year, season)

    # Row selection filtering
    detail_rows = _filter_rows_by_selection(detail_rows, lambda row: str(row[1]))
    if _selection_has_filter():
        grand_row = [
            "",
            "Grand Total",
            fmt(sum(parse_number(str(r[2]).replace(",", "")) for r in detail_rows)),
            fmt(sum(parse_number(str(r[3]).replace(",", "")) for r in detail_rows)),
            fmt(sum(parse_number(str(r[4]).replace(",", "")) for r in detail_rows)),
            fmt(sum(parse_number(str(r[5]).replace(",", "")) for r in detail_rows)),
            fmt(sum(parse_number(str(r[6]).replace(",", "")) for r in detail_rows)),
        ]

    # Column filtering
    _sel = []
    if cols_param:
        _sel = [k.strip() for k in cols_param.split(",") if k.strip() in SEASON_SECTOR_COL_MAP]
    _pdf_cols = sorted([SEASON_SECTOR_COL_MAP[k] for k in _sel]) if _sel else list(range(7))
    _filtered_headers = [headers[i] for i in _pdf_cols]
    _left_cols = {_pdf_cols.index(1)} if 1 in _pdf_cols else set()

    # PDF
    if fmt_type == "pdf":
        buf = io.BytesIO()
        margins = 8 * mm
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=6*mm, bottomMargin=6*mm, leftMargin=6*mm, rightMargin=6*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("SeasonSectorTitle", parent=styles["Heading1"], fontSize=20, textColor=ACCENT, alignment=1, spaceAfter=6*mm, fontName="Helvetica-Bold")
        elements = [Paragraph(report_title, title_style)]
        page_w = landscape(A4)[0] - 2 * margins
        _all_pw = [page_w*0.05, page_w*0.30, page_w*0.10, page_w*0.12, page_w*0.12, page_w*0.15, page_w*0.16]
        _pw = [_all_pw[i] for i in _pdf_cols]

        def wrap_left(v):
            return wrap_pdf_body_cells([[v]], font_size=9, left_columns={0})[0][0]

        body_rows = []
        for r in detail_rows:
            full = [r[0], wrap_left(r[1]), r[2], r[3], r[4], r[5], r[6]]
            body_rows.append([full[i] for i in _pdf_cols])
        gt_full = [grand_row[0], wrap_left(grand_row[1]), grand_row[2], grand_row[3], grand_row[4], grand_row[5], grand_row[6]]
        body_rows.append([gt_full[i] for i in _pdf_cols])
        table_data = [_filtered_headers] + body_rows
        elements.append(_make_pdf_table(table_data, col_widths=_pw, left_cols=_left_cols, header_font_size=10, body_font_size=9, cell_padding=6))
        doc.build(elements)
        buf.seek(0)
        return Response(buf.getvalue(), mimetype="application/pdf", headers={"Content-Disposition": f"attachment; filename={file_slug}.pdf"})

    # CSV
    if fmt_type == "csv":
        _csv_rows = [[r[i] for i in _pdf_cols] for r in detail_rows]
        _csv_gt = [grand_row[i] for i in _pdf_cols]
        _csv_rows.append(_csv_gt)
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(_filtered_headers)
        writer.writerows(_csv_rows)
        return Response(out.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={file_slug}.csv"})

    # Excel
    if fmt_type == "xlsx":
        _ex_rows = [[r[i] for i in _pdf_cols] for r in detail_rows]
        _ex_gt = [grand_row[i] for i in _pdf_cols]
        _ex_rows.append(_ex_gt)
        df = pd.DataFrame(_ex_rows, columns=_filtered_headers)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="SectorWise", index=False)
        buf.seek(0)
        return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={file_slug}.xlsx"})

    flash("Unknown export format.")
    return redirect(url_for("bill_list"))


@app.route("/bill-list/advanced-filter/<fmt_type>")
def export_advanced_bills(fmt_type: str):
    outstanding_amount = request.args.get("outstanding_amount")
    outstanding_amount = parse_number(outstanding_amount) if outstanding_amount else None

    outstanding_operator = request.args.get("outstanding_operator", "gt")

    bill_status = request.args.get("bill_status") or ""

    sector = request.args.get("sector") or None
    zone = request.args.get("zone") or None
    staff_id = request.args.get("staff_id")
    staff_id = int(staff_id) if staff_id and staff_id.isdigit() else None
    show_summary = request.args.get("show_summary", "0") == "1"
    cols_param = request.args.get("cols")
    group_by = request.args.get("group_by", "normal")

    bills = get_filtered_bills(
        outstanding_amount=outstanding_amount,
        outstanding_operator=outstanding_operator,
        bill_status=bill_status,
        sector=sector,
        zone=zone,
        staff_id=staff_id,
    )

    staff_name = None
    if staff_id:
        with get_db() as conn:
            row = conn.execute("SELECT name FROM staff WHERE id = ?", (staff_id,)).fetchone()
            staff_name = row["name"] if row else f"Staff #{staff_id}"

    filters_list = []
    if bill_status:
        filters_list.append("Paid" if bill_status == "paid" else "Unpaid")
    if outstanding_amount:
        op_label = ">" if outstanding_operator == "gt" else "<"
        filters_list.append(f"Outstanding {op_label} {fmt(outstanding_amount)}")
    if sector:
        filters_list.append(f"Sector: {sector}")
    if zone:
        filters_list.append(f"Zone: {zone}")
    if staff_name:
        filters_list.append(f"Staff: {staff_name}")

    filters_applied = ", ".join(filters_list) if filters_list else "None"

    if fmt_type == "zip":
        if not bills:
            flash("No bills match the selected filters.")
            return redirect(url_for("bill_list"))
        if group_by not in ("sector", "zone", "staff"):
            flash("ZIP export is available for Sector-wise, Zone-wise, and Staff-wise reports only.")
            return redirect(url_for("bill_list"))
        groups = group_bills(bills, group_by)
        if not groups:
            flash("No bills match the selected filters for ZIP export.")
            return redirect(url_for("bill_list"))
        zip_buf = generate_zip_of_group_pdfs(group_by, groups, filters_applied, cols_param=cols_param)
        zip_parts = [f"{group_by.capitalize()}_Wise"]
        if bill_status:
            zip_parts.append("Paid" if bill_status == "paid" else "Unpaid")
        if outstanding_amount:
            op_str = "Gt" if outstanding_operator == "gt" else "Lt"
            zip_parts.append(f"Outstanding{op_str}_{int(outstanding_amount)}")
        if sector:
            zip_parts.append(f"Sector_{sector}")
        if zone:
            zip_parts.append(f"Zone_{zone}")
        if staff_name:
            zip_parts.append(f"Staff_{staff_name.replace(' ', '_')}")
        zip_filename = sanitize_filename("_".join(zip_parts)) + ".zip"
        return Response(zip_buf.getvalue(), mimetype="application/zip", headers={"Content-Disposition": f"attachment; filename={zip_filename}"})

    return export_advanced_bills_response(fmt_type, bills, filters_applied, show_summary=show_summary, cols_param=cols_param, group_by=group_by)


# ---------------------------------------------------------------------------
# Summary Reports (Zones, Sectors, Staff)
# ---------------------------------------------------------------------------

def get_zones_summary():
    init_bill_list_db()
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT
                zone,
                COUNT(*) AS bill_count,
                SUM(total_bill) AS total_amount
            FROM bills
            GROUP BY zone
            ORDER BY CASE zone WHEN 'A' THEN 1 WHEN 'B' THEN 2 WHEN 'C' THEN 3 WHEN 'Commercial' THEN 4 ELSE 5 END
            """
        ).fetchall()
    return [
        {"name": row["zone"], "count": int(row["bill_count"] or 0), "total_amount": float(row["total_amount"] or 0)}
        for row in rows
    ]


def get_sectors_summary():
    init_bill_list_db()
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT
                sector,
                COUNT(*) AS bill_count,
                SUM(total_bill) AS total_amount
            FROM bills
            GROUP BY sector
            ORDER BY sector
            """
        ).fetchall()
    return [
        {"name": row["sector"], "count": int(row["bill_count"] or 0), "total_amount": float(row["total_amount"] or 0)}
        for row in rows
    ]


def get_staff_summary():
    init_bill_list_db()
    with get_db() as conn:
        has_auto_rules = conn.execute("SELECT COUNT(*) AS cnt FROM auto_assignment_rules").fetchone()["cnt"] > 0

        if has_auto_rules:
            auto_exclude = "AND NOT EXISTS (SELECT 1 FROM auto_assignment_rules aar WHERE aar.sector = b.sector AND aar.locality = b.locality)"
            auto_assign_sql = """
            UNION ALL
            SELECT
                s.name AS staff_name,
                COUNT(b.id) AS bill_count,
                SUM(b.total_bill) AS total_amount
            FROM bills b
            JOIN auto_assignment_rules aar
                ON aar.sector = b.sector AND aar.locality = b.locality
                AND CAST(b.connection_no AS TEXT) >= aar.connection_min
                AND (aar.connection_max IS NULL OR CAST(b.connection_no AS TEXT) <= aar.connection_max)
            JOIN staff s ON UPPER(TRIM(s.name)) = UPPER(TRIM(aar.staff_name))
            GROUP BY s.id, s.name
            """
        else:
            auto_exclude = ""
            auto_assign_sql = ""

        rows = conn.execute(
            f"""
            SELECT
                s.name AS staff_name,
                COUNT(b.id) AS bill_count,
                SUM(b.total_bill) AS total_amount
            FROM staff_assignments sa
            JOIN staff s ON s.id = sa.staff_id
            LEFT JOIN bills b
                ON b.zone = sa.zone
                AND (sa.sector IS NULL OR b.sector = sa.sector)
                AND (sa.locality IS NULL OR b.locality = sa.locality)
                {auto_exclude}
            GROUP BY s.id, s.name

            {auto_assign_sql}

            UNION ALL
            SELECT
                'Unassigned' AS staff_name,
                COUNT(b.id) AS bill_count,
                SUM(b.total_bill) AS total_amount
            FROM bills b
            WHERE NOT EXISTS (
                SELECT 1
                FROM staff_assignments sa
                WHERE sa.zone = b.zone
                  AND (sa.sector IS NULL OR sa.sector = b.sector)
                  AND (sa.locality IS NULL OR sa.locality = b.locality)
            )
            {('AND NOT EXISTS (SELECT 1 FROM auto_assignment_rules aar WHERE aar.sector = b.sector AND aar.locality = b.locality)' if has_auto_rules else '')}
            HAVING COUNT(b.id) > 0
            ORDER BY staff_name
            """
        ).fetchall()
    return [
        {"name": fmt_staff_name(row["staff_name"]), "count": int(row["bill_count"] or 0), "total_amount": float(row["total_amount"] or 0)}
        for row in rows
    ]


def generate_summary_pdf(title: str, headers: list[str], rows: list[list], total_row: list[str] = None, show_summary: bool = True) -> bytes:
    page_w = A4[0] - 30 * mm
    col_widths = [page_w / len(headers)] * len(headers)
    summary_lines = [f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}"] if show_summary else []

    return generate_card_pdf(
        title,
        summary_lines,
        headers,
        rows,
        total_row,
        pagesize=A4,
        col_widths=col_widths,
        left_cols=[0],
        header_font_size=11,
        body_font_size=10,
    )


def export_summary_response(fmt_type: str, title: str, data: list[dict], filename_prefix: str, show_summary: bool = True, cols_param: str = None):
    # Summary card exports are filtered by the row checkboxes before totals are generated.
    data = _filter_rows_by_selection(data, lambda item: str(item["name"]))
    all_headers = ["Name", "Count", "Total Amount"]
    all_rows = [[item["name"], fmt(item["count"]), fmt(item["total_amount"])] for item in data]
    headers, rows = parse_export_cols(cols_param, SUMMARY_COL_MAP, all_headers, all_rows)

    total_count = sum(item["count"] for item in data)
    total_amount = sum(item["total_amount"] for item in data)
    total_row = ["Total", fmt(total_count), fmt(total_amount)]
    # Filter total_row to match selected columns (same key order as headers)
    if cols_param:
        _sel = [k.strip() for k in cols_param.split(",") if k.strip() in SUMMARY_COL_MAP]
        _idx = [SUMMARY_COL_MAP[k] for k in _sel]
        total_row = [total_row[i] for i in _idx] if _idx else total_row

    if fmt_type == "pdf":
        pdf_bytes = generate_summary_pdf(title, headers, rows, total_row, show_summary=show_summary)
        return Response(pdf_bytes, mimetype="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename_prefix}.pdf"})

    if fmt_type == "csv":
        df = pd.DataFrame(rows, columns=headers)
        total_df = pd.DataFrame([total_row], columns=headers)
        combined_df = pd.concat([df, total_df], ignore_index=True)
        csv_data = combined_df.to_csv(index=False)
        return Response(csv_data, mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={filename_prefix}.csv"})

    if fmt_type == "xlsx":
        buf = io.BytesIO()
        df = pd.DataFrame(rows, columns=headers)
        total_df = pd.DataFrame([total_row], columns=headers)
        combined_df = pd.concat([df, total_df], ignore_index=True)
        combined_df.to_excel(buf, index=False)
        buf.seek(0)
        return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename_prefix}.xlsx"})

    flash("Unknown export format.")
    return redirect(url_for("bill_list"))


@app.route("/bill-list/summary/zones/<fmt_type>")
def export_zones_summary(fmt_type: str):
    show_summary = request.args.get("show_summary", "0") == "1"
    cols_param = request.args.get("cols")
    data = get_zones_summary()
    return export_summary_response(fmt_type, "All Zones Summary Report", data, "all_zones_summary", show_summary=show_summary, cols_param=cols_param)


@app.route("/bill-list/summary/sectors/<fmt_type>")
def export_sectors_summary(fmt_type: str):
    show_summary = request.args.get("show_summary", "0") == "1"
    cols_param = request.args.get("cols")
    data = get_sectors_summary()
    return export_summary_response(fmt_type, "All Sectors Summary Report", data, "all_sectors_summary", show_summary=show_summary, cols_param=cols_param)


@app.route("/bill-list/summary/staff/<fmt_type>")
def export_staff_summary(fmt_type: str):
    show_summary = request.args.get("show_summary", "0") == "1"
    cols_param = request.args.get("cols")
    data = get_staff_summary()
    return export_summary_response(fmt_type, "All Staff Summary Report", data, "all_staff_summary", show_summary=show_summary, cols_param=cols_param)


# ---------------------------------------------------------------------------
# Card download endpoints (PDF / CSV / Excel)
# ---------------------------------------------------------------------------

def _card_rows_to_df(headers, rows):
    return pd.DataFrame(rows, columns=headers)


def commercial_daily_income_export_rows(results: dict):
    metric_header = results.get("commercial_daily_metric_label") or "Arrears Received"
    headers = [
        "Sr",
        "Date",
        "Consumer Name",
        "Connection No",
        "Sector",
        "Locality",
        metric_header,
        "Amount Received",
    ]
    rows = []
    current_date = None
    day_metric = 0
    day_amount = 0
    total_metric = 0
    total_amount = 0
    sr = 1

    def append_day_total(date_label: str):
        rows.append(["", "", f"{date_label} Total", "", "", "", fmt(day_metric), fmt(day_amount)])

    for row in results.get("commercial_daily_income") or []:
        date_label = row.get("date", "")
        if current_date is not None and date_label != current_date:
            append_day_total(current_date)
            day_metric = 0
            day_amount = 0
        current_date = date_label
        metric_total = row.get("metric_total", 0)
        amount_total = row.get("amount_total", 0)
        rows.append(
            [
                sr,
                date_label,
                row.get("consumer_name", ""),
                row.get("connection_no", ""),
                row.get("sector", ""),
                row.get("locality", ""),
                fmt(metric_total),
                fmt(amount_total),
            ]
        )
        sr += 1
        day_metric += metric_total
        day_amount += amount_total
        total_metric += metric_total
        total_amount += amount_total

    if current_date is not None:
        append_day_total(current_date)
    grand = ["", "", "Grand Total", "", "", "", fmt(total_metric), fmt(total_amount)] if rows else None
    return headers, rows, grand


def daily_staff_receive_export_tables(results: dict, sort_order: str = "default"):
    report = results.get("daily_staff_receive") or {}
    summary_headers = ["Sr", "Staff Name", "No. of Bills Received", "Arrears Received", "Total Amount Received"]
    detail_headers = ["Staff Name", "Zone", "Sr", "Sector", "Locality", "Bills", "Arrears Received", "Received Amount"]

    summary_rows_data = list(report.get("summary_rows") or [])
    if sort_order == "desc":
        summary_rows_data.sort(key=lambda r: r.get("amount_total", 0), reverse=True)
    elif sort_order == "asc":
        summary_rows_data.sort(key=lambda r: r.get("amount_total", 0), reverse=False)

    summary_rows = []
    total_bills = 0
    total_metric = 0
    total_amount = 0
    for idx, row in enumerate(summary_rows_data, start=1):
        bills = row.get("bills", 0)
        metric_total = row.get("metric_total", 0)
        amount_total = row.get("amount_total", 0)
        summary_rows.append([idx, fmt_staff_name(row.get("staff_name", "")), fmt(bills), fmt(metric_total), fmt(amount_total)])
        total_bills += bills
        total_metric += metric_total
        total_amount += amount_total
    summary_grand = ["", "Grand Total", fmt(total_bills), fmt(total_metric), fmt(total_amount)] if summary_rows else None

    detail_rows: list[list] = []
    for group in report.get("grouped_detail") or []:
        staff_name = fmt_staff_name(group.get("staff_name", ""))
        zone = group.get("zone", "")
        # Merge sub_rows by normalised sector using shared helper
        merged_subs = merge_sector_rows(group.get("sub_rows") or [])
        for idx, merged in enumerate(merged_subs, start=1):
            detail_rows.append(
                [
                    staff_name,
                    zone,
                    idx,
                    merged["sector"],
                    merged["locality"],
                    fmt(merged["bills"]),
                    fmt(merged["metric_total"]),
                    fmt(merged["amount_total"]),
                ]
            )
    return summary_headers, summary_rows, summary_grand, detail_headers, detail_rows


def _calc_daily_summary_col_widths(page_w, n_cols):
    """Calculate column widths for the daily staff receive summary table."""
    if n_cols == 5:
        return [page_w * 0.08, page_w * 0.27, page_w * 0.19, page_w * 0.22, page_w * 0.24]
    elif n_cols == 4:
        return [page_w * 0.10, page_w * 0.30, page_w * 0.25, page_w * 0.35]
    elif n_cols == 3:
        return [page_w * 0.12, page_w * 0.38, page_w * 0.50]
    elif n_cols == 2:
        return [page_w * 0.40, page_w * 0.60]
    elif n_cols == 1:
        return [page_w]
    else:
        return [page_w / n_cols] * n_cols


def _calc_daily_detail_col_widths(page_w, display_headers):
    """Calculate column widths for the daily staff receive detail table.

    Layout strategy based on which columns are visible:
    - Sr (narrow), Sector/Locality (wide, left-aligned), numeric cols (equal share).
    """
    n = len(display_headers)
    if n == 0:
        return []

    has_sr = "Sr" in display_headers
    has_sector = "Sector" in display_headers
    has_locality = "Locality" in display_headers
    has_bills = "Bills" in display_headers
    has_arrears = "Arrears Received" in display_headers
    has_amount = "Received Amount" in display_headers

    numeric_count = sum(1 for h in display_headers if h in ("Bills", "Arrears Received", "Received Amount"))

    if n == 6:
        return [page_w * 0.06, page_w * 0.22, page_w * 0.27, page_w * 0.10, page_w * 0.15, page_w * 0.20]
    elif n == 5:
        return [page_w * 0.07, page_w * 0.25, page_w * 0.30, page_w * 0.14, page_w * 0.24]
    elif n == 4:
        return [page_w * 0.08, page_w * 0.30, page_w * 0.30, page_w * 0.32]
    elif n == 3:
        return [page_w * 0.10, page_w * 0.40, page_w * 0.50]
    elif n == 2:
        return [page_w * 0.40, page_w * 0.60]
    elif n == 1:
        return [page_w]
    else:
        return [page_w / n] * n


def generate_daily_staff_receive_pdf(results: dict, sort_order: str = "default", summary_cols: str = "", detail_cols: str = "") -> bytes:
    summary_headers, summary_rows, summary_grand, detail_headers, detail_rows = daily_staff_receive_export_tables(results, sort_order=sort_order)

    # Apply column selection filtering for summary (skip UI-only keys like "avatar", "perf")
    # Save original headers before filtering so grand total can be filtered correctly
    orig_summary_headers = summary_headers
    if summary_cols:
        summary_headers, summary_rows = parse_export_cols(summary_cols, DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP, summary_headers, summary_rows)
        if summary_grand:
            _, g = parse_export_cols(summary_cols, DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP, orig_summary_headers, [summary_grand])
            summary_grand = g[0] if g else summary_grand

    # Apply column selection filtering for detail (skip UI-only keys like "avatar", "perf")
    orig_detail_headers = detail_headers
    if detail_cols:
        detail_headers, detail_rows = parse_export_cols(detail_cols, DAILY_STAFF_RECEIVE_DETAIL_COL_MAP, detail_headers, detail_rows)

    report = results.get("daily_staff_receive") or {}
    buf = io.BytesIO()

    portrait_size = A4
    top_margin = 15 * mm
    bottom_margin = 15 * mm
    left_margin = 15 * mm
    right_margin = 15 * mm
    portrait_frame = Frame(left_margin, bottom_margin, portrait_size[0] - left_margin - right_margin, portrait_size[1] - top_margin - bottom_margin, id="portrait")
    doc = BaseDocTemplate(
        buf,
        pagesize=portrait_size,
        pageTemplates=[
            PageTemplate(id="portrait", frames=[portrait_frame], pagesize=portrait_size),
        ],
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("DailyStaffTitle", parent=styles["Heading1"], fontSize=19, textColor=ACCENT, alignment=1, spaceAfter=5 * mm, fontName="Helvetica-Bold")
    summary_style = ParagraphStyle("DailyStaffSummary", parent=styles["Normal"], fontSize=11, leading=15, spaceAfter=2 * mm)
    group_style = ParagraphStyle("DailyStaffGroup", parent=styles["Heading3"], fontSize=13, leading=16, spaceBefore=5 * mm, spaceAfter=2 * mm, fontName="Helvetica-Bold")
    progress_style_small = ParagraphStyle(
        "DailyStaffProgressSmall",
        parent=styles["Normal"],
        fontSize=9,
        leading=13,
        spaceBefore=1 * mm,
        spaceAfter=3 * mm,
        textColor=colors.HexColor("#222222"),
    )

    elements = [
        Paragraph("Daily Receive Amount of Staff", title_style),
        Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}", summary_style),
    ]
    if report.get("date_range"):
        elements.append(Paragraph(f"<b>Report Date:</b> {report['date_range']}", summary_style))

    page_w = portrait_size[0] - left_margin - right_margin
    report_total_bills = report_total_metric = report_total_amount = 0
    for row in report.get("summary_rows") or []:
        report_total_bills += row.get("bills", 0)
        report_total_metric += row.get("metric_total", 0)
        report_total_amount += row.get("amount_total", 0)
    elements.append(Spacer(1, 3 * mm))
    elements.append(Paragraph(
        f"<b>Total Received Connections:</b> {fmt(report_total_bills)} &nbsp;&nbsp;&nbsp; "
        f"<b>Arrears Received:</b> {fmt(report_total_metric)} &nbsp;&nbsp;&nbsp; "
        f"<b>Total Amount Received:</b> {fmt(report_total_amount)}",
        progress_style_small,
    ))
    elements.append(Spacer(1, 5 * mm))

    # ── FIRST PAGE: Summary table with full-page layout ──
    summary_data = [wrap_pdf_header_cells(summary_headers, font_size=10)] + wrap_pdf_body_cells(summary_rows, font_size=10, left_columns={1})
    if summary_grand:
        summary_data.append(wrap_pdf_body_cells([summary_grand], font_size=10, left_columns={1})[0])

    n_summary_cols = len(summary_headers)
    summary_col_widths = _calc_daily_summary_col_widths(page_w, n_summary_cols)

    # Find staff name column index for left-alignment
    summary_left_cols = []
    for key, idx in DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP.items():
        if key == "staffName" and idx < n_summary_cols:
            summary_left_cols.append(idx)

    # Calculate row heights to fill the A4 page (same approach as summary PDF)
    num_summary_rows = len(summary_data)
    total_page_h = portrait_size[1]
    available_h = total_page_h - top_margin - bottom_margin
    header_content_h = 85 * mm
    table_target_h = available_h - header_content_h
    if table_target_h < 100 * mm:
        table_target_h = 100 * mm
    row_h = table_target_h / num_summary_rows
    row_h = max(row_h, 12 * mm)
    row_h = min(row_h, 35 * mm)
    row_heights = [row_h] * num_summary_rows

    t = Table(summary_data, colWidths=summary_col_widths, repeatRows=1, rowHeights=row_heights)

    cell_pad = min(6, max(4, int((row_h - 10) / 2)))
    table_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
        ("TEXTCOLOR", (0, 0), (-1, 0), HEADER_FG),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_CLR),
        ("TOPPADDING", (0, 0), (-1, -1), cell_pad),
        ("BOTTOMPADDING", (0, 0), (-1, -1), cell_pad),
        ("LEFTPADDING", (0, 0), (-1, -1), max(3, cell_pad - 2)),
        ("RIGHTPADDING", (0, 0), (-1, -1), max(3, cell_pad - 2)),
    ]
    for i in range(1, num_summary_rows):
        if i % 2 == 0:
            table_style_cmds.append(("BACKGROUND", (0, i), (-1, i), ALT_ROW))
    if num_summary_rows > 2:
        last = num_summary_rows - 1
        table_style_cmds.append(("FONTNAME", (0, last), (-1, last), "Helvetica-Bold"))
        table_style_cmds.append(("FONTSIZE", (0, last), (-1, last), 10))
        table_style_cmds.append(("BACKGROUND", (0, last), (-1, last), colors.HexColor("#e6d8c8")))
    for idx in range(1, num_summary_rows):
        row_text = " ".join(str(cell) for cell in summary_data[idx])
        if " Total" in row_text or "Grand Total" in row_text:
            table_style_cmds.append(("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold"))
            table_style_cmds.append(("FONTSIZE", (0, idx), (-1, idx), 10))
            table_style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#e6d8c8")))
            table_style_cmds.append(("TEXTCOLOR", (0, idx), (-1, idx), colors.black))
    if summary_left_cols:
        for lc in summary_left_cols:
            table_style_cmds.append(("ALIGN", (lc, 1), (lc, -1), "LEFT"))

    t.setStyle(TableStyle(table_style_cmds))
    elements.append(t)

    # ── DETAIL PAGES: One page per staff ──
    if detail_rows:
        detail_page_w = portrait_size[0] - left_margin - right_margin
        n_detail_cols = len(detail_headers)

        # Build a map from header name to column index for safe lookups
        detail_header_idx = {h: i for i, h in enumerate(detail_headers)}

        # Identify key column indices by header name (safe even after filtering)
        idx_staff_name = detail_header_idx.get("Staff Name")
        idx_zone = detail_header_idx.get("Zone")
        idx_sr = detail_header_idx.get("Sr")
        idx_sector = detail_header_idx.get("Sector")
        idx_locality = detail_header_idx.get("Locality")
        idx_bills = detail_header_idx.get("Bills")
        idx_arrears = detail_header_idx.get("Arrears Received")
        idx_amount = detail_header_idx.get("Received Amount")

        # Detail columns to display (everything except Staff Name and Zone which are shown in the group header)
        detail_display_indices = [i for i, h in enumerate(detail_headers) if h not in ("Staff Name", "Zone")]
        detail_display_headers = [detail_headers[i] for i in detail_display_indices]
        n_display_cols = len(detail_display_headers)

        # Determine left-aligned column indices in the display table
        display_left_cols = []
        for display_i, orig_i in enumerate(detail_display_indices):
            if detail_headers[orig_i] in ("Sector", "Locality"):
                display_left_cols.append(display_i)

        # Map numeric column positions in the display table for summing
        display_bills_pos = None
        display_arrears_pos = None
        display_amount_pos = None
        for display_i, orig_i in enumerate(detail_display_indices):
            h = detail_headers[orig_i]
            if h == "Bills":
                display_bills_pos = display_i
            elif h == "Arrears Received":
                display_arrears_pos = display_i
            elif h == "Received Amount":
                display_amount_pos = display_i

        # Detail column widths: distribute proportionally based on column count and content type
        detail_col_widths = _calc_daily_detail_col_widths(detail_page_w, detail_display_headers)

        elements.append(PageBreak())
        current_group = None
        group_rows = []

        def flush_group():
            if not group_rows or current_group is None:
                return
            staff_elements = [
                Paragraph(f"Staff: {fmt_staff_name(current_group).replace(chr(10), '<br/>')}", group_style),
            ]
            data_rows = []
            total_bills = total_metric = total_amount = 0
            for row in group_rows:
                # Build display row from filtered columns
                display_row = [row[i] for i in detail_display_indices]
                data_rows.append(display_row)
                # Safely sum numeric totals using the display positions
                if display_bills_pos is not None and display_bills_pos < len(row):
                    total_bills += parse_number(str(row[detail_display_indices[display_bills_pos]]).replace(",", ""))
                if display_arrears_pos is not None and display_arrears_pos < len(row):
                    total_metric += parse_number(str(row[detail_display_indices[display_arrears_pos]]).replace(",", ""))
                if display_amount_pos is not None and display_amount_pos < len(row):
                    total_amount += parse_number(str(row[detail_display_indices[display_amount_pos]]).replace(",", ""))

            # Progress summary line: only include totals for visible numeric columns
            progress_parts = []
            if display_bills_pos is not None:
                progress_parts.append(f"<b>Total Received Connections:</b> {fmt(total_bills)}")
            if display_arrears_pos is not None:
                progress_parts.append(f"<b>Arrears Received:</b> {fmt(total_metric)}")
            if display_amount_pos is not None:
                progress_parts.append(f"<b>Total Amount Received:</b> {fmt(total_amount)}")
            if progress_parts:
                staff_elements.append(Paragraph(" &nbsp;&nbsp;&nbsp; ".join(progress_parts), progress_style_small))
            staff_elements.append(Spacer(1, 3 * mm))

            # Grand Total row: empty cells for text cols, totals for numeric cols
            grand_row = [""] * n_display_cols
            if display_bills_pos is not None:
                grand_row[display_bills_pos] = fmt(total_bills)
            if display_arrears_pos is not None:
                grand_row[display_arrears_pos] = fmt(total_metric)
            if display_amount_pos is not None:
                grand_row[display_amount_pos] = fmt(total_amount)
            # Place "Grand Total" label in the Locality column; fall back to
            # the previous text column if Locality is filtered out.
            label_pos = 0
            prev_text_col = 0
            for ci, orig_i in enumerate(detail_display_indices):
                h = detail_headers[orig_i]
                if h not in ("Bills", "Arrears Received", "Received Amount"):
                    if h == "Locality":
                        label_pos = ci
                        break
                    prev_text_col = ci
            if label_pos == 0 and prev_text_col != 0:
                label_pos = prev_text_col
            grand_row[label_pos] = "Grand Total"
            data_rows.append(grand_row)

            # Wrap all cells for PDF rendering
            # Grand Total is always the last row
            grand_row_idx = len(data_rows) - 1 if data_rows else -1
            # Identify Sector/Locality column positions in the display table
            bracket_col_indices = set()
            for ci, orig_i in enumerate(detail_display_indices):
                if detail_headers[orig_i] in ("Sector", "Locality"):
                    bracket_col_indices.add(ci)
            # Use wrap_pdf_body_cells for consistent styling with bold Grand Total
            # and bracket text rendering for Sector/Locality columns
            text_row_indices = set(range(len(data_rows)))
            wrapped_rows = wrap_pdf_body_cells(
                data_rows,
                font_size=10,
                left_columns=display_left_cols,
                bold_rows={grand_row_idx} if grand_row_idx >= 0 else set(),
                bracket_cols=bracket_col_indices,
            )

            staff_elements.append(
                _make_pdf_table(
                    [wrap_pdf_header_cells(detail_display_headers, font_size=11)] + wrapped_rows,
                    col_widths=detail_col_widths,
                    left_cols=display_left_cols,
                    header_font_size=11,
                    body_font_size=10,
                    cell_padding=3,
                )
            )
            staff_elements.append(Spacer(1, 4 * mm))
            elements.append(KeepTogether(staff_elements))

        for row in detail_rows:
            group = row[0] if row else ""
            if current_group is not None and group != current_group:
                flush_group()
                group_rows = []
            current_group = group
            group_rows.append(row)
        flush_group()

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


def daily_staff_receive_export_response(fmt_type: str, results: dict, cols: str = "", detail_cols: str = "", sort_order: str = "default"):
    summary_headers, summary_rows, summary_grand, detail_headers, detail_rows = daily_staff_receive_export_tables(results, sort_order=sort_order)
    orig_summary_headers = summary_headers
    orig_detail_headers = detail_headers
    if cols:
        summary_headers, summary_rows = parse_export_cols(cols, DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP, summary_headers, summary_rows)
        if summary_grand:
            _, g = parse_export_cols(cols, DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP, orig_summary_headers, [summary_grand])
            summary_grand = g[0] if g else summary_grand
    if detail_cols:
        detail_headers, detail_rows = parse_export_cols(detail_cols, DAILY_STAFF_RECEIVE_DETAIL_COL_MAP, orig_detail_headers, detail_rows)
    filename = "daily_staff_receive_report"
    if fmt_type == "pdf":
        return Response(
            generate_daily_staff_receive_pdf(results, sort_order=sort_order, summary_cols=cols, detail_cols=detail_cols),
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
        )
    if fmt_type == "csv":
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(["Summary", *summary_headers])
        for row in summary_rows:
            writer.writerow(["Summary", *row])
        if summary_grand:
            writer.writerow(["Summary", *summary_grand])
        writer.writerow([])
        writer.writerow(["Details", *detail_headers])
        for row in detail_rows:
            writer.writerow(["Details", *row])
        return Response(
            out.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )
    if fmt_type == "xlsx":
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            pd.DataFrame(summary_rows + ([summary_grand] if summary_grand else []), columns=summary_headers).to_excel(writer, sheet_name="Summary", index=False)
            pd.DataFrame(detail_rows, columns=detail_headers).to_excel(writer, sheet_name="Details", index=False)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )
    flash("Unknown export format.")
    return redirect(url_for("daily_staff_receive"))


@app.route("/download-card/<card>/<fmt_type>")
def download_card(card: str, fmt_type: str):
    r = _last_results
    if not r:
        flash("No data available. Please upload files first.")
        return redirect(url_for("index"))

    sort_order = request.args.get("sort", "default")
    if sort_order not in ("default", "asc", "desc"):
        sort_order = "default"

    if card == "monthly":
        title = "Month-wise Report"
        fiscal_start, current_period = get_fiscal_window()
        period_str = f"{format_calendar_month(fiscal_start)} to {format_calendar_month(current_period)}"
        headers = ["Month", "No. of Bills", "Arrears Received", "Current Amount Received", "Amount Received"]
        rows = []
        gt_count, gt_arrears, gt_amount = 0, 0, 0
        for row in r.get("monthly_rows", []):
            c = row.get("count", 0)
            ar = row.get("arrears_total", 0)
            am = row.get("amount_total", 0)
            current = am - ar
            rows.append([row["label"], fmt(c), fmt(ar), fmt(current), fmt(am)])
            gt_count += c
            gt_arrears += ar
            gt_amount += am
        gt_current = gt_amount - gt_arrears
        summary = [
            f"<b>Period:</b> {period_str}",
            f"<b>Total Amount Received:</b> Rs. {fmt(gt_amount)}",
            f"<b>Total Arrears Received:</b> Rs. {fmt(gt_arrears)}",
            f"<b>Total Current Amount Received:</b> Rs. {fmt(gt_current)}",
        ]
        grand = ["Grand Total", fmt(gt_count), fmt(gt_arrears), fmt(gt_current), fmt(gt_amount)]

    elif card == "daily":
        title = "Date-wise Report"
        summary = [f"<b>Total Connections:</b> {fmt(r.get('fiscal_row_count'))}"]
        headers = ["Date", "No. of Bills", "Arrears Received", "Amount Received"]
        rows = []
        gt_count, gt_arrears, gt_amount = 0, 0, 0
        for row in r.get("daily_rows", []):
            c = row.get("count", 0)
            ar = row.get("arrears_total", 0)
            am = row.get("amount_total", 0)
            rows.append([row["label"], fmt(c), fmt(ar), fmt(am)])
            gt_count += c
            gt_arrears += ar
            gt_amount += am
        grand = ["Grand Total", fmt(gt_count), fmt(gt_arrears), fmt(gt_amount)]

    elif card == "sector":
        title = "Sector-wise Report"
        summary = [f"<b>Total Amount:</b> Rs. {r.get('total_amount_formatted', '0')}"]
        headers = ["Sector", "Arrears Received", "Amount Received"]
        rows = []
        gt_arrears, gt_amount = 0, 0
        for row in r.get("sector_rows", []):
            ar = row.get("arrears_total", 0)
            am = row.get("amount_total", 0)
            rows.append([row["label"], fmt(ar), fmt(am)])
            gt_arrears += ar
            gt_amount += am
        grand = ["Grand Total", fmt(gt_arrears), fmt(gt_amount)]

    elif card in ("commercial", "commercial-total"):
        title = "Commercial Sector — Locality Report"
        summary = [f"<b>Total (July to May)</b>"]
        headers = ["Locality", "No. of Bills"]
        if r.get("has_arrears"):
            headers.append("Arrears Received")
        headers.append("Amount Received")
        rows = []
        gt_count, gt_amount, gt_arrears = 0, 0, 0
        for row in r.get("commercial_total", []):
            c = row.get("count", 0)
            am = row.get("amount_total", 0)
            ar = row.get("arrears_total", 0)
            row_vals = [row["label"], fmt(c)]
            if r.get("has_arrears"):
                row_vals.append(fmt(ar))
                gt_arrears += ar
            row_vals.append(fmt(am))
            rows.append(row_vals)
            gt_count += c
            gt_amount += am
        grand = ["Grand Total", fmt(gt_count)]
        if r.get("has_arrears"):
            grand.append(fmt(gt_arrears))
        grand.append(fmt(gt_amount))

    elif card == "commercial-monthly":
        title = "Commercial Sector - Month-wise Locality Report"
        summary = [f"<b>Month-wise locality breakdown</b>"]
        headers = ["Month", "Locality", "No. of Bills"]
        if r.get("has_arrears"):
            headers.append("Arrears Received")
        headers.append("Amount Received")
        rows = []
        gt_count, gt_amount, gt_arrears = 0, 0, 0
        for month_label, month_rows in r.get("commercial_monthly", {}).items():
            for row in month_rows:
                c = row.get("count", 0)
                am = row.get("amount_total", 0)
                ar = row.get("arrears_total", 0)
                row_vals = [month_label, row["label"], fmt(c)]
                if r.get("has_arrears"):
                    row_vals.append(fmt(ar))
                    gt_arrears += ar
                row_vals.append(fmt(am))
                rows.append(row_vals)
                gt_count += c
                gt_amount += am
        grand = ["Grand Total", "", fmt(gt_count)]
        if r.get("has_arrears"):
            grand.append(fmt(gt_arrears))
        grand.append(fmt(gt_amount))

    elif card == "commercial-daily-income":
        title = "Daily Income for Commercials"
        headers, rows, grand = commercial_daily_income_export_rows(r)
        summary = [
            f"<b>Total Commercial Records:</b> {fmt(len(r.get('commercial_daily_income') or []))}",
            f"<b>{headers[-2]}:</b> Rs. {grand[-2] if grand else '0'}",
            f"<b>Total Amount Received:</b> Rs. {grand[-1] if grand else '0'}",
        ]

    elif card == "commercial-month-wise":
        title = "Commercial Month-wise Report"
        fiscal_start, current_period = get_fiscal_window()
        period_str = f"{format_calendar_month(fiscal_start)} to {format_calendar_month(current_period)}"
        headers = ["Month", "No. of Bills", "Arrears Received", "Current Amount Received", "Amount Received"]
        rows = []
        gt_count, gt_arrears, gt_amount = 0, 0, 0
        for row in r.get("commercial_month_wise_summary", []):
            c = row.get("count", 0)
            ar = row.get("arrears_total", 0)
            am = row.get("amount_total", 0)
            current = am - ar
            rows.append([row["label"], fmt(c), fmt(ar), fmt(current), fmt(am)])
            gt_count += c
            gt_arrears += ar
            gt_amount += am
        gt_current = gt_amount - gt_arrears
        summary = [
            f"<b>Period:</b> {period_str}",
            f"<b>Total Amount Received:</b> Rs. {fmt(gt_amount)}",
            f"<b>Total Arrears Received:</b> Rs. {fmt(gt_arrears)}",
            f"<b>Total Current Amount Received:</b> Rs. {fmt(gt_current)}",
        ]
        grand = ["Grand Total", fmt(gt_count), fmt(gt_arrears), fmt(gt_current), fmt(gt_amount)]

    elif card == "daily-staff-receive":
        title = "Daily Receive Amount of Staff"
        summary_headers, summary_rows, summary_grand, detail_headers, detail_rows = daily_staff_receive_export_tables(r, sort_order=sort_order)
        headers = ["Section", *summary_headers]
        rows = [["Summary", *row] for row in summary_rows]
        grand = ["Summary", *summary_grand] if summary_grand else None
        rows.extend([["", "", "", "", "", ""], ["Details", *detail_headers[:5]]])
        rows.extend([["Details", *row[:5]] for row in detail_rows])
        summary = [f"<b>Report Date:</b> {(r.get('daily_staff_receive') or {}).get('date_range') or 'Uploaded data'}"]

    elif card == "receipt-monthly":
        title = "Month-wise Report"
        headers = ["Month", "No. of Bills", "Arrears Received", "Current Amount Received", "Amount Received"]
        rows = []
        gt_count, gt_arrears, gt_amount = 0, 0, 0
        for row in r.get("receipt_monthly_rows", []):
            c = row.get("count", 0)
            ar = row.get("arrears_total", 0)
            am = row.get("amount_total", 0)
            current = am - ar
            rows.append([row["label"], fmt(c), fmt(ar), fmt(current), fmt(am)])
            gt_count += c
            gt_arrears += ar
            gt_amount += am
        gt_current = gt_amount - gt_arrears
        summary = [
            f"<b>Period:</b> {r.get('receipt_period_start', '')} to {r.get('receipt_period_end', '')}",
            f"<b>Total Amount Received:</b> Rs. {fmt(gt_amount)}",
            f"<b>Total Arrears Received:</b> Rs. {fmt(gt_arrears)}",
            f"<b>Total Current Amount Received:</b> Rs. {fmt(gt_current)}",
        ]
        grand = ["Grand Total", fmt(gt_count), fmt(gt_arrears), fmt(gt_current), fmt(gt_amount)]

    elif card == "connection-type-summary":
        title = "Summary"
        summary = ["<b>Connection type breakdown of total collections</b>"]
        source_rows = r.get("receipt_monthly_rows") or r.get("monthly_rows") or []
        overall_count = sum(row.get("count", 0) for row in source_rows)
        overall_arrears = sum(row.get("arrears_total", 0) for row in source_rows)
        overall_amount = sum(row.get("amount_total", 0) for row in source_rows)
        overall_current = overall_amount - overall_arrears
        comm_count = sum(row.get("count", 0) for row in (r.get("commercial_total") or []))
        comm_arrears = sum(row.get("arrears_total", 0) for row in (r.get("commercial_total") or []))
        comm_amount = sum(row.get("amount_total", 0) for row in (r.get("commercial_total") or []))
        comm_current = comm_amount - comm_arrears
        res_count = overall_count - comm_count
        res_arrears = overall_arrears - comm_arrears
        res_amount = overall_amount - comm_amount
        res_current = res_amount - res_arrears
        headers = ["Name", "No. of Bills", "Arrears Received", "Current Amount Received", "Amount Received"]
        rows = [
            ["Normal / Residential Connections", fmt(res_count), fmt(res_arrears), fmt(res_current), fmt(res_amount)],
            ["Commercial Connections", fmt(comm_count), fmt(comm_arrears), fmt(comm_current), fmt(comm_amount)],
        ]
        grand = ["Grand Total", fmt(overall_count), fmt(overall_arrears), fmt(overall_current), fmt(overall_amount)]

    else:
        flash("Unknown card type.")
        return redirect(url_for("index"))

    # Apply per-card column selection filtering for dashboard exports
    if card != "daily-staff-receive":
        cols_param = request.args.get("cols")
        col_map = _get_card_col_map(card, r)
        if cols_param and col_map:
            headers, rows, grand = _filter_card_export(cols_param, col_map, headers, rows, grand)

    # Build connection type summary if it should be attached to month-wise export
    conn_summary = None
    if request.args.get("include_connection_summary") == "true":
        conn_headers, conn_rows, conn_grand = build_connection_summary(r)
        conn_cols_param = request.args.get("connection_cols")
        if conn_cols_param:
            conn_headers, conn_rows, conn_grand = _filter_card_export(
                conn_cols_param, CONN_SUMMARY_COL_MAP, conn_headers, conn_rows, conn_grand)
        conn_summary = {"headers": conn_headers, "rows": conn_rows, "grand": conn_grand}

    if fmt_type == "pdf":
        if card == "commercial":
            month_rows = {}
            for month_label, m_rows in r.get("commercial_monthly", {}).items():
                section_rows = []
                for m_row in m_rows:
                    vals = [m_row["label"], fmt(m_row.get("count", 0))]
                    if r.get("has_arrears"):
                        vals.append(fmt(m_row.get("arrears_total", 0)))
                    vals.append(fmt(m_row.get("amount_total", 0)))
                    section_rows.append(vals)
                month_rows[fiscal_label_to_calendar_label(month_label)] = section_rows
            pdf_headers, pdf_rows, pdf_grand = remove_pdf_column(headers, rows, "No. of Bills", grand)
            for month_label, section_rows in list(month_rows.items()):
                _, section_pdf_rows, _ = remove_pdf_column(headers, section_rows, "No. of Bills")
                month_rows[month_label] = section_pdf_rows
            pdf_bytes = generate_commercial_pdf(
                summary,
                pdf_headers,
                pdf_rows,
                pdf_grand,
                month_rows,
                r.get("has_arrears"),
            )
        else:
            pdf_kwargs = {}
            pdf_headers, pdf_rows, pdf_grand = headers, rows, grand
            if card == "monthly":
                pdf_rows = [[fiscal_label_to_calendar_label(row[0]), *row[1:]] for row in pdf_rows]
            if card == "commercial-month-wise":
                pdf_rows = [[fiscal_label_to_calendar_label(row[0]), *row[1:]] for row in pdf_rows]
            if card == "sector":
                page_w = landscape(A4)[0] - 30 * mm
                col_widths = [page_w * 0.58, page_w * 0.21, page_w * 0.21]
                pdf_kwargs = {
                    "pagesize": landscape(A4),
                    "col_widths": col_widths,
                    "first_col_left": True,
                }
            elif card == "commercial-total":
                page_w = A4[0] - 30 * mm
                if r.get("has_arrears"):
                    col_widths = [page_w * 0.38, page_w * 0.20, page_w * 0.21, page_w * 0.21]
                else:
                    col_widths = [page_w * 0.48, page_w * 0.26, page_w * 0.26]
                pdf_kwargs = {
                    "pagesize": A4,
                    "col_widths": col_widths,
                    "first_col_left": True,
                }
            elif card == "commercial-monthly":
                monthly_sections = []
                overall_arrears = 0
                overall_amount = 0
                overall_count = 0
                for month_label, month_rows in r.get("commercial_monthly", {}).items():
                    section_rows = []
                    month_arrears = 0
                    month_amount = 0
                    month_count = 0
                    for row in month_rows:
                        c = row.get("count", 0)
                        arrears_total = row.get("arrears_total", 0)
                        amount_total = row.get("amount_total", 0)
                        vals = [row["label"], fmt(c)]
                        if r.get("has_arrears"):
                            vals.append(fmt(arrears_total))
                            month_arrears += arrears_total
                        vals.append(fmt(amount_total))
                        month_amount += amount_total
                        month_count += c
                        section_rows.append(vals)
                    total_vals = ["Grand Total", fmt(month_count)]
                    if r.get("has_arrears"):
                        total_vals.append(fmt(month_arrears))
                        overall_arrears += month_arrears
                    total_vals.append(fmt(month_amount))
                    overall_amount += month_amount
                    overall_count += month_count
                    monthly_sections.append(
                        {
                            "month": fiscal_label_to_calendar_full_label(month_label),
                            "rows": section_rows,
                            "total": total_vals,
                        }
                    )
                overall_vals = ["Overall Total", fmt(overall_count)]
                if r.get("has_arrears"):
                    overall_vals.append(fmt(overall_arrears))
                overall_vals.append(fmt(overall_amount))

                # Apply column filtering to monthly_sections for PDF
                pdf_headers = ["Locality", "No. of Bills"]
                if r.get("has_arrears"):
                    pdf_headers.append("Arrears Received")
                pdf_headers.append("Amount Received")
                cols_param = request.args.get("cols")
                col_map = _get_card_col_map(card, r)
                if cols_param and col_map:
                    selected_keys = [k.strip() for k in cols_param.split(",") if k.strip()]
                    pdf_key_idx = {"locality": 0, "count": 1}
                    if r.get("has_arrears"):
                        pdf_key_idx["arrearsReceived"] = 2
                        pdf_key_idx["amountReceived"] = 3
                    else:
                        pdf_key_idx["amountReceived"] = 2
                    pdf_selected = [pdf_key_idx[k] for k in selected_keys if k in pdf_key_idx]
                    if pdf_selected and pdf_selected != list(range(len(pdf_headers))):
                        pdf_headers = [pdf_headers[i] for i in pdf_selected]
                        for section in monthly_sections:
                            section["rows"] = [[row[j] for j in pdf_selected] for row in section["rows"]]
                            section["total"] = [section["total"][0]] + [section["total"][j+1] for j in pdf_selected]
                        overall_vals = [overall_vals[0]] + [overall_vals[j+1] for j in pdf_selected]

                # Build PDF column widths dynamically
                n_cols = len(pdf_headers)
                page_w = A4[0] - 30 * mm
                if n_cols == 4:
                    pdf_col_widths = [page_w * 0.38, page_w * 0.20, page_w * 0.21, page_w * 0.21]
                elif n_cols == 3:
                    pdf_col_widths = [page_w * 0.48, page_w * 0.26, page_w * 0.26]
                elif n_cols == 2:
                    pdf_col_widths = [page_w * 0.62, page_w * 0.38]
                else:
                    pdf_col_widths = [page_w / n_cols] * n_cols

                pdf_bytes = generate_commercial_monthly_pdf(
                    summary,
                    monthly_sections,
                    overall_vals if monthly_sections else None,
                    r.get("has_arrears"),
                    pdf_headers=pdf_headers,
                    col_widths=pdf_col_widths,
                )
            elif card == "commercial-daily-income":
                page_w = landscape(A4)[0] - 18 * mm
                col_widths = [
                    page_w * 0.03,
                    page_w * 0.09,
                    page_w * 0.24,
                    page_w * 0.10,
                    page_w * 0.14,
                    page_w * 0.18,
                    page_w * 0.11,
                    page_w * 0.11,
                ]
                pdf_rows = wrap_pdf_table_cells(rows, font_size=10)
                pdf_kwargs = {
                    "pagesize": landscape(A4),
                    "col_widths": col_widths,
                    "first_col_left": False,
                    "left_cols": [],
                    "header_font_size": 11,
                    "body_font_size": 10,
                    "cell_padding": 8,
                }
            elif card == "connection-type-summary":
                page_w = A4[0] - 30 * mm
                col_widths = [page_w * 0.38, page_w * 0.20, page_w * 0.21, page_w * 0.21]
                pdf_kwargs = {
                    "pagesize": A4,
                    "col_widths": col_widths,
                    "first_col_left": True,
                    "left_cols": [0],
                }
            if card == "daily-staff-receive":
                ds_cols = request.args.get("cols", "")
                ds_detail_cols = request.args.get("detail_cols", "")
                pdf_bytes = generate_daily_staff_receive_pdf(r, sort_order=sort_order, summary_cols=ds_cols, detail_cols=ds_detail_cols)
            elif card == "commercial-monthly":
                pass
            else:
                # Attach connection summary under month-wise/monthly report when checkbox checked
                if card in ("receipt-monthly", "monthly") and conn_summary is not None:
                    pdf_kwargs["extra_section"] = conn_summary
                    pdf_kwargs["compact"] = True
                pdf_bytes = generate_card_pdf(title, summary, pdf_headers, pdf_rows, pdf_grand, **pdf_kwargs)
        dl_filename = {"commercial-month-wise": "Commercial_Month_Wise_Report"}.get(card, card)
        return Response(pdf_bytes, mimetype="application/pdf",
                        headers={"Content-Disposition": f"attachment; filename={dl_filename}.pdf"})
    elif fmt_type == "csv":
        if card == "daily-staff-receive":
            ds_cols = request.args.get("cols", "")
            ds_detail_cols = request.args.get("detail_cols", "")
            summary_headers, summary_rows, summary_grand, detail_headers, detail_rows = daily_staff_receive_export_tables(r, sort_order=sort_order)
            orig_sh = summary_headers
            orig_dh = detail_headers
            if ds_cols:
                summary_headers, summary_rows = parse_export_cols(ds_cols, DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP, summary_headers, summary_rows)
                if summary_grand:
                    _, g = parse_export_cols(ds_cols, DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP, orig_sh, [summary_grand])
                    summary_grand = g[0] if g else summary_grand
            if ds_detail_cols:
                detail_headers, detail_rows = parse_export_cols(ds_detail_cols, DAILY_STAFF_RECEIVE_DETAIL_COL_MAP, orig_dh, detail_rows)
            csv_rows = [["Summary", *summary_headers]]
            csv_rows.extend([["Summary", *row] for row in summary_rows])
            if summary_grand:
                csv_rows.append(["Summary", *summary_grand])
            csv_rows.append([])
            csv_rows.append(["Details", *detail_headers])
            csv_rows.extend([["Details", *row] for row in detail_rows])
            csv_data = "\n".join(",".join(f'"{str(value).replace(chr(34), chr(34) + chr(34))}"' for value in row) for row in csv_rows)
            return Response(csv_data, mimetype="text/csv",
                            headers={"Content-Disposition": f"attachment; filename={dl_filename}.csv"})
        all_rows = rows + [grand] if grand else rows
        if headers and headers[0] == "Month":
            out = io.StringIO()
            out.write('\ufeff')
            out.write(",".join('"' + h + '"' for h in headers) + "\n")
            for row_data in all_rows:
                parts = []
                for i, val in enumerate(row_data):
                    val_str = str(val)
                    if i == 0:
                        parts.append('="' + val_str + '"')
                    else:
                        if ',' in val_str or '"' in val_str:
                            parts.append('"' + val_str.replace('"', '""') + '"')
                        else:
                            parts.append(val_str)
                out.write(",".join(parts) + "\n")
            csv_data = out.getvalue()
            # Append connection summary to monthly CSV when included
            if card in ("receipt-monthly", "monthly") and conn_summary is not None:
                csv_data += "\n"
                csv_data += "Summary\n"
                csv_data += ",".join('"' + h + '"' for h in conn_summary["headers"]) + "\n"
                conn_all_rows = conn_summary["rows"] + [conn_summary["grand"]]
                for row_data in conn_all_rows:
                    parts = []
                    for val in row_data:
                        val_str = str(val)
                        if ',' in val_str or '"' in val_str:
                            parts.append('"' + val_str.replace('"', '""') + '"')
                        else:
                            parts.append(val_str)
                    csv_data += ",".join(parts) + "\n"
        else:
            df = _card_rows_to_df(headers, all_rows)
            csv_data = df.to_csv(index=False)
        return Response(csv_data, mimetype="text/csv",
                        headers={"Content-Disposition": f"attachment; filename={dl_filename}.csv"})
    elif fmt_type == "xlsx":
        if card == "daily-staff-receive":
            ds_cols = request.args.get("cols", "")
            ds_detail_cols = request.args.get("detail_cols", "")
            summary_headers, summary_rows, summary_grand, detail_headers, detail_rows = daily_staff_receive_export_tables(r, sort_order=sort_order)
            orig_sh = summary_headers
            orig_dh = detail_headers
            if ds_cols:
                summary_headers, summary_rows = parse_export_cols(ds_cols, DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP, summary_headers, summary_rows)
                if summary_grand:
                    _, g = parse_export_cols(ds_cols, DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP, orig_sh, [summary_grand])
                    summary_grand = g[0] if g else summary_grand
            if ds_detail_cols:
                detail_headers, detail_rows = parse_export_cols(ds_detail_cols, DAILY_STAFF_RECEIVE_DETAIL_COL_MAP, orig_dh, detail_rows)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                pd.DataFrame(summary_rows + ([summary_grand] if summary_grand else []), columns=summary_headers).to_excel(writer, sheet_name="Summary", index=False)
                pd.DataFrame(detail_rows, columns=detail_headers).to_excel(writer, sheet_name="Details", index=False)
            buf.seek(0)
            return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename={dl_filename}.xlsx"})
        buf.seek(0)
        return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename={dl_filename}.xlsx"})

    flash("Unknown format type.")
    return redirect(url_for("index"))


def generate_daily_staff_receive_summary_pdf(results: dict, sort_order: str = "default", cols: str = "") -> bytes:
    summary_headers, summary_rows, summary_grand, _, _ = daily_staff_receive_export_tables(results, sort_order=sort_order)

    # Apply column selection filtering - save original headers for grand total
    orig_summary_headers = summary_headers
    if cols:
        summary_headers, summary_rows = parse_export_cols(cols, DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP, summary_headers, summary_rows)
        if summary_grand:
            _, g = parse_export_cols(cols, DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP, orig_summary_headers, [summary_grand])
            summary_grand = g[0] if g else summary_grand

    report = results.get("daily_staff_receive") or {}
    buf = io.BytesIO()

    portrait_size = A4
    top_margin = 15 * mm
    bottom_margin = 15 * mm
    left_margin = 15 * mm
    right_margin = 15 * mm
    portrait_frame = Frame(left_margin, bottom_margin, portrait_size[0] - left_margin - right_margin, portrait_size[1] - top_margin - bottom_margin, id="portrait")
    doc = BaseDocTemplate(
        buf,
        pagesize=portrait_size,
        pageTemplates=[
            PageTemplate(id="portrait", frames=[portrait_frame], pagesize=portrait_size),
        ],
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("DailyStaffTitle", parent=styles["Heading1"], fontSize=19, textColor=ACCENT, alignment=1, spaceAfter=5 * mm, fontName="Helvetica-Bold")
    summary_style = ParagraphStyle("DailyStaffSummary", parent=styles["Normal"], fontSize=11, leading=15, spaceAfter=2 * mm)
    progress_style_small = ParagraphStyle(
        "DailyStaffProgressSmall",
        parent=styles["Normal"],
        fontSize=9,
        leading=13,
        spaceBefore=1 * mm,
        spaceAfter=3 * mm,
        textColor=colors.HexColor("#222222"),
    )

    elements = [
        Paragraph("Daily Receive Amount of Staff", title_style),
        Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}", summary_style),
    ]
    if report.get("date_range"):
        elements.append(Paragraph(f"<b>Report Date:</b> {report['date_range']}", summary_style))

    page_w = portrait_size[0] - left_margin - right_margin
    report_total_bills = report_total_metric = report_total_amount = 0
    for row in report.get("summary_rows") or []:
        report_total_bills += row.get("bills", 0)
        report_total_metric += row.get("metric_total", 0)
        report_total_amount += row.get("amount_total", 0)
    elements.append(Spacer(1, 3 * mm))
    elements.append(Paragraph(
        f"<b>Total Received Connections:</b> {fmt(report_total_bills)} &nbsp;&nbsp;&nbsp; "
        f"<b>Arrears Received:</b> {fmt(report_total_metric)} &nbsp;&nbsp;&nbsp; "
        f"<b>Total Amount Received:</b> {fmt(report_total_amount)}",
        progress_style_small,
    ))
    elements.append(Spacer(1, 5 * mm))

    num_rows = len(summary_rows) + (1 if summary_grand else 0)
    num_table_rows = 1 + num_rows  # header + data + grand total

    # Calculate available height for the table to fill the A4 page
    total_page_h = portrait_size[1]
    available_h = total_page_h - top_margin - bottom_margin
    # Header content uses ~85mm (title + dates + totals + spacers)
    header_content_h = 85 * mm
    table_target_h = available_h - header_content_h
    if table_target_h < 100 * mm:
        table_target_h = 100 * mm

    row_h = table_target_h / num_table_rows
    row_h = max(row_h, 12 * mm)
    row_h = min(row_h, 35 * mm)

    # Set explicit row heights on all rows
    row_heights = [row_h] * num_table_rows

    summary_data = [wrap_pdf_header_cells(summary_headers, font_size=10)] + wrap_pdf_body_cells(summary_rows, font_size=10, left_columns={1})
    if summary_grand:
        summary_data.append(wrap_pdf_body_cells([summary_grand], font_size=10, left_columns={1})[0])

    # Calculate column widths dynamically based on visible columns
    n_cols = len(summary_headers)
    # Find the index of the staff name column (key "staffName" = index 1 in full map)
    staff_name_idx = None
    for key, idx in DAILY_STAFF_RECEIVE_SUMMARY_COL_MAP.items():
        if key == "staffName" and idx < n_cols:
            staff_name_idx = idx
            break

    if n_cols == 5:
        col_widths = [page_w * 0.08, page_w * 0.27, page_w * 0.19, page_w * 0.22, page_w * 0.24]
    elif n_cols == 4:
        col_widths = [page_w * 0.10, page_w * 0.30, page_w * 0.25, page_w * 0.35]
    elif n_cols == 3:
        col_widths = [page_w * 0.12, page_w * 0.38, page_w * 0.50]
    elif n_cols == 2:
        col_widths = [page_w * 0.40, page_w * 0.60]
    elif n_cols == 1:
        col_widths = [page_w]
    else:
        col_widths = [page_w / n_cols] * n_cols

    t = Table(summary_data, colWidths=col_widths, repeatRows=1, rowHeights=row_heights)

    cell_pad = min(6, max(4, int((row_h - 10) / 2)))
    table_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
        ("TEXTCOLOR", (0, 0), (-1, 0), HEADER_FG),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_CLR),
        ("TOPPADDING", (0, 0), (-1, -1), cell_pad),
        ("BOTTOMPADDING", (0, 0), (-1, -1), cell_pad),
        ("LEFTPADDING", (0, 0), (-1, -1), max(3, cell_pad - 2)),
        ("RIGHTPADDING", (0, 0), (-1, -1), max(3, cell_pad - 2)),
    ]
    # Alternate row colors
    for i in range(1, num_table_rows):
        if i % 2 == 0:
            table_style_cmds.append(("BACKGROUND", (0, i), (-1, i), ALT_ROW))
    # Bold grand total row (last row)
    if num_table_rows > 2:
        last = num_table_rows - 1
        table_style_cmds.append(("FONTNAME", (0, last), (-1, last), "Helvetica-Bold"))
        table_style_cmds.append(("FONTSIZE", (0, last), (-1, last), 10))
        table_style_cmds.append(("BACKGROUND", (0, last), (-1, last), colors.HexColor("#e6d8c8")))
    # Bold any intermediate total rows
    for idx in range(1, num_table_rows):
        row_text = " ".join(str(cell) for cell in summary_data[idx])
        if " Total" in row_text or "Grand Total" in row_text:
            table_style_cmds.append(("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold"))
            table_style_cmds.append(("FONTSIZE", (0, idx), (-1, idx), 10))
            table_style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#e6d8c8")))
            table_style_cmds.append(("TEXTCOLOR", (0, idx), (-1, idx), colors.black))
    # Left-align staff name column when visible
    if staff_name_idx is not None:
        table_style_cmds.append(("ALIGN", (staff_name_idx, 1), (staff_name_idx, -1), "LEFT"))

    t.setStyle(TableStyle(table_style_cmds))
    elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


@app.route("/daily-staff-receive/export/summary-pdf")
def export_daily_staff_receive_summary_pdf():
    if not _last_daily_staff_results:
        flash("No daily staff report is available. Please upload files first.")
        return redirect(url_for("daily_staff_receive"))
    sort_order = request.args.get("sort", "default")
    if sort_order not in ("default", "asc", "desc"):
        sort_order = "default"
    cols = request.args.get("cols", "")
    pdf_bytes = generate_daily_staff_receive_summary_pdf(_last_daily_staff_results, sort_order=sort_order, cols=cols)
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=daily_staff_receive_summary.pdf"},
    )


@app.route("/daily-staff-receive/export/<fmt_type>")
def export_daily_staff_receive(fmt_type: str):
    if not _last_daily_staff_results:
        flash("No daily staff report is available. Please upload files first.")
        return redirect(url_for("daily_staff_receive"))
    cols = request.args.get("cols", "")
    detail_cols = request.args.get("detail_cols", "")
    sort_order = request.args.get("sort", "default")
    if sort_order not in ("default", "asc", "desc"):
        sort_order = "default"
    return daily_staff_receive_export_response(fmt_type, _last_daily_staff_results, cols=cols, detail_cols=detail_cols, sort_order=sort_order)


# ---------------------------------------------------------------------------
# Consumer Report — Sector-based connection summary from uploaded CSV
# ---------------------------------------------------------------------------

_consumer_report_data: list[dict] | None = None
_consumer_report_filename: str | None = None
_last_consumer_summary: dict | None = None
_last_connection_rate_report: dict | None = None
_connection_rate_report_filename: str | None = None


def _save_consumer_summary_cache(summary: dict, filename: str | None, total_rows: int = 0) -> None:
    """Persist the consumer summary to disk so it survives serverless cold starts.
    Vercel keeps /tmp between invocations of the same function instance."""
    try:
        os.makedirs(os.path.dirname(CONSUMER_REPORT_CACHE), exist_ok=True)
        payload = {
            "summary": summary,
            "filename": filename,
            "total_rows": total_rows,
        }
        with open(CONSUMER_REPORT_CACHE, "w") as f:
            json.dump(payload, f)
    except Exception:
        pass  # Non-fatal: in-memory state still works for same-instance requests


def _load_consumer_summary_cache() -> tuple[dict | None, str | None, int]:
    """Load a previously saved consumer summary from disk. Returns (summary, filename, total_rows)."""
    try:
        if os.path.exists(CONSUMER_REPORT_CACHE):
            with open(CONSUMER_REPORT_CACHE) as f:
                payload = json.load(f)
            return payload.get("summary"), payload.get("filename"), payload.get("total_rows", 0)
    except Exception:
        pass
    return None, None, 0


def _clear_consumer_summary_cache() -> None:
    try:
        if os.path.exists(CONSUMER_REPORT_CACHE):
            os.remove(CONSUMER_REPORT_CACHE)
    except Exception:
        pass

# Flexible column aliases: maps a canonical key to a list of normalised
# substrings that identify the column.  The first match wins.
_CONSUMER_COL_ALIASES: dict[str, list[str]] = {
    "serial": ["sr #", "sr#", "serial", "s no", "s.no", "serial no", "serial number", "sr no"],
    "consumer_name": ["consumer name", "consumer", "name"],
    "father_name": ["f/h name", "f h name", "father name", "father", "f/h", "guardian"],
    "mobile": ["mobile", "phone", "cell", "contact"],
    "sector": ["sector"],
    "locality": ["locality"],
    "address": ["address"],
    "order_number": ["order no", "order number", "register no", "order no. / register no", "register"],
    "rate_type": ["rate type", "rate"],
    "connection": ["connection no", "connection no.", "connection number", "connection"],
    "old_connection": ["old connection no", "old connection no.", "old connection number", "old connection"],
    "connection_status": ["status", "connection status", "conn status"],
    "consumer_status": ["consumer status"],
    "action": ["action", "actions"],
}


def _normalize_consumer_col(name: str) -> str:
    return " ".join(str(name).strip().lower().replace("_", " ").split())


def _resolve_consumer_columns(columns: list[str]) -> dict[str, str]:
    """Map each canonical key to the actual CSV column name that matched."""
    norm_cols = [_normalize_consumer_col(c) for c in columns]
    resolved: dict[str, str] = {}
    for key, aliases in _CONSUMER_COL_ALIASES.items():
        for alias in aliases:
            for i, nc in enumerate(norm_cols):
                if alias in nc and key not in resolved:
                    resolved[key] = columns[i]
                    break
            if key in resolved:
                break
    return resolved


def _classify_connection_status(value: str) -> str:
    """Classify both 'Status' column values and 'Consumer Status' into Active/Closed/Suspended."""
    text = str(value or "").strip().lower()
    # Suspended is displayed in its own Consumer Report column, but it still
    # remains non-active so it never receives an active-only budget.
    if text in ("suspended", "suspend"):
        return "Suspended"
    if text in ("closed", "c", "close", "inactive", "disconnected", "terminated", "in-active", "dead"):
        return "Closed"
    if text in ("active", "a", "open", "connected", "live", "running", "regular connection",
                "new connection", "regular", "new"):
        return "Active"
    return "Active"


def _canonical_consumer_sector_locality(sector: str, locality: str) -> tuple[str, str]:
    """Normalize known duplicate sector labels that refer to the same society."""
    sector_text = str(sector or "").strip()
    locality_text = str(locality or "").strip()
    normalized = _normalize_consumer_col(sector_text)
    compact = normalized.replace(" ", "")
    # Keep the official Zain City private society label consistent.
    if "zaincity" in compact and "private" in normalized:
        return "ZAIN CITY (PRIVATE SOCITIES)", "ZAIN CITY Zone A"
    # Keep the official Noor Mohalla sector/locality label consistent.  Similar
    # address text can appear elsewhere, but it must not become a second sector.
    if normalized == "noor mohalla melad chowk mian road":
        return "Noor Mohalla Melad Chowk Mian Road", "Noor Mohalla Zone A"
    return sector_text, locality_text


def _is_extra_zain_city_13g_sector(sector: str) -> bool:
    """Skip the extra Zain City CHACK NO 13/G block; official Zain City has 50 rows."""
    normalized = _normalize_consumer_col(sector)
    compact = normalized.replace(" ", "")
    return "zaincity" in compact and "13" in normalized and "private" not in normalized


def _is_extra_noor_mohalla_main_road_sector(sector: str) -> bool:
    """Skip duplicate Noor Mohalla sector text generated from address wording."""
    normalized = _normalize_consumer_col(sector)
    return normalized == "noor mohala millad chowk main road"


def _is_faulty_empty_consumer_sector(sector: str) -> bool:
    """Skip placeholder/faulty Empty sectors that do not represent real connections."""
    return _normalize_consumer_col(sector) == "empty"


def _normalize_rate_title(value: str) -> str:
    """Canonical key for rate matching; tolerates case/spacing drift without changing display text."""
    return " ".join(str(value or "").strip().split()).upper()


def _clean_rate_type_name(value: str) -> str:
    """Return only the rate type label for warnings, never a mixed/full CSV row."""
    text = " ".join(str(value or "").replace("\r", " ").replace("\n", " ").split()).strip(" ,")
    if not text or text.lower().replace(" ", "") == "ratetype":
        return ""
    return text[:120]


def _add_rate_alias(lookup: dict, alias: str, target: str) -> None:
    """Map known legacy consumer rate labels to the active rate title."""
    alias_key = _normalize_rate_title(alias)
    target_key = _normalize_rate_title(target)
    if alias_key and target_key and target_key in lookup and alias_key not in lookup:
        lookup[alias_key] = lookup[target_key]


def _parse_consumer_csv(file_storage) -> tuple[list[dict], list[str]]:
    """Read uploaded CSV/XLSX and return (rows, errors).
    Uses flexible column matching so files with varying header names work."""
    filename = secure_filename(file_storage.filename or "")
    ext = os.path.splitext(filename)[1].lower()
    try:
        if ext == ".csv":
            df = pd.read_csv(file_storage, dtype=str)
        else:
            df = pd.read_excel(file_storage, dtype=str)
    except Exception as exc:
        return [], [f"Could not read file: {exc}"]

    df = df.fillna("")
    raw_columns = list(df.columns)

    # Resolve canonical keys from whatever column names the file uses
    col_map = _resolve_consumer_columns(raw_columns)

    errors: list[str] = []
    for required in ("sector", "locality", "connection_status"):
        if required not in col_map:
            errors.append(f"Missing required column: looked for '{required}' (e.g. Sector, Locality, Status)")
    if errors:
        return [], errors

    def _get(row, key, default=""):
        csv_col = col_map.get(key)
        if csv_col and csv_col in row.index:
            return str(row[csv_col])
        return str(default)

    rows: list[dict] = []
    for _, row in df.iterrows():
        # Fix: joined exports can contain repeated header rows; skip them so
        # they do not become fake sectors or unmatched "Rate Type" warnings.
        if _normalize_consumer_col(_get(row, "sector", "")) == "sector" or \
                _normalize_consumer_col(_get(row, "rate_type", "")) == "rate type":
            continue

        # Suspended is stored in the Status column in consumer exports, while
        # Consumer Status may still say In-Active.  Read explicit Suspended
        # from Status first, then keep the existing Consumer Status preference.
        connection_status_val = _get(row, "connection_status", "")
        consumer_status_val = _get(row, "consumer_status", "")
        status_val = connection_status_val if _classify_connection_status(connection_status_val) == "Suspended" else consumer_status_val
        if not status_val.strip():
            status_val = connection_status_val
        rows.append({
            "serial_number": _get(row, "serial"),
            "consumer_name": _get(row, "consumer_name"),
            "father_name": _get(row, "father_name"),
            "mobile_number": _get(row, "mobile"),
            "sector": _get(row, "sector").strip(),
            "locality": _get(row, "locality").strip(),
            "address": _get(row, "address"),
            "order_number": _get(row, "order_number"),
            "rate_type": _get(row, "rate_type"),
            "connection": _get(row, "connection"),
            "old_connection": _get(row, "old_connection"),
            "connection_status": _classify_connection_status(status_val),
            "consumer_status": _get(row, "consumer_status"),
            "action": _get(row, "action"),
        })
    return rows, []


def _build_consumer_sector_summary(rows: list[dict]) -> dict:
    """Group rows by Sector, then Locality. Return summary data.

    SECTOR NORMALIZATION AND GROUPING:
    - Sector names are normalized (trimmed, lowercased, extra spaces removed)
      before grouping so that variations in spacing, capitalization, or minor
      spelling differences are merged into a single sector row.
    - All localities under the same normalized sector are combined into a
      comma-separated string in the Locality column.
    - One row is created per unique normalized sector.

    RATE IMPORT:
    - Rates are loaded from the bundled rates.json file (a fixed, version-
      controlled data file derived from rates.csv).  This file ships with
      the app and works on Vercel without requiring a CSV upload.
    - Fallback: if rates.json is missing, rates.csv is loaded instead
      (local development convenience).

    RATE MATCHING:
    - Each consumer row has a "Rate Type" field (e.g. "DOMESTIC (NEW CONNECTION)").
    - This is matched against "Rate Title" in rates.json using an exact
      string comparison (trimmed of whitespace).

    RATE NORMALISATION:
    - Water Rate is parsed as a float with comma removal.
    - Billing Period is lowercased for keyword matching.

    ANNUAL BUDGET CALCULATION:
    - Monthly   → Water Rate × 12
    - Quarterly → Water Rate × 4
    - Six Month → Water Rate × 2
    - Yearly    → Water Rate × 1

    BUDGET RULE: Only Active consumers/connections get budget calculated.
    Closed/inactive/disconnected records contribute 0 to the budget.
    """
    from collections import OrderedDict

    def _normalize_sector(name: str) -> str:
        """Normalize sector name for grouping: trim, lowercase, collapse spaces."""
        return " ".join(str(name or "").strip().lower().split())

    # --- Build rate lookup from bundled rates.json ---
    # Returns a dict: { "Rate Title (normalised)": {"period": str, "rate": float} }
    # Only the first entry per Rate Title is used (deduplicates by title).
    def _build_rate_lookup() -> dict:
        rates = _load_rates_csv()
        lookup = {}
        for r in rates:
            title = (r.get("Rate Title") or "").strip()
            title_key = _normalize_rate_title(title)
            period = (r.get("Billing Period") or "").strip().lower()
            rate_str = (r.get("Water Rate (Rs.)") or "0").replace(",", "").strip()
            try:
                rate = float(rate_str)
            except (ValueError, TypeError):
                rate = 0
            if title_key and rate > 0 and title_key not in lookup:
                lookup[title_key] = {"period": period, "rate": rate, "title": title}
        _add_rate_alias(lookup, 'PRIVATE SOCIETY(3" DIA)NEW CONNECTION',
                        '3" Dia Connection For Private Socities (New Connection)')
        return lookup

    # --- Annual budget multiplier based on Billing Period ---
    def _calc_annual_budget(rate: float, period: str) -> float:
        p = period.lower()
        if "six" in p or "half" in p or "semi" in p:
            return rate * 2
        if "quarter" in p:
            return rate * 4
        if "year" in p or "annual" in p:
            return rate
        if "monthly" in p or "month" in p:
            return rate * 12
        return rate * 12  # Default: treat as monthly

    def _domestic_annual_rate_override(rate_type: str, sector: str, locality: str) -> float | None:
        """Use approved annual domestic tariffs instead of inconsistent period labels."""
        combined = _normalize_rate_title(" ".join([rate_type or "", sector or "", locality or ""]))
        rate_key = _normalize_rate_title(rate_type)
        if "COMMERCIAL" in rate_key or "COMERCIAL" in rate_key:
            return None
        if "PRIVATE" in combined and ("SOCIETY" in combined or "SOCIETIES" in combined or "SOCITIES" in combined or "SOCITES" in combined or "SOCITY" in combined):
            return 9600.0
        if "DOMESTIC" in rate_key or "MONTH JAN26 TO JUNE26" in rate_key:
            return 4800.0
        return None

    rate_lookup = _build_rate_lookup()

    # sector_map: normalized_key → { original, localities_set, closed, active, budget }
    sector_map: OrderedDict[str, dict] = OrderedDict()
    commercial_locality_map: OrderedDict[str, dict] = OrderedDict()
    unmatched_rate_types: list[str] = []
    unmatched_budget_count = 0

    for row in rows:
        sector_raw = (row.get("sector") or "Unspecified").strip()
        locality_raw = (row.get("locality") or "Unspecified").strip()
        if (
            _is_faulty_empty_consumer_sector(sector_raw)
            or _is_extra_zain_city_13g_sector(sector_raw)
            or _is_extra_noor_mohalla_main_road_sector(sector_raw)
        ):
            continue
        sector_raw, locality_raw = _canonical_consumer_sector_locality(sector_raw, locality_raw)
        status = row.get("connection_status", "Active")
        rate_type = (row.get("rate_type") or "").strip()

        # Fix: protect this summary builder from repeated CSV header rows too,
        # because cached/raw rows can reach this path outside _parse_consumer_csv.
        if _normalize_consumer_col(sector_raw) == "sector" or _normalize_consumer_col(rate_type) == "rate type":
            continue

        norm_key = _normalize_sector(sector_raw)

        if norm_key not in sector_map:
            sector_map[norm_key] = {
                "original": sector_raw,
                "localities": [],  # list of unique locality names
                "closed": 0,
                "suspended": 0,
                "active": 0,
                "budget": 0.0,
                "annual_rate": 0.0,
            }

        # Keep the shortest original sector name for cleaner display
        if len(sector_raw) < len(sector_map[norm_key]["original"]):
            sector_map[norm_key]["original"] = sector_raw

        # -------------------------------------------------------------------
        # LOCALITY MERGING — CLEAN VERSION
        # Keeps locality strings short and clean by:
        # 1. Skipping values that are substrings of existing ones (keeps longer)
        # 2. Replacing existing values that are substrings of the new one
        # 3. Capping at 3 entries to prevent overly long Locality column
        # If no clean locality is found, falls back to the sector name later.
        # -------------------------------------------------------------------
        if locality_raw and len(locality_raw) <= 40:
            existing = sector_map[norm_key]["localities"]
            norm_new = locality_raw.lower()
            # Check if new value is a substring of any existing one (skip it)
            is_substring = any(norm_new in el for el in [l.lower() for l in existing])
            if not is_substring:
                # Remove existing values that are substrings of the new one
                existing = [l for l in existing if norm_new not in l.lower()]
                # Cap at 3 entries
                if len(existing) < 3:
                    existing.append(locality_raw)
                sector_map[norm_key]["localities"] = existing

        # Fix: calculate the per-row annual budget once from the consumer
        # Rate Type -> rate-list Rate Title match, and apply it only to Active rows.
        row_budget = 0.0
        annual_rate = 0.0
        clean_rate_type = _clean_rate_type_name(rate_type)
        rate_key = _normalize_rate_title(rate_type)
        if status == "Active" and rate_key and rate_key in rate_lookup:
            rl = rate_lookup[rate_key]
            annual_rate = _domestic_annual_rate_override(rate_type, sector_raw, locality_raw)
            if annual_rate is None:
                annual_rate = _calc_annual_budget(rl["rate"], rl["period"])
            row_budget = annual_rate
        elif status == "Active" and clean_rate_type:
            if clean_rate_type not in unmatched_rate_types:
                unmatched_rate_types.append(clean_rate_type)
            unmatched_budget_count += 1

        if status == "Suspended":
            sector_map[norm_key]["suspended"] += 1
        elif status == "Closed":
            sector_map[norm_key]["closed"] += 1
        else:
            sector_map[norm_key]["active"] += 1
        sector_map[norm_key]["budget"] += row_budget

        # Rate display must show the annual tariff, never an average derived
        # from Budget/counts. Domestic rows have one tariff per grouped sector.
        if status == "Active" and annual_rate:
            sector_map[norm_key]["annual_rate"] = annual_rate

        # Fix: commercial records also keep a locality-level map so they never
        # collapse into one COMMERCIAL row in the Commercial tab or exports.
        if sector_raw.upper().startswith("COMMERCIAL"):
            # Include annual tariff in the key so mixed-diameter/category
            # commercial records are split instead of averaged into one row.
            loc_key = norm_key + "|||" + _normalize_sector(locality_raw) + "|||" + str(int(annual_rate or 0))
            if loc_key not in commercial_locality_map:
                commercial_locality_map[loc_key] = {
                    "sector": sector_raw,
                    "locality": locality_raw,
                    "closed": 0,
                    "suspended": 0,
                    "active": 0,
                    "budget": 0.0,
                    "annual_rate": annual_rate,
                }
            if status == "Suspended":
                commercial_locality_map[loc_key]["suspended"] += 1
            elif status == "Closed":
                commercial_locality_map[loc_key]["closed"] += 1
            else:
                commercial_locality_map[loc_key]["active"] += 1
            commercial_locality_map[loc_key]["budget"] += row_budget
            if status == "Active" and annual_rate:
                commercial_locality_map[loc_key]["annual_rate"] = annual_rate

    # Build summary rows — one row per normalized sector
    summary_rows: list[dict] = []
    serial = 0
    sector_totals: dict[str, dict] = {}
    grand_closed, grand_suspended, grand_active, grand_budget = 0, 0, 0, 0.0

    for norm_key, s_data in sector_map.items():
        serial += 1
        closed = s_data["closed"]
        suspended = s_data.get("suspended", 0)
        active = s_data["active"]
        total = closed + suspended + active
        budget = s_data["budget"]
        original_sector = s_data["original"]
        # Combine localities into comma-separated string.
        # Fallback: if no locality found, use the sector name itself.
        locality_str = ", ".join(s_data["localities"]) if s_data["localities"] else original_sector

        summary_rows.append({
            "serial": serial,
            "sector": original_sector,
            "locality": locality_str,
            "closed": closed,
            "suspended": suspended,
            "active": active,
            "total": total,
            "budget": budget,
            "rate": s_data.get("annual_rate", 0),
        })
        sector_totals[original_sector] = {"closed": closed, "suspended": suspended, "active": active, "total": total, "budget": budget}
        grand_closed += closed
        grand_suspended += suspended
        grand_active += active
        grand_budget += budget

    # Fix: expose commercial locality rows to the tab and export pipeline so
    # commercial localities remain separate after server-side uploads/caches.
    commercial_detailed_rows: list[dict] = []
    commercial_closed = commercial_suspended = commercial_active = 0
    commercial_budget = 0.0
    for i, c_data in enumerate(commercial_locality_map.values(), 1):
        c_closed = c_data["closed"]
        c_suspended = c_data.get("suspended", 0)
        c_active = c_data["active"]
        c_budget = c_data["budget"]
        commercial_detailed_rows.append({
            "serial": i,
            "sector": c_data["sector"],
            "locality": c_data["locality"],
            "closed": c_closed,
            "suspended": c_suspended,
            "active": c_active,
            "total": c_closed + c_suspended + c_active,
            "budget": c_budget,
            "rate": c_data.get("annual_rate", 0),
        })
        commercial_closed += c_closed
        commercial_suspended += c_suspended
        commercial_active += c_active
        commercial_budget += c_budget

    return {
        "summary_rows": summary_rows,
        "sector_totals": sector_totals,
        "grand_total": {
            "closed": grand_closed,
            "suspended": grand_suspended,
            "active": grand_active,
            "total": grand_closed + grand_suspended + grand_active,
            "budget": grand_budget,
        },
        "sector_count": len(sector_map),
        "locality_count": len(summary_rows),
        "total_connections": grand_closed + grand_suspended + grand_active,
        "total_budget": grand_budget,
        "commercial_detailed_rows": commercial_detailed_rows,
        "commercial_grand_total": {
            "closed": commercial_closed,
            "suspended": commercial_suspended,
            "active": commercial_active,
            "total": commercial_closed + commercial_suspended + commercial_active,
            "budget": commercial_budget,
        },
        "unmatched_rate_types": unmatched_rate_types,
        "unmatched_budget_count": unmatched_budget_count,
    }


# ---------------------------------------------------------------------------
# SPLIT SUMMARY: Separate COMMERCIAL rows from normal (non-commercial) rows.
# Returns (normal_summary, commercial_summary) — each a dict compatible with
# the consumer_report.html template.  Used by the GET handler to render two
# sub-tabs: one for normal sectors, one for COMMERCIAL.
# ---------------------------------------------------------------------------
def _sort_summary_rows(rows: list[dict], sort_priority: str, sort_order: str) -> list[dict]:
    """Shared sorting for the Consumer Sector Report (preview, PDF, CSV, Excel).

    Applies the user's PRIORITY (active_first | closed_first) and ORDER
    (desc = Highest→Lowest, asc = Lowest→Highest) to EVERY row using its own
    count — never a sector-averaged value.  sector then locality are stable
    secondary keys so the result is deterministic.

    Commercial locality rows stay separate (never merged); they are simply
    reordered by their individual counts.  Returns a NEW list with serial
    numbers re-assigned in the final sorted order.
    """
    if sort_priority not in ("active_first", "closed_first"):
        sort_priority = "active_first"
    if sort_order not in ("desc", "asc"):
        sort_order = "desc"

    sort_key_field = "active" if sort_priority == "active_first" else "closed"
    primary_sign = -1 if sort_order == "desc" else 1

    def _key(r):
        return (primary_sign * (r.get(sort_key_field) or 0),
                (r.get("sector") or ""),
                (r.get("locality") or ""))

    final = sorted(rows, key=_key)
    for i, r in enumerate(final, 1):
        r["serial"] = i
    return final


def _filter_active_rows(summary: dict) -> dict:
    """Return a copy of `summary` with all rows having zero active
    connections removed.  Applies to both `summary_rows` and
    `commercial_detailed_rows`, and recomputes the affected grand totals.

    This keeps the preview tables, checkbox selection, grand totals, and all
    export formats consistent: only rows with Active > 0 are ever shown.
    """
    if not summary:
        return summary
    out = dict(summary)

    def _keep(rows):
        return [r for r in (rows or []) if (r.get("active") or 0) > 0]

    out["summary_rows"] = _keep(summary.get("summary_rows", []))
    out["commercial_detailed_rows"] = _keep(summary.get("commercial_detailed_rows", []))

    # Recompute grand totals strictly from surviving rows.
    gt = summary.get("grand_total", {})
    out["grand_total"] = {
        "closed": sum(r.get("closed", 0) for r in out["summary_rows"]),
        "suspended": sum(r.get("suspended", 0) for r in out["summary_rows"]),
        "active": sum(r.get("active", 0) for r in out["summary_rows"]),
        "total": sum(r.get("total", 0) for r in out["summary_rows"]),
        "budget": sum(r.get("budget", 0) for r in out["summary_rows"]),
    }
    if "commercial_grand_total" in summary:
        cdr = out["commercial_detailed_rows"]
        out["commercial_grand_total"] = {
            "closed": sum(r.get("closed", 0) for r in cdr),
            "suspended": sum(r.get("suspended", 0) for r in cdr),
            "active": sum(r.get("active", 0) for r in cdr),
            "total": sum(r.get("total", 0) for r in cdr),
            "budget": sum(r.get("budget", 0) for r in cdr),
        }
    out["total_connections"] = out["grand_total"]["total"]
    out["total_budget"] = out["grand_total"]["budget"]
    out["sector_count"] = len(out["summary_rows"])
    out["locality_count"] = len(out["summary_rows"])
    return out


def _split_summary_by_type(summary: dict) -> tuple[dict, dict]:
    """Split a summary dict into (normal_summary, commercial_summary).

    COMMERCIAL sectors are those whose sector name starts with 'COMMERCIAL'
    (case-insensitive).  All other sectors go into the normal summary.
    Each sub-summary gets its own serial numbers (re-numbered from 1),
    grand totals, and metadata counts.
    """
    if not summary:
        return ({}, {})

    # Exclude rows with zero active connections from the displayed/split view.
    summary = _filter_active_rows(summary)

    all_rows = summary.get("summary_rows", [])

    # --- Partition rows by type ---
    normal_rows: list[dict] = []
    commercial_rows: list[dict] = []
    for row in all_rows:
        sector_name = (row.get("sector") or "").strip()
        if sector_name.upper().startswith("COMMERCIAL"):
            commercial_rows.append(row)
        else:
            normal_rows.append(row)

    # --- Build sub-summary helper ---
    def _build_sub(rows: list[dict], label: str) -> dict:
        serial = 0
        grand_closed, grand_suspended, grand_active, grand_budget = 0, 0, 0, 0
        sector_totals: dict[str, dict] = {}
        result_rows: list[dict] = []
        for r in rows:
            serial += 1
            budget = r.get("budget", 0)
            suspended = r.get("suspended", 0)
            result_rows.append({
                "serial": serial,
                "sector": r["sector"],
                "locality": r["locality"],
                "closed": r["closed"],
                "suspended": suspended,
                "active": r["active"],
                "total": r["total"],
                "budget": budget,
                "rate": r.get("rate", 0),
            })
            sector_totals[r["sector"]] = {
                "closed": r["closed"],
                "suspended": suspended,
                "active": r["active"],
                "total": r["total"],
                "budget": budget,
            }
            grand_closed += r["closed"]
            grand_suspended += suspended
            grand_active += r["active"]
            grand_budget += budget
        return {
            "summary_rows": result_rows,
            "sector_totals": sector_totals,
            "grand_total": {
                "closed": grand_closed,
                "suspended": grand_suspended,
                "active": grand_active,
                "total": grand_closed + grand_suspended + grand_active,
                "budget": grand_budget,
            },
            "sector_count": len(rows),
            "locality_count": len(rows),
            "total_connections": grand_closed + grand_suspended + grand_active,
            "total_budget": grand_budget,
        }

    normal = _build_sub(normal_rows, "normal") if normal_rows else {}

    # --- Commercial: prefer locality-level rows if available ---
    # The client sends commercial_detailed_rows (one row per locality) which
    # preserves individual locality names.  Fall back to merged rows if absent.
    commercial_detail_rows = summary.get("commercial_detailed_rows", [])
    if commercial_detail_rows:
        # Re-number serials and build summary from locality-level rows
        for i, r in enumerate(commercial_detail_rows, 1):
            r["serial"] = i
        commercial = _build_sub(commercial_detail_rows, "commercial")
        # Use client-provided grand total if available (more accurate)
        client_gt = summary.get("commercial_grand_total")
        if client_gt:
            commercial["grand_total"] = client_gt
            commercial["total_connections"] = client_gt.get("total", 0)
            commercial["total_budget"] = client_gt.get("budget", 0)
    else:
        commercial = _build_sub(commercial_rows, "commercial") if commercial_rows else {}
    return (normal, commercial)


def _load_rates_csv() -> list[dict]:
    """Load rate data from the provided rates CSV or bundled rates.json.

    RATE SOURCE PRIORITY:
    1. User-provided rates CSV at USER_RATES_CSV_PATH (e.g. C:\\Users\\Rising\\Downloads\\rates.csv)
    2. Bundled rates.json (version-controlled, ships with app for Vercel)
    3. Bundled rates.csv (local development fallback)

    Each dict has keys: Rate Title, Connection Type, Billing Period,
    Water Rate (Rs.), Status — matching the CSV headers the client expects.

    RATE MATCHING:
    - Consumer's "Rate Type" field is matched against "Rate Title" using
      exact string comparison (trimmed of whitespace).
    - Only Active connections contribute to budget calculation.

    ANNUAL BUDGET CALCULATION:
    - Monthly   -> Water Rate x 12
    - Quarterly -> Water Rate x 4
    - Six Month -> Water Rate x 2
    - Yearly    -> Water Rate x 1
    """
    import json as _json
    import csv as _csv

    # --- Priority 1: User-provided rates CSV ---
    # This is the main source for rate values when available.
    user_rates_path = r"C:\Users\Rising\Downloads\rates.csv"
    if os.path.exists(user_rates_path):
        try:
            with open(user_rates_path, "r", encoding="utf-8-sig") as f:
                reader = _csv.DictReader(f)
                rows = []
                for row in reader:
                    title = (row.get("Rate Title") or "").strip()
                    if not title:
                        continue
                    # Normalize Water Rate: remove commas, strip whitespace
                    rate_str = (row.get("Water Rate (Rs.)") or "0").replace(",", "").strip()
                    try:
                        rate_val = float(rate_str)
                    except (ValueError, TypeError):
                        rate_val = 0
                    row["Water Rate (Rs.)"] = str(int(rate_val)) if rate_val == int(rate_val) else str(rate_val)
                    rows.append(row)
                if rows:
                    return rows
        except Exception:
            pass  # Fall through to bundled sources

    # --- Priority 2: Bundled rates.json (Vercel deployments) ---
    json_path = os.path.join(os.path.dirname(__file__), "rates.json")
    if os.path.exists(json_path):
        with open(json_path, "r", encoding="utf-8") as f:
            data = _json.load(f)
        # Ensure numeric Water Rate values (JSON stores them as numbers,
        # but normalise to strings for backward compatibility with client code)
        for row in data:
            rate_val = row.get("Water Rate (Rs.)", 0)
            if isinstance(rate_val, (int, float)):
                row["Water Rate (Rs.)"] = str(int(rate_val)) if rate_val == int(rate_val) else str(rate_val)
        return [row for row in data if (row.get("Rate Title") or "").strip()]

    # --- Priority 3: Bundled rates.csv (local development fallback) ---
    rates_path = os.path.join(os.path.dirname(__file__), "rates.csv")
    if not os.path.exists(rates_path):
        return []
    with open(rates_path, "r", encoding="utf-8-sig") as f:
        reader = _csv.DictReader(f)
        return [row for row in reader if (row.get("Rate Title") or "").strip()]


CONNECTION_RATE_CATEGORIES = [
    ("domestic", "Domestic", 300),
    ("domestic_private_societies", "Domestic private societies", 800),
    ("bank_hamam_tea_stall_samosa", "BANK/HAMAM/TEA STALL/SAMOSA SHOP ETC", 500),
    ("college_schools", "College And Schools", 500),
    ("petrol_pump", "Petrol Pump", 500),
    ("private_hospital", "Private Hospital", 500),
    ("hotel_marriage_sweet_bakery", "HOTEL/ MARRIAGE HALL/ SWEET SHOP /BAKERY", 1000),
    ("park", "Park", 1000),
    ("service_station", "Service Station", 2500),
    ("factory_commercial", "Factory Commercial", 3000),
]


def _norm_report_text(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", str(value or "").upper()).strip()


def _connection_rate_category(row: dict) -> str:
    sector = row.get("sector", "")
    locality = row.get("locality", "")
    rate_type = row.get("rate_type", "")
    text = _norm_report_text(" ".join([sector, locality, rate_type]))
    rate_key = _normalize_rate_title(rate_type)
    if "PRIVATE" in text and any(word in text for word in ("SOCIETY", "SOCIETIES", "SOCITIES", "SOCITES", "SOCITY")):
        return "domestic_private_societies"
    if "COMMERCIAL" not in rate_key and "COMERCIAL" not in rate_key and "COMMERCIAL" not in text and "COMERCIAL" not in text:
        return "domestic"
    if any(word in text for word in ("SCHOOL", "COLLEGE")):
        return "college_schools"
    if "PETROL" in text:
        return "petrol_pump"
    if "HOSPITAL" in text:
        return "private_hospital"
    if any(word in text for word in ("HOTEL", "MARRIAGE", "SWEET", "BAKERY", "BAKERIES")):
        return "hotel_marriage_sweet_bakery"
    if "PARK" in text:
        return "park"
    if "SERVICE" in text and "STATION" in text:
        return "service_station"
    if "FACTORY" in text:
        return "factory_commercial"
    return "bank_hamam_tea_stall_samosa"


def _build_connection_rate_report(rows: list[dict]) -> dict:
    counts = {key: 0 for key, _, _ in CONNECTION_RATE_CATEGORIES}
    for row in rows:
        if row.get("connection_status", "Active") != "Active":
            continue
        counts[_connection_rate_category(row)] += 1
    report_rows = []
    for idx, (key, description, rate) in enumerate(CONNECTION_RATE_CATEGORIES, start=1):
        count = counts.get(key, 0)
        report_rows.append({
            "sr": idx,
            "key": key,
            "description": description,
            "connections": count,
            "rate": rate,
            "total": count * rate * 12,
        })
    return {
        "rows": report_rows,
        "grand_total": {
            "connections": sum(r["connections"] for r in report_rows),
            "total": sum(r["total"] for r in report_rows),
        },
    }


def _connection_rate_rows_from_payload(payload: dict) -> list[dict]:
    rows = []
    for idx, row in enumerate(payload.get("rows", []), start=1):
        connections = int(parse_number(row.get("connections", 0)))
        rate = float(parse_number(row.get("rate", 0)))
        rows.append({
            "sr": idx,
            "description": str(row.get("description", "")).strip(),
            "connections": connections,
            "rate": rate,
            "total": connections * rate * 12,
        })
    return rows


@app.route("/connection-rate-report", methods=["GET", "POST"])
def connection_rate_report():
    global _last_connection_rate_report, _connection_rate_report_filename
    if request.method == "POST":
        if request.form.get("action") == "clear":
            _last_connection_rate_report = None
            _connection_rate_report_filename = None
            return redirect(url_for("connection_rate_report"))
        file = request.files.get("consumer_file")
        if not file or not file.filename:
            flash("Please choose a consumer CSV or Excel file first.")
            return redirect(url_for("connection_rate_report"))
        rows, errors = _parse_consumer_csv(file)
        if errors:
            flash("Upload failed: " + "; ".join(errors))
            return redirect(url_for("connection_rate_report"))
        _last_connection_rate_report = _build_connection_rate_report(rows)
        _connection_rate_report_filename = secure_filename(file.filename or "upload.csv")
        flash("Connection rate report generated.")
        return redirect(url_for("connection_rate_report"))
    return render_template(
        "connection_rate_report.html",
        report=_last_connection_rate_report,
        filename=_connection_rate_report_filename,
        active_page="connection_rate_report",
    )


@app.route("/connection-rate-report/export/<fmt_type>", methods=["POST"])
def export_connection_rate_report(fmt_type: str):
    payload = request.get_json(silent=True) or {}
    rows = _connection_rate_rows_from_payload(payload)
    if not rows:
        flash("No connection rate report data available.")
        return redirect(url_for("connection_rate_report"))

    headers = ["Sr No.", "Description", "No of Connections", "Rate per Connection", "Total amount"]
    data_rows = [[r["sr"], r["description"], r["connections"], fmt(r["rate"]), fmt(r["total"])] for r in rows]
    total_row = ["", "Total", sum(r["connections"] for r in rows), "", fmt(sum(r["total"] for r in rows))]

    if fmt_type == "csv":
        df = pd.DataFrame(data_rows + [total_row], columns=headers)
        return Response(df.to_csv(index=False), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=connection_rate_report.csv"})
    if fmt_type == "xlsx":
        buf = io.BytesIO()
        pd.DataFrame(data_rows + [total_row], columns=headers).to_excel(buf, index=False)
        buf.seek(0)
        return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=connection_rate_report.xlsx"})
    if fmt_type == "pdf":
        pdf_bytes = generate_summary_pdf("Connection Rate Report", headers, data_rows, total_row, show_summary=True)
        return Response(pdf_bytes, mimetype="application/pdf", headers={"Content-Disposition": "attachment; filename=connection_rate_report.pdf"})
    flash("Unknown export format.")
    return redirect(url_for("connection_rate_report"))


@app.route("/consumer-report", methods=["GET", "POST"])
def consumer_report():
    global _consumer_report_data, _consumer_report_filename, _last_consumer_summary

    if request.method == "POST":
        action = request.form.get("action", "")

        # -------------------------------------------------------------------
        # JSON upload: client-side CSV parser sends pre-computed sector summary
        # as JSON to bypass Vercel's 4.5MB serverless body limit.  The payload
        # contains summary_rows, sector_totals, grand_total, and metadata.
        # -------------------------------------------------------------------
        if request.is_json:
            data = request.get_json(silent=True)
            if not data or "summary_rows" not in data:
                if is_ajax():
                    return ajax_error("Invalid data payload.")
                flash("Invalid upload data.")
                return redirect(url_for("consumer_report"))

            # Store the pre-computed summary directly for export routes
            _consumer_report_data = data.get("summary_rows", [])
            _consumer_report_filename = data.get("filename", "upload.csv")

            # Build a summary dict compatible with the template and exports.
            # Include commercial_detailed_rows (locality-level) for commercial PDF export.
            summary = {
                "summary_rows": data.get("summary_rows", []),
                "sector_totals": data.get("sector_totals", {}),
                "grand_total": data.get("grand_total", {"closed": 0, "suspended": 0, "active": 0, "total": 0, "budget": 0}),
                "sector_count": data.get("sector_count", 0),
                "locality_count": data.get("locality_count", 0),
                "total_connections": data.get("total_connections", 0),
                "total_budget": data.get("total_budget", 0),
                "commercial_detailed_rows": data.get("commercial_detailed_rows", []),
                "commercial_grand_total": data.get("commercial_grand_total", {"closed": 0, "suspended": 0, "active": 0, "total": 0, "budget": 0}),
                "unmatched_rate_types": data.get("unmatched_rate_types", []),
                "unmatched_budget_count": data.get("unmatched_budget_count", 0),
            }
            # Exclude rows with zero active connections from the stored summary
            # so every downstream view/export is consistent.
            summary = _filter_active_rows(summary)
            _last_consumer_summary = summary
            # Persist to disk so the data survives Vercel serverless cold starts
            _save_consumer_summary_cache(summary, _consumer_report_filename, data.get("total_rows", 0))

            msg = f"File uploaded. Found {summary['total_connections']:,} connections across {summary['sector_count']} sectors."
            if is_ajax():
                return ajax_ok(message=msg, redirect_url=url_for("consumer_report"))
            flash(msg)
            normal_summary, commercial_summary = _split_summary_by_type(summary)
            return render_template("consumer_report.html", summary=summary,
                                   normal_summary=normal_summary, commercial_summary=commercial_summary,
                                   filename=_consumer_report_filename, total_rows=data.get("total_rows", 0),
                                   rates_json=json.dumps(_load_rates_csv()))

        if action == "clear":
            _consumer_report_data = None
            _consumer_report_filename = None
            _last_consumer_summary = None
            _clear_consumer_summary_cache()
            flash("Consumer report data cleared.")
            return redirect(url_for("consumer_report"))

        if action == "upload":
            file = request.files.get("consumer_file")
            if not file or not file.filename:
                if is_ajax():
                    return ajax_error("Please choose a CSV or Excel file first.")
                flash("Please choose a CSV or Excel file first.")
                return redirect(url_for("consumer_report"))
            if not allowed_file(file.filename):
                if is_ajax():
                    return ajax_error(f"Unsupported file type: {file.filename}")
                flash(f"Unsupported file type: {file.filename}")
                return redirect(url_for("consumer_report"))

            rows, errors = _parse_consumer_csv(file)
            if errors:
                msg = "Upload failed: " + "; ".join(errors)
                if is_ajax():
                    return ajax_error(msg)
                flash(msg)
                return redirect(url_for("consumer_report"))

            _consumer_report_data = rows
            _consumer_report_filename = secure_filename(file.filename or "upload.csv")

            summary = _build_consumer_sector_summary(rows)
            # Fix: form uploads must update the same filtered cache as browser
            # JSON uploads; otherwise GET/export can reuse an older report.
            summary = _filter_active_rows(summary)
            _last_consumer_summary = summary
            _save_consumer_summary_cache(summary, _consumer_report_filename, len(rows))
            msg = f"File uploaded. Found {summary['total_connections']:,} connections across {summary['sector_count']} sectors."
            if is_ajax():
                return ajax_ok(message=msg, redirect_url=url_for("consumer_report"))
            flash(msg)
            normal_summary, commercial_summary = _split_summary_by_type(summary)
            return render_template("consumer_report.html", summary=summary,
                                   normal_summary=normal_summary, commercial_summary=commercial_summary,
                                   filename=_consumer_report_filename, total_rows=len(rows),
                                   rates_json=json.dumps(_load_rates_csv()))

    # GET — restore from cached summary (JSON upload) or raw data (form upload).
    # Load from disk cache first so data survives Vercel serverless cold starts.
    rates_json_str = json.dumps(_load_rates_csv())
    cached_summary, cached_filename, cached_rows = _load_consumer_summary_cache()
    if cached_summary:
        normal_summary, commercial_summary = _split_summary_by_type(cached_summary)
        return render_template("consumer_report.html", summary=cached_summary,
                               normal_summary=normal_summary, commercial_summary=commercial_summary,
                               filename=cached_filename, total_rows=cached_rows,
                               rates_json=rates_json_str)
    if _last_consumer_summary:
        normal_summary, commercial_summary = _split_summary_by_type(_last_consumer_summary)
        return render_template("consumer_report.html", summary=_last_consumer_summary,
                               normal_summary=normal_summary, commercial_summary=commercial_summary,
                               filename=_consumer_report_filename, total_rows=_last_consumer_summary.get("total_rows", 0),
                               rates_json=rates_json_str)
    if _consumer_report_data:
        summary = _build_consumer_sector_summary(_consumer_report_data)
        normal_summary, commercial_summary = _split_summary_by_type(summary)
        return render_template("consumer_report.html", summary=summary,
                               normal_summary=normal_summary, commercial_summary=commercial_summary,
                               filename=_consumer_report_filename, total_rows=len(_consumer_report_data),
                               rates_json=rates_json_str)
    return render_template("consumer_report.html", summary=None, normal_summary={}, commercial_summary={},
                           filename=None, total_rows=0, rates_json=rates_json_str)


@app.route("/consumer-report/export/<fmt_type>", methods=["GET", "POST"])
def export_consumer_report(fmt_type: str):
    global _last_consumer_summary, _consumer_report_filename, _consumer_report_data
    # -------------------------------------------------------------------------
    # Vercel serverless functions are stateless: the summary parsed in the browser
    # is POSTed back with each export request as `summary_data` (JSON string).
    # If present, use it directly. Otherwise fall back to disk cache / in-memory.
    # -------------------------------------------------------------------------
    summary = None
    if request.method == "POST":
        raw = request.form.get("summary_data") or ""
        if raw:
            try:
                payload = json.loads(raw)
                if isinstance(payload, dict) and "summary_rows" in payload:
                    summary = {
                        "summary_rows": payload.get("summary_rows", []),
                        "sector_totals": payload.get("sector_totals", {}),
                        "grand_total": payload.get("grand_total", {"closed": 0, "suspended": 0, "active": 0, "total": 0, "budget": 0}),
                        "sector_count": payload.get("sector_count", 0),
                        "locality_count": payload.get("locality_count", 0),
                        "total_connections": payload.get("total_connections", 0),
                        # Fix: keep export POST payloads lossless so commercial
                        # locality rows, budgets, and warnings survive downloads.
                        "total_budget": payload.get("total_budget", 0),
                        "commercial_detailed_rows": payload.get("commercial_detailed_rows", []),
                        "commercial_grand_total": payload.get("commercial_grand_total", {"closed": 0, "suspended": 0, "active": 0, "total": 0, "budget": 0}),
                        "unmatched_rate_types": payload.get("unmatched_rate_types", []),
                        "unmatched_budget_count": payload.get("unmatched_budget_count", 0),
                    }
                    _last_consumer_summary = summary
                    _consumer_report_filename = payload.get("filename", "upload.csv")
                    _consumer_report_data = payload.get("summary_rows", [])
            except Exception:
                summary = None

    # Fall back to disk cache so exports also work after a serverless cold start
    if summary is None:
        cached_summary, cached_filename, _ = _load_consumer_summary_cache()
        if cached_summary:
            summary = cached_summary
            _last_consumer_summary = cached_summary
            _consumer_report_filename = cached_filename

    if summary is None and _consumer_report_data is not None:
        summary = _build_consumer_sector_summary(_consumer_report_data)

    if summary is None:
        flash("No consumer report data available. Please upload a file first.")
        return redirect(url_for("consumer_report"))

    # -----------------------------------------------------------------------
    # UNIFIED EXPORT PIPELINE
    # All export formats (PDF, Commercial PDF, CSV, Excel) use ONE shared,
    # fully-filtered dataset so that row counts and grand totals match across
    # every output.  The pipeline applies, in order:
    #   1. Report scope (domestic vs commercial) based on fmt_type / tab param
    #   2. Active > 0 filter  (rows with 0 active connections are excluded)
    #   3. Sorting (priority + order)
    #   4. Row checkbox selection
    #   5. Column visibility (selected columns only)
    # Checkbox selection is applied client-side from the visible sorted table
    # and POSTed as `summary_data`; never rebuild exports from the full cache
    # when selected rows were provided.
    # -----------------------------------------------------------------------
    from collections import OrderedDict

    sort_priority = request.args.get("priority", "active_first")
    sort_order = request.args.get("order", "desc")
    if sort_priority not in ("active_first", "closed_first"):
        sort_priority = "active_first"
    if sort_order not in ("desc", "asc"):
        sort_order = "desc"

    tab_param = request.args.get("tab", "normal")
    if tab_param not in ("normal", "commercial"):
        tab_param = "normal"

    # `summary` is already resolved above (from POST data, cache, or raw data).
    # Do NOT rebuild here — `_consumer_report_data` may already be summary rows.
    if "summary_rows" not in summary:
        summary = _build_consumer_sector_summary(_consumer_report_data)

    # -----------------------------------------------------------------------
    # STEP 1 — Report scope: choose the base row set.
    #   - commercial-pdf OR tab=commercial  -> COMMERCIAL rows (locality-level)
    #   - everything else                   -> DOMESTIC (non-COMMERCIAL) rows
    # -----------------------------------------------------------------------
    if fmt_type == "commercial-pdf" or tab_param == "commercial":
        detailed = summary.get("commercial_detailed_rows", [])
        if detailed:
            base_rows = list(detailed)
        else:
            base_rows = [r for r in summary["summary_rows"] if r["sector"].upper().startswith("COMMERCIAL")]
        scope_label = "commercial"
    else:
        base_rows = [r for r in summary["summary_rows"] if not r["sector"].upper().startswith("COMMERCIAL")]
        scope_label = "domestic"

    # -----------------------------------------------------------------------
    # STEP 2 — Active > 0 filter.
    # A row is excluded when it has zero active connections.
    # -----------------------------------------------------------------------
    filtered_rows = [r for r in base_rows if (r.get("active") or 0) > 0]

    # -----------------------------------------------------------------------
    # STEP 3 — Shared sorting logic (used by preview, PDF, CSV, and Excel).
    # Delegated to _sort_summary_rows so the exact same ordering is guaranteed
    # across every output format.  See _sort_summary_rows for the full rules.
    # -----------------------------------------------------------------------
    final_sorted = _sort_summary_rows(filtered_rows, sort_priority, sort_order)

    # -----------------------------------------------------------------------
    # STEP 4 — Column visibility.
    # -----------------------------------------------------------------------
    cols_param = request.args.get("cols", "")
    if cols_param:
        selected_cols = [c.strip().lower() for c in cols_param.split(",") if c.strip()]
    else:
        selected_cols = ["sr", "sector", "locality", "rate", "closed", "suspended", "active", "total", "budget"]

    COL_DEFS = {
        "sr":     ("SR",     lambda r: r["serial"],             12),
        "sector": ("Sector", lambda r: r["sector"],             58),
        "locality": ("Locality", lambda r: r["locality"],       54),
        "rate":   ("Rate (Rs./Year)", lambda r: int(r.get("rate", 0)), 18),
        "closed": ("Closed", lambda r: r["closed"],             16),
        "suspended": ("Suspended", lambda r: r.get("suspended", 0), 18),
        "active": ("Active", lambda r: r["active"],             16),
        "total":  ("Total Connections", lambda r: r["total"],   18),
        "budget": ("Budget (Rs.)", lambda r: r.get("budget", 0), 22),
    }

    active_cols = [c for c in ["sr", "sector", "locality", "rate", "closed", "suspended", "active", "total", "budget"] if c in selected_cols]

    # -----------------------------------------------------------------------
    # Shared export dataset: headers, rows, and grand total computed ONLY from
    # the final filtered/sorted rows so CSV/Excel/PDF all agree.
    # -----------------------------------------------------------------------
    headers = [COL_DEFS[c][0] for c in active_cols]
    rows = [
        [COL_DEFS[c][1](r) for c in active_cols]
        for r in final_sorted
    ]

    # Grand total derived strictly from visible/filtered rows.
    gt_vals = {
        "closed": sum(r.get("closed", 0) for r in final_sorted),
        "suspended": sum(r.get("suspended", 0) for r in final_sorted),
        "active": sum(r.get("active", 0) for r in final_sorted),
        "total":  sum(r.get("total", 0) for r in final_sorted),
        "budget": sum(r.get("budget", 0) for r in final_sorted),
    }
    grand = []
    for c in active_cols:
        if c == "sr":
            grand.append("")
        elif c == "sector":
            grand.append("GRAND TOTAL")
        elif c == "locality":
            grand.append("")
        elif c == "rate":
            grand.append("")
        elif c == "budget":
            grand.append(gt_vals.get("budget", 0))
        else:
            grand.append(gt_vals.get(c, 0))

    filename = "commercial_sector_report" if scope_label == "commercial" else "consumer_sector_report"

    # -----------------------------------------------------------------------
    # PDF export — Professional A4 portrait print layout with light colours
    # Clean A4 portrait design with soft teal header, light alternating rows,
    # and a bold Grand Total row at the bottom. No sector subtotal rows.
    # -----------------------------------------------------------------------
    if fmt_type == "pdf":
        # -- Colour palette (light, print-friendly) --
        PDF_HEADER_BG = colors.HexColor("#e8f5f3")
        PDF_HEADER_FG = colors.HexColor("#1a5c52")
        PDF_ALT_ROW = colors.HexColor("#f7faf9")
        PDF_WHITE_ROW = colors.white
        PDF_GRAND_BG = colors.HexColor("#d4edda")
        PDF_GRAND_FG = colors.HexColor("#155724")
        # Darker grid lines keep the printed PDF visibly table-shaped.
        PDF_GRID = colors.HexColor("#6fa8a0")
        PDF_BODY_FG = colors.HexColor("#2c3e50")

        # -- Page setup: A4 portrait with tighter side margins --
        # Fix: Domestic/Consumer PDF uses the real printable width so the table
        # feels wider while still staying inside the A4 page.
        page_w, page_h = A4
        top_m = 14 * mm
        bottom_m = 12 * mm
        left_m = 8 * mm
        right_m = 8 * mm
        usable_w = page_w - left_m - right_m

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=left_m,
            rightMargin=right_m,
            topMargin=top_m,
            bottomMargin=bottom_m,
        )

        styles = getSampleStyleSheet()

        # -- Heading styles --
        report_title_style = ParagraphStyle(
            "ConsumerReportTitle",
            parent=styles["Heading1"],
            fontSize=18,
            fontName="Helvetica-Bold",
            textColor=PDF_HEADER_FG,
            alignment=1,
            spaceAfter=2 * mm,
        )
        report_subtitle_style = ParagraphStyle(
            "ConsumerReportSubtitle",
            parent=styles["Normal"],
            fontSize=10,
            fontName="Helvetica",
            textColor=colors.HexColor("#5a7a74"),
            alignment=1,
            spaceAfter=3 * mm,
        )
        meta_style = ParagraphStyle(
            "ConsumerReportMeta",
            parent=styles["Normal"],
            fontSize=9,
            fontName="Helvetica",
            textColor=colors.HexColor("#6c8a84"),
            alignment=1,
            spaceAfter=6 * mm,
        )

        # -- Style for wrapping Sector/Locality text inside cells --
        cell_wrap_style = ParagraphStyle(
            "CellWrap",
            parent=styles["Normal"],
            fontSize=8.2,
            fontName="Helvetica",
            textColor=PDF_BODY_FG,
            leading=10.8,
            alignment=0,
        )
        cell_center_style = ParagraphStyle(
            "CellCenter",
            parent=styles["Normal"],
            fontSize=8.2,
            fontName="Helvetica",
            textColor=PDF_BODY_FG,
            alignment=1,
            leading=10.8,
        )
        header_cell_style = ParagraphStyle(
            "HeaderCell",
            parent=styles["Normal"],
            fontSize=9,
            fontName="Helvetica-Bold",
            textColor=PDF_HEADER_FG,
            alignment=1,
        )

        elements = []

        # -- Report heading block --
        elements.append(Paragraph("Consumer Sector Report", report_title_style))
        elements.append(Paragraph(
            "Sector-Based Connection Summary — Active &amp; Closed Status",
            report_subtitle_style,
        ))

        # -- Shared filtered dataset is already domestic + Active>0 + sorted --
        non_commercial_rows = final_sorted

        # -- Metadata: only Sectors, Localities, Total Connections + Generated by AI --
        # Count non-commercial rows for metadata
        non_commercial_count = len(non_commercial_rows)
        non_commercial_sectors = len(set(r["sector"] for r in non_commercial_rows))
        non_commercial_closed = sum(r["closed"] for r in non_commercial_rows)
        non_commercial_suspended = sum(r.get("suspended", 0) for r in non_commercial_rows)
        non_commercial_active = sum(r["active"] for r in non_commercial_rows)
        non_commercial_total = non_commercial_closed + non_commercial_suspended + non_commercial_active
        non_commercial_budget = sum(r.get("budget", 0) for r in non_commercial_rows)

        meta_parts = []
        meta_parts.append(f"<b>Sectors:</b> {non_commercial_sectors}")
        meta_parts.append(f"<b>Localities:</b> {non_commercial_count}")
        meta_parts.append(f"<b>Total Connections:</b> {non_commercial_total:,}")
        meta_parts.append("<b>Generated by AI</b>")
        elements.append(Paragraph(" &nbsp;&nbsp;|&nbsp;&nbsp; ".join(meta_parts), meta_style))

        # -- Build table data with Paragraph wrapping for Sector/Locality --
        # Uses active_cols list to only include selected columns
        table_data = []
        row_types = []  # "header" | "data" | "grand_total"

        # Column style mapping: which style to use for each column type
        col_styles = {
            "sr": cell_center_style,
            "sector": cell_wrap_style,
            "locality": cell_wrap_style,
            "rate": ParagraphStyle("CellRate", parent=cell_center_style, alignment=1),
            "closed": cell_center_style,
            "suspended": cell_center_style,
            "active": cell_center_style,
            "total": cell_center_style,
            "budget": ParagraphStyle("CellBudget", parent=cell_center_style, alignment=1),
        }

        # Header row — only include selected columns
        table_data.append([Paragraph(COL_DEFS[c][0], header_cell_style) for c in active_cols])
        row_types.append("header")

        # Data rows — filter to non-commercial, then build rows with selected columns
        serial_counter = 0
        for r in non_commercial_rows:
            serial_counter += 1
            # Update serial in the row for display
            r_display = dict(r)
            r_display["serial"] = serial_counter
            row_cells = []
            for c in active_cols:
                val = COL_DEFS[c][1](r_display)
                row_cells.append(Paragraph(str(val), col_styles[c]))
            table_data.append(row_cells)
            row_types.append("data")

        # Grand Total row — only include selected columns
        gt_cells = []
        for c in active_cols:
            if c == "sr":
                gt_cells.append(Paragraph("", cell_center_style))
            elif c == "sector":
                gt_cells.append(Paragraph("GRAND TOTAL", ParagraphStyle("GrandLabel", parent=cell_wrap_style, fontName="Helvetica-Bold", fontSize=9, textColor=PDF_GRAND_FG)))
            elif c == "locality":
                gt_cells.append(Paragraph("", cell_center_style))
            elif c == "rate":
                gt_cells.append(Paragraph("", cell_center_style))
            else:
                gt_val = {"closed": non_commercial_closed, "suspended": non_commercial_suspended, "active": non_commercial_active, "total": non_commercial_total, "budget": non_commercial_budget}.get(c, 0)
                gt_cells.append(Paragraph(str(int(gt_val)) if c == "budget" else str(gt_val), ParagraphStyle(f"Grand{c}", parent=cell_center_style, fontName="Helvetica-Bold", fontSize=9, textColor=PDF_GRAND_FG)))
        table_data.append(gt_cells)
        row_types.append("grand_total")

        # -- Column widths (mm) — dynamic based on selected columns --
        # Available width: 180mm (A4 with 15mm margins on each side)
        # Distribute extra space proportionally to flexible columns
        # Fix: use the full A4 printable width without overflow. Text columns
        # get most of the room for wrapping; numeric columns stay compact.
        consumer_col_weights = {
            "sr": 11,
            "sector": 50,
            "locality": 56,
            "rate": 18,
            "closed": 16,
            "suspended": 19,
            "active": 16,
            "total": 21,
            "budget": 28,
        }
        total_weight = sum(consumer_col_weights.get(c, COL_DEFS[c][2]) for c in active_cols)
        col_widths = [
            usable_w * (consumer_col_weights.get(c, COL_DEFS[c][2]) / total_weight)
            for c in active_cols
        ]

        # -- Create table --
        t = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign="CENTER")

        # -- Build table style commands --
        style_cmds = [
            # Header row styling
            ("BACKGROUND", (0, 0), (-1, 0), PDF_HEADER_BG),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # Grid lines (subtle light teal)
            ("GRID", (0, 0), (-1, -1), 0.75, PDF_GRID),
            ("BOX", (0, 0), (-1, -1), 0.9, PDF_GRID),
            # Padding for all cells
            # More padding gives wrapped text breathing room and avoids a tight grid.
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]

        # -- Alternate row backgrounds for data rows only --
        data_row_indices = [i for i, rt in enumerate(row_types) if rt == "data"]
        for idx in data_row_indices:
            if data_row_indices.index(idx) % 2 == 1:
                style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), PDF_ALT_ROW))
            else:
                style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), PDF_WHITE_ROW))

        # -- Grand Total row styling --
        grand_idx = len(table_data) - 1
        style_cmds.append(("BACKGROUND", (0, grand_idx), (-1, grand_idx), PDF_GRAND_BG))
        style_cmds.append(("LINEABOVE", (0, grand_idx), (-1, grand_idx), 1.2, PDF_GRAND_FG))
        style_cmds.append(("LINEBELOW", (0, grand_idx), (-1, grand_idx), 1.2, PDF_GRAND_FG))

        t.setStyle(TableStyle(style_cmds))
        elements.append(t)

        doc.build(elements)
        return Response(buf.getvalue(), mimetype="application/pdf",
                        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"})

    # -----------------------------------------------------------------------
    # COMMERCIAL PDF export — Separate PDF for COMMERCIAL sector records only.
    # Uses the same A4 portrait design with light colours and proper wrapping.
    # -----------------------------------------------------------------------
    if fmt_type == "commercial-pdf":
        # ---------------------------------------------------------------
        # COMMERCIAL PDF EXPORT
        # Uses locality-level detail when available (commercial_detailed_rows).
        # Falls back to sector-aggregated summary rows for backward compat.
        # Each commercial locality appears as a separate row in the PDF.
        # ---------------------------------------------------------------
        # Shared filtered dataset is already commercial + Active>0 + sorted
        commercial_rows = final_sorted
        # Selection rule: if every commercial row is unchecked, the POSTed
        # selected dataset is empty.  Still generate the report shell and zero
        # grand total rather than falling back to cached/full rows.

        # Compute commercial-specific totals (from the shared filtered rows)
        commercial_closed = sum(r["closed"] for r in commercial_rows)
        commercial_suspended = sum(r.get("suspended", 0) for r in commercial_rows)
        commercial_active = sum(r["active"] for r in commercial_rows)
        commercial_total = commercial_closed + commercial_suspended + commercial_active
        commercial_budget = sum(r.get("budget", 0) for r in commercial_rows)

        # Recalculate sector_count and locality_count for commercial only
        commercial_sectors = set(r["sector"] for r in commercial_rows)
        commercial_localities = len(commercial_rows)

        # -- Colour palette (same as main PDF for consistency) --
        PDF_HEADER_BG = colors.HexColor("#e8f5f3")
        PDF_HEADER_FG = colors.HexColor("#1a5c52")
        PDF_ALT_ROW = colors.HexColor("#f7faf9")
        PDF_WHITE_ROW = colors.white
        PDF_GRAND_BG = colors.HexColor("#d4edda")
        PDF_GRAND_FG = colors.HexColor("#155724")
        # Darker grid lines keep the printed PDF visibly table-shaped.
        PDF_GRID = colors.HexColor("#6fa8a0")
        PDF_BODY_FG = colors.HexColor("#2c3e50")

        # -- Page setup: A4 portrait (match Consumer PDF margins) --
        page_w, page_h = A4
        top_m = 14 * mm
        bottom_m = 12 * mm
        left_m = 8 * mm
        right_m = 8 * mm
        usable_w = page_w - left_m - right_m

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=left_m,
            rightMargin=right_m,
            topMargin=top_m,
            bottomMargin=bottom_m,
        )

        styles = getSampleStyleSheet()

        # -- Heading styles --
        commercial_title_style = ParagraphStyle(
            "CommercialReportTitle",
            parent=styles["Heading1"],
            fontSize=18,
            fontName="Helvetica-Bold",
            textColor=PDF_HEADER_FG,
            alignment=1,
            spaceAfter=2 * mm,
        )
        commercial_subtitle_style = ParagraphStyle(
            "CommercialReportSubtitle",
            parent=styles["Normal"],
            fontSize=10,
            fontName="Helvetica",
            textColor=colors.HexColor("#5a7a74"),
            alignment=1,
            spaceAfter=3 * mm,
        )
        commercial_meta_style = ParagraphStyle(
            "CommercialReportMeta",
            parent=styles["Normal"],
            fontSize=9,
            fontName="Helvetica",
            textColor=colors.HexColor("#6c8a84"),
            alignment=1,
            spaceAfter=6 * mm,
        )

        # -- Styles for wrapping text inside cells (match Consumer PDF) --
        cell_wrap_style = ParagraphStyle(
            "CommCellWrap",
            parent=styles["Normal"],
            fontSize=8.2,
            fontName="Helvetica",
            textColor=PDF_BODY_FG,
            leading=10.8,
            alignment=0,
        )
        cell_center_style = ParagraphStyle(
            "CommCellCenter",
            parent=styles["Normal"],
            fontSize=8.2,
            fontName="Helvetica",
            textColor=PDF_BODY_FG,
            alignment=1,
            leading=10.8,
        )
        header_cell_style = ParagraphStyle(
            "CommHeaderCell",
            parent=styles["Normal"],
            fontSize=9,
            fontName="Helvetica-Bold",
            textColor=PDF_HEADER_FG,
            alignment=1,
        )

        elements = []

        # -- Report heading block --
        elements.append(Paragraph("Commercial Sector Report", commercial_title_style))
        elements.append(Paragraph(
            "COMMERCIAL Connections — Active &amp; Closed Status",
            commercial_subtitle_style,
        ))

        # -- Metadata: Sectors, Localities, Total Connections + Generated by AI --
        meta_parts = []
        meta_parts.append(f"<b>Sectors:</b> {len(commercial_sectors)}")
        meta_parts.append(f"<b>Localities:</b> {commercial_localities}")
        meta_parts.append(f"<b>Total Connections:</b> {commercial_total:,}")
        meta_parts.append("<b>Generated by AI</b>")
        elements.append(Paragraph(" &nbsp;&nbsp;|&nbsp;&nbsp; ".join(meta_parts), commercial_meta_style))

        # -- Build table data with Paragraph wrapping for Sector/Locality --
        # Uses active_cols list to only include selected columns
        table_data = []
        row_types = []  # "header" | "data" | "grand_total"

        # Column style mapping: which style to use for each column type
        col_styles = {
            "sr": cell_center_style,
            "sector": cell_wrap_style,
            "locality": cell_wrap_style,
            "rate": ParagraphStyle("CellRateComm", parent=cell_center_style, alignment=1),
            "closed": cell_center_style,
            "suspended": cell_center_style,
            "active": cell_center_style,
            "total": cell_center_style,
            "budget": ParagraphStyle("CellBudgetComm", parent=cell_center_style, alignment=1),
        }

        # Header row — only include selected columns
        table_data.append([Paragraph(COL_DEFS[c][0], header_cell_style) for c in active_cols])
        row_types.append("header")

        # Data rows — re-assign serial numbers for commercial-only report
        serial_counter = 0
        for r in commercial_rows:
            serial_counter += 1
            r_display = dict(r)
            r_display["serial"] = serial_counter
            row_cells = []
            for c in active_cols:
                val = COL_DEFS[c][1](r_display)
                row_cells.append(Paragraph(str(val), col_styles[c]))
            table_data.append(row_cells)
            row_types.append("data")

        # Grand Total row — only include selected columns
        gt_cells = []
        for c in active_cols:
            if c == "sr":
                gt_cells.append(Paragraph("", cell_center_style))
            elif c == "sector":
                gt_cells.append(Paragraph("GRAND TOTAL", ParagraphStyle("CommGrandLabel", parent=cell_wrap_style, fontName="Helvetica-Bold", fontSize=9, textColor=PDF_GRAND_FG)))
            elif c == "locality":
                gt_cells.append(Paragraph("", cell_center_style))
            elif c == "rate":
                gt_cells.append(Paragraph("", cell_center_style))
            else:
                gt_val = {"closed": commercial_closed, "suspended": commercial_suspended, "active": commercial_active, "total": commercial_total, "budget": commercial_budget}.get(c, 0)
                gt_cells.append(Paragraph(str(int(gt_val)) if c == "budget" else str(gt_val), ParagraphStyle(f"CommGrand{c}", parent=cell_center_style, fontName="Helvetica-Bold", fontSize=9, textColor=PDF_GRAND_FG)))
        table_data.append(gt_cells)
        row_types.append("grand_total")

        # -- Column widths (mm) — match Consumer PDF proportional layout --
        consumer_col_weights = {
            "sr": 11,
            "sector": 50,
            "locality": 56,
            "rate": 18,
            "closed": 16,
            "suspended": 19,
            "active": 16,
            "total": 21,
            "budget": 28,
        }
        total_weight = sum(consumer_col_weights.get(c, COL_DEFS[c][2]) for c in active_cols)
        col_widths = [
            usable_w * (consumer_col_weights.get(c, COL_DEFS[c][2]) / total_weight)
            for c in active_cols
        ]

        t = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign="CENTER")

        # -- Build table style commands --
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), PDF_HEADER_BG),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.75, PDF_GRID),
            ("BOX", (0, 0), (-1, -1), 0.9, PDF_GRID),
            # More padding gives wrapped text breathing room and avoids a tight grid.
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]

        # Alternate row backgrounds
        data_row_indices = [i for i, rt in enumerate(row_types) if rt == "data"]
        for idx in data_row_indices:
            if data_row_indices.index(idx) % 2 == 1:
                style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), PDF_ALT_ROW))
            else:
                style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), PDF_WHITE_ROW))

        # Grand Total row styling
        grand_idx = len(table_data) - 1
        style_cmds.append(("BACKGROUND", (0, grand_idx), (-1, grand_idx), PDF_GRAND_BG))
        style_cmds.append(("LINEABOVE", (0, grand_idx), (-1, grand_idx), 1.2, PDF_GRAND_FG))
        style_cmds.append(("LINEBELOW", (0, grand_idx), (-1, grand_idx), 1.2, PDF_GRAND_FG))

        t.setStyle(TableStyle(style_cmds))
        elements.append(t)

        doc.build(elements)
        return Response(buf.getvalue(), mimetype="application/pdf",
                        headers={"Content-Disposition": "attachment; filename=commercial_sector_report.pdf"})

    if fmt_type == "csv":
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        writer.writerow(grand)
        return Response(out.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition": f"attachment; filename={filename}.csv"})

    if fmt_type == "xlsx":
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            all_rows = rows + [grand]
            pd.DataFrame(all_rows, columns=headers).to_excel(writer, sheet_name="Sector Summary", index=False)
        buf.seek(0)
        return Response(buf.getvalue(),
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"})

    flash("Unknown export format.")
    return redirect(url_for("consumer_report"))


# ---------------------------------------------------------------------------
# File Merger Route
# ---------------------------------------------------------------------------
# The File Merger page uses client-side JavaScript (SheetJS + pdf-lib) for
# all merging. Only the page render route is needed on the server.
# No server-side merge endpoints are required.


@app.route("/file-merger")
def file_merger():
    """Render the File Merger page with CSV/Excel and PDF merge sections."""
    return render_template("file_merger.html", active_page="file_merger")


# ---------------------------------------------------------------------------
# File Column Matcher
# Client-side page for matching and merging columns between two files.
# No server-side processing required — all logic runs in the browser.
# ---------------------------------------------------------------------------


@app.route("/file-column-matcher")
def file_column_matcher():
    """Render the File Column Matcher page for matching columns between two files."""
    return render_template("file_column_matcher.html", active_page="file_column_matcher")


# ---------------------------------------------------------------------------
# Arrear Calculator
# ---------------------------------------------------------------------------
_arrear_calc_data = None
_arrear_calc_filename = None


@app.route("/arrear-calculator", methods=["GET", "POST"])
def arrear_calculator():
    global _arrear_calc_data, _arrear_calc_filename

    if request.method == "POST":
        action = request.form.get("action", "")

        # JSON upload: client-side parser sends pre-computed summary
        if request.is_json:
            data = request.get_json(silent=True)
            if not data or "rows" not in data:
                if is_ajax():
                    return ajax_error("Invalid data payload.")
                flash("Invalid upload data.")
                return redirect(url_for("arrear_calculator"))

            _arrear_calc_data = data
            _arrear_calc_filename = data.get("filename", "upload.csv")
            # Persist in session so the data survives the redirect on
            # serverless platforms (stateless globals).
            session["arrear_calc_data"] = data
            session["arrear_calc_filename"] = _arrear_calc_filename

            msg = (
                f"File uploaded. Found {data.get('total_connections', 0):,} connections "
                f"across {data.get('sector_count', 0)} sectors and "
                f"{data.get('locality_count', 0)} localities."
            )
            if is_ajax():
                return ajax_ok(message=msg, redirect_url=url_for("arrear_calculator"))
            flash(msg)
            return render_template(
                "arrear_calculator.html",
                summary=data,
                filename=_arrear_calc_filename,
                total_rows=data.get("total_rows", 0),
                active_page="arrear_calc",
            )

        if action == "clear":
            _arrear_calc_data = None
            _arrear_calc_filename = None
            session.pop("arrear_calc_data", None)
            session.pop("arrear_calc_filename", None)
            flash("Arrear Calculator data cleared.")
            return redirect(url_for("arrear_calculator"))

    # GET — fall back to session data (serverless-safe)
    if _arrear_calc_data:
        summary = _arrear_calc_data
        filename = _arrear_calc_filename
    elif session.get("arrear_calc_data"):
        summary = session.get("arrear_calc_data")
        filename = session.get("arrear_calc_filename")
    else:
        summary = None
        filename = None

    if summary:
        return render_template(
            "arrear_calculator.html",
            summary=summary,
            filename=filename,
            total_rows=summary.get("total_rows", 0),
            active_page="arrear_calc",
        )
    return render_template(
        "arrear_calculator.html",
        summary=None,
        filename=None,
        total_rows=0,
        active_page="arrear_calc",
    )


def _parse_arrear_export_cols(cols_param):
    """Parse comma-separated column keys into an ordered list.

    Fixed column order: SR, Sector, Locality, Closed, Suspended, Active,
    Open, Total Connections, 2023-2024, 2024-2025, 2025-2026, Total Arrears.
    """
    COL_ORDER = ["sr", "sector", "locality", "closed", "suspended", "active",
                 "open", "total", "fy2023", "fy2024", "fy2025", "arrears"]
    COL_NAMES = {
        "sr": "SR", "sector": "Sector", "locality": "Locality",
        "closed": "Closed", "suspended": "Suspended", "active": "Active",
        "open": "Open", "total": "Total Conn.", "fy2023": "2023-2024",
        "fy2024": "2024-2025", "fy2025": "2025-2026", "arrears": "Total Arrears",
    }
    if cols_param:
        selected = [c.strip() for c in cols_param.split(",") if c.strip() in COL_NAMES]
    else:
        selected = list(COL_ORDER)
    return [(k, COL_NAMES[k]) for k in selected if k in COL_NAMES]


def _build_arrear_export_rows(rows, selected_keys):
    """Build export rows from summary data, selecting only requested columns."""
    out_rows = []
    for row in rows:
        r = []
        for key in selected_keys:
            if key == "sr":
                r.append(str(row.get("serial", "")))
            elif key == "sector":
                r.append(row.get("sector", ""))
            elif key == "locality":
                r.append(row.get("locality", ""))
            elif key == "fy2023":
                r.append(f"{row.get('fy2023', 0):,.0f}")
            elif key == "fy2024":
                r.append(f"{row.get('fy2024', 0):,.0f}")
            elif key == "fy2025":
                r.append(f"{row.get('fy2025', 0):,.0f}")
            elif key == "closed":
                r.append(str(row.get("closed", 0)))
            elif key == "suspended":
                r.append(str(row.get("suspended", 0)))
            elif key == "active":
                r.append(str(row.get("active", 0)))
            elif key == "open":
                r.append(str(row.get("open", 0)))
            elif key == "total":
                r.append(str(row.get("total_connections", 0)))
            elif key == "arrears":
                r.append(f"{row.get('total_arrears', 0):,.0f}")
        out_rows.append(r)
    return out_rows


def _sort_arrear_rows(rows, priority, order):
    """Sort rows by the given status priority and order."""
    sign = -1 if order == "desc" else 1
    if priority == "none":
        return sorted(
            rows,
            key=lambda r: (
                sign * r.get("total_arrears", 0),
                r.get("sector", "").lower(),
                r.get("locality", "").lower(),
            ),
        )
    key_map = {
        "active_first": "active",
        "suspended_first": "suspended",
        "closed_first": "closed",
        "open_first": "open",
    }
    sort_key = key_map.get(priority, "active")
    return sorted(
        rows,
        key=lambda r: (
            sign * r.get(sort_key, 0),
            r.get("sector", "").lower(),
            r.get("locality", "").lower(),
        ),
    )


@app.route("/arrear-calculator/export/<fmt_type>", methods=["GET", "POST"])
def export_arrear_calculator(fmt_type: str):
    # Load summary from POST body, in-memory, or fail
    summary = None
    if request.method == "POST":
        raw = request.form.get("summary_data") or ""
        if raw:
            try:
                payload = json.loads(raw)
                if isinstance(payload, dict) and "rows" in payload:
                    summary = payload
            except (json.JSONDecodeError, TypeError):
                pass
    if not summary:
        summary = _arrear_calc_data or session.get("arrear_calc_data")

    if not summary or not summary.get("rows"):
        flash("No arrear data to export. Upload a CSV first.")
        return redirect(url_for("arrear_calculator"))

    # Tab filter — "commercial" or "domestic" (default: all)
    tab = request.args.get("tab", "").strip().lower()
    tab_label = ""
    if tab == "commercial":
        tab_label = " Commercial"
    elif tab == "domestic":
        tab_label = " Domestic"

    # Sort rows
    priority = request.args.get("priority", "none")
    order = request.args.get("order", "desc")
    rows = _sort_arrear_rows(summary["rows"], priority, order)

    # Column selection
    cols_param = request.args.get("cols", "")
    col_defs = _parse_arrear_export_cols(cols_param)
    selected_keys = [k for k, _ in col_defs]
    headers = [h for _, h in col_defs]
    export_rows = _build_arrear_export_rows(rows, selected_keys)

    # Grand total row — must match the column order from _parse_arrear_export_cols
    gt_map = {
        "sr": "", "sector": "GRAND TOTAL", "locality": "",
        "closed": str(summary.get("grand_status", {}).get("closed", 0)),
        "suspended": str(summary.get("grand_status", {}).get("suspended", 0)),
        "active": str(summary.get("grand_status", {}).get("active", 0)),
        "open": str(summary.get("grand_status", {}).get("open", 0)),
        "total": str(summary.get("total_connections", 0)),
        "fy2023": f"{summary.get('grand_fy2023', 0):,.0f}",
        "fy2024": f"{summary.get('grand_fy2024', 0):,.0f}",
        "fy2025": f"{summary.get('grand_fy2025', 0):,.0f}",
        "arrears": f"{summary.get('grand_total_arrears', 0):,.0f}",
    }
    grand = [gt_map.get(k, "") for k in selected_keys]
    export_rows.append(grand)

    filename_base = summary.get("filename", "arrear_report")
    if filename_base.lower().endswith(".csv"):
        filename_base = filename_base[:-4]
    filename_base += tab_label.replace(" ", "_") if tab_label else ""

    if fmt_type == "csv":
        import csv as csv_mod

        buf = io.StringIO()
        writer = csv_mod.writer(buf)
        writer.writerow(headers)
        writer.writerows(export_rows)
        csv_bytes = buf.getvalue().encode("utf-8-sig")
        return Response(
            csv_bytes,
            mimetype="text/csv",
            headers={"Content-Disposition": f'attachment;filename="{filename_base}_arrears.csv"'},
        )

    if fmt_type == "xlsx":
        df = pd.DataFrame(export_rows, columns=headers)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Arrears")
        xlsx_bytes = buf.getvalue()
        return Response(
            xlsx_bytes,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment;filename="{filename_base}_arrears.xlsx"'},
        )

    if fmt_type == "pdf":
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer)
        from reportlab.lib.enums import TA_LEFT, TA_CENTER

        buf = io.BytesIO()
        page_w, page_h = landscape(A4)
        margin = 10 * mm
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=margin,
            rightMargin=margin,
            topMargin=12 * mm,
            bottomMargin=10 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ArrearTitle", parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=16, leading=19, spaceAfter=6, spaceBefore=0,
            textColor=rl_colors.HexColor("#10243f"),
        )
        meta_style = ParagraphStyle(
            "ArrearMeta", parent=styles["Normal"],
            fontSize=7.5, textColor=rl_colors.HexColor("#555"),
            spaceAfter=4, spaceBefore=0,
        )
        cell_style = ParagraphStyle(
            "Cell", parent=styles["Normal"],
            fontSize=7.2, leading=9, spaceBefore=0, spaceAfter=0,
        )
        cell_left = ParagraphStyle(
            "CellLeft", parent=cell_style, alignment=TA_LEFT,
        )
        cell_center = ParagraphStyle(
            "CellCenter", parent=cell_style, alignment=TA_CENTER,
        )
        hdr_style = ParagraphStyle(
            "HdrCell", parent=styles["Normal"],
            fontSize=7.5, leading=9, alignment=TA_CENTER,
            textColor=rl_colors.white, fontName="Helvetica-Bold",
            spaceBefore=0, spaceAfter=0,
        )

        elements = []
        elements.append(Paragraph(f"Arrear Report{tab_label}", title_style))
        gen = datetime.now().strftime("%d %b %Y, %H:%M")
        meta_parts = [
            f"Generated: {gen}",
            f"Records: {summary.get('total_rows', 0):,}",
            f"Sectors: {summary.get('sector_count', 0)}",
            f"Localities: {summary.get('locality_count', 0)}",
            f"Grand Total Arrears: {summary.get('grand_total_arrears', 0):,.0f}",
        ]
        elements.append(Paragraph("   |   ".join(meta_parts), meta_style))
        elements.append(Spacer(1, 6))

        # Build table data with Paragraph objects for wrapping
        left_align_cols = {"sector", "locality"}
        table_data = []
        # Header row
        hdr_row = [Paragraph(h, hdr_style) for _, h in col_defs]
        table_data.append(hdr_row)
        # Data rows
        for row in export_rows:
            cells = []
            for idx, (key, _) in enumerate(col_defs):
                txt = str(row[idx]) if idx < len(row) else ""
                st = cell_left if key in left_align_cols else cell_center
                cells.append(Paragraph(txt, st))
            table_data.append(cells)

        # Adaptive column widths: flexible columns absorb spare space.
        # Priority: Sector > Locality > Total Arrears > Year cols > others.
        usable = page_w - 2 * margin

        # Minimum base widths for each column type
        MIN_WIDTHS = {
            "sr": 20, "closed": 26, "suspended": 30, "active": 26,
            "open": 22, "total": 30,
        }
        # Flexible columns with proportional weights — extra space is
        # distributed roughly in proportion to these weights.
        FLEX_WEIGHTS = {
            "sector": 56, "locality": 64,
            "fy2023": 40, "fy2024": 40, "fy2025": 40,
            "arrears": 48,
        }

        col_widths = []
        flex_idx = []
        fixed_total = 0
        for k, _ in col_defs:
            if k in FLEX_WEIGHTS:
                flex_idx.append(len(col_widths))
                col_widths.append(0)
            else:
                w = MIN_WIDTHS.get(k, 30)
                col_widths.append(w)
                fixed_total += w

        # Distribute remaining width among flexible columns by weight
        remaining = usable - fixed_total
        if flex_idx and remaining > 0:
            total_weight = sum(FLEX_WEIGHTS[col_defs[i][0]] for i in flex_idx)
            for i in flex_idx:
                k = col_defs[i][0]
                col_widths[i] = remaining * (FLEX_WEIGHTS[k] / total_weight)
            # Scale so the table exactly fills the usable width
            total = sum(col_widths)
            if total > 0:
                scale = usable / total
                col_widths = [w * scale for w in col_widths]
        else:
            total = sum(col_widths)
            if total > 0 and total < usable:
                scale = usable / total
                col_widths = [w * scale for w in col_widths]

        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        accent = rl_colors.HexColor("#2f6f6d")
        dark_border = rl_colors.HexColor("#555555")
        tbl.setStyle(TableStyle([
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), accent),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7.5),
            # Body
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 7.2),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # Sector + Locality left-aligned
            ("ALIGN", (1, 1), (2, -1), "LEFT"),
            # Grid — dark visible borders
            ("GRID", (0, 0), (-1, -1), 0.5, dark_border),
            ("LINEBELOW", (0, 0), (-1, 0), 1, accent),
            # Alternating row colors
            ("ROWBACKGROUNDS", (0, 1), (-1, -2),
             [rl_colors.white, rl_colors.HexColor("#f4f7fa")]),
            # Grand total row
            ("BACKGROUND", (0, -1), (-1, -1), rl_colors.HexColor("#e8f5e9")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            # Comfortable padding for readability
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(tbl)
        doc.build(elements)
        pdf_bytes = buf.getvalue()
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f'attachment;filename="{filename_base}_arrears.pdf"'},
        )

    flash("Unsupported export format.")
    return redirect(url_for("arrear_calculator"))


# ---------------------------------------------------------------------------
# Consumer Sector Remaining Report
# Combines consumer sector statistics with six-month pending amounts.
# ---------------------------------------------------------------------------

def _normalize_sector_key(name: str) -> str:
    """Normalize sector name for matching between consumer and bill data."""
    return " ".join(str(name or "").strip().lower().split())


def build_consumer_sector_remaining_report(year: int, season: str) -> list[dict]:
    """Build combined consumer sector + remaining amount rows.

    Joins consumer sector summary (from cached upload) with bill pending
    amounts (from bills database) by normalized sector name.

    Returns a list of dicts with keys:
        serial, sector, locality, active, budget, remaining
    """
    # --- Step 1: Load consumer sector summary ---
    # Try disk cache first, then fall back to in-memory globals
    cached_summary, _, _ = _load_consumer_summary_cache()
    if not cached_summary or not cached_summary.get("summary_rows"):
        # Fallback to in-memory consumer report data
        if _last_consumer_summary and _last_consumer_summary.get("summary_rows"):
            cached_summary = _last_consumer_summary
        elif _consumer_report_data:
            cached_summary = _build_consumer_sector_summary(_consumer_report_data)
            cached_summary = _filter_active_rows(cached_summary)
    if not cached_summary or not cached_summary.get("summary_rows"):
        return []

    # Build lookup: normalized_sector → { sector, locality, active, budget }
    consumer_lookup: dict[str, dict] = {}
    for row in cached_summary.get("summary_rows", []):
        sector_name = (row.get("sector") or "").strip()
        if not sector_name:
            continue
        # Budget Suggestion must use the same cleaned Consumer Sector names as
        # the main Consumer Report; otherwise address-derived rows appear as
        # duplicate sectors in the remaining-budget table and exports.
        if (
            _is_faulty_empty_consumer_sector(sector_name)
            or _is_extra_zain_city_13g_sector(sector_name)
            or _is_extra_noor_mohalla_main_road_sector(sector_name)
        ):
            continue
        sector_name, locality_name = _canonical_consumer_sector_locality(
            sector_name,
            (row.get("locality") or "").strip(),
        )
        norm_key = _normalize_sector_key(sector_name)
        active = int(row.get("active", 0))
        if active <= 0:
            continue  # skip zero-active rows
        consumer_lookup[norm_key] = {
            "sector": sector_name,
            "locality": locality_name,
            "active": active,
            "budget": float(row.get("budget", 0)),
        }

    if not consumer_lookup:
        return []

    # --- Step 2: Get pending amounts from bills database OR cache ---
    bill_sector_pending: dict[str, float] = {}

    # First try bill pending cache (summary format upload)
    cache_path = os.path.join(UPLOAD_FOLDER, "bill_pending_cache.json")
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                cache_data = json.load(f)
            pending_lookup = cache_data.get("pending", {})
            for sector_name, amount in pending_lookup.items():
                # Keep pending-amount sector keys aligned with the cleaned
                # consumer-sector keys so the join does not create duplicates.
                if (
                    _is_faulty_empty_consumer_sector(sector_name)
                    or _is_extra_zain_city_13g_sector(sector_name)
                    or _is_extra_noor_mohalla_main_road_sector(sector_name)
                ):
                    continue
                sector_name, _ = _canonical_consumer_sector_locality(sector_name, "")
                norm_key = _normalize_sector_key(sector_name)
                bill_sector_pending[norm_key] = float(amount or 0)
        except (json.JSONDecodeError, OSError):
            pass

    # If no cache, try bills database (raw bill format upload)
    if not bill_sector_pending:
        season_bill_ids = _get_season_bill_ids(year, season)
        if season_bill_ids:
            id_list = ",".join(str(i) for i in season_bill_ids)
            with get_db() as conn:
                bill_rows = conn.execute(
                    f"""
                    SELECT
                        id,
                        COALESCE(NULLIF(TRIM(sector), ''), 'Unassigned Sector') AS sector,
                        amount_received,
                        raw_data
                    FROM bills
                    WHERE id IN ({id_list})
                    """
                ).fetchall()

            for row in bill_rows:
                sector = row["sector"]
                amount_received = float(row["amount_received"] or 0)
                if amount_received > 0:
                    continue  # only count remaining bills
                try:
                    data = json.loads(row["raw_data"])
                    water_fee = 0.0
                    wf = data.get("water fee")
                    if wf is not None:
                        water_fee = parse_number(str(wf).replace(",", ""))
                except (json.JSONDecodeError, ValueError):
                    water_fee = 0.0
                # Raw bill imports can carry locality/address text as sector
                # names; skip those known faulty duplicates before grouping.
                if (
                    _is_faulty_empty_consumer_sector(sector)
                    or _is_extra_zain_city_13g_sector(sector)
                    or _is_extra_noor_mohalla_main_road_sector(sector)
                ):
                    continue
                sector, _ = _canonical_consumer_sector_locality(sector, "")
                norm_sector = _normalize_sector_key(sector)
                bill_sector_pending[norm_sector] = bill_sector_pending.get(norm_sector, 0.0) + water_fee

    # --- Step 3: Join consumer data with bill pending amounts ---
    result: list[dict] = []
    serial = 0
    for norm_key, cdata in sorted(consumer_lookup.items(), key=lambda x: x[1]["sector"].lower()):
        serial += 1
        remaining = bill_sector_pending.get(norm_key, 0.0)
        result.append({
            "serial": serial,
            "sector": cdata["sector"],
            "locality": cdata["locality"],
            "active": cdata["active"],
            "budget": cdata["budget"],
            "remaining": remaining,
        })

    return result


@app.route("/consumer-sector-remaining-report", methods=["GET", "POST"])
def consumer_sector_remaining_report():
    """Display the combined Consumer Sector Remaining Report.

    Supports two upload modes:
    - JSON body (Content-Type: application/json): client-side parsed consumer CSV
    - Multipart form (action=upload_bills): server-side bill file import
    """
    global _consumer_report_data, _consumer_report_filename, _last_consumer_summary

    year_param = request.args.get("year", "").strip()
    season = request.args.get("season", "").strip().lower()

    # Validate season
    if season not in ("jan-jun", "jul-dec"):
        season = "jan-jun"

    # Validate year
    try:
        year = int(year_param)
    except (TypeError, ValueError):
        year = datetime.now().year

    season_label = "January to June" if season == "jan-jun" else "July to December"

    # ---- POST handling ----
    if request.method == "POST":
        # Mode 1: JSON consumer CSV upload (client-side parsed)
        if request.is_json:
            data = request.get_json(silent=True)
            if not data or "summary_rows" not in data:
                return ajax_error("Invalid consumer data payload.")

            _consumer_report_data = data.get("summary_rows", [])
            _consumer_report_filename = data.get("filename", "upload.csv")

            summary = {
                "summary_rows": data.get("summary_rows", []),
                "sector_totals": data.get("sector_totals", {}),
                "grand_total": data.get("grand_total", {"closed": 0, "suspended": 0, "active": 0, "total": 0, "budget": 0}),
                "sector_count": data.get("sector_count", 0),
                "locality_count": data.get("locality_count", 0),
                "total_connections": data.get("total_connections", 0),
                "total_budget": data.get("total_budget", 0),
                "commercial_detailed_rows": data.get("commercial_detailed_rows", []),
                "commercial_grand_total": data.get("commercial_grand_total", {"closed": 0, "suspended": 0, "active": 0, "total": 0, "budget": 0}),
                "unmatched_rate_types": data.get("unmatched_rate_types", []),
                "unmatched_budget_count": data.get("unmatched_budget_count", 0),
            }
            summary = _filter_active_rows(summary)
            _last_consumer_summary = summary
            _save_consumer_summary_cache(summary, _consumer_report_filename, data.get("total_rows", 0))

            msg = f"Consumer data uploaded. {summary['sector_count']} sectors, {summary['total_connections']:,} connections."
            return ajax_ok(message=msg, redirect_url=url_for("consumer_sector_remaining_report",
                                                             season=season, year=year))

        # Mode 2: Bill file upload (multipart form)
        action = request.form.get("action", "")
        if action == "upload_bills":
            file = request.files.get("bill_file")
            if not file or not file.filename:
                flash("Please choose a bill file first.")
                return redirect(url_for("consumer_sector_remaining_report", season=season, year=year))
            if not allowed_file(file.filename):
                flash(f"Unsupported file type: {file.filename}")
                return redirect(url_for("consumer_sector_remaining_report", season=season, year=year))

            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
            save_path = os.path.join(app.config["UPLOAD_FOLDER"], f"bill_remaining_{timestamp}_{filename}")
            file.save(save_path)
            try:
                df = read_dataframe(save_path)
                cols_lower = [str(c).strip().lower() for c in df.columns]

                # Detect summary format: has "sector name" and "pending amount" columns
                has_sector_name = any("sector" in c and "name" in c for c in cols_lower)
                has_pending = any("pending" in c and "amount" in c for c in cols_lower)

                if has_sector_name and has_pending:
                    # Summary format: store pending amounts directly as JSON
                    import re

                    def _sector_base(name: str) -> str:
                        """Extract base sector name by removing phase/block/section numbers."""
                        n = re.sub(r"\s*\(?\s*phase\s*\d+\s*\)?", "", name, flags=re.IGNORECASE)
                        n = re.sub(r"\s*\(?\s*block\s*[a-z]?\s*\)?", "", n, flags=re.IGNORECASE)
                        n = re.sub(r"\s*\(?\s*section\s*\d+\s*\)?", "", n, flags=re.IGNORECASE)
                        return " ".join(n.split()).strip().lower()

                    pending_lookup = {}
                    base_to_original: dict[str, str] = {}
                    for _, row in df.iterrows():
                        sector_col = [c for c in df.columns if "sector" in str(c).lower() and "name" in str(c).lower()][0]
                        pending_col = [c for c in df.columns if "pending" in str(c).lower() and "amount" in str(c).lower()][0]
                        sector_name = str(row[sector_col]).strip()
                        pending_val = str(row[pending_col]).replace(",", "").strip()
                        try:
                            pending_val = float(pending_val)
                        except (ValueError, TypeError):
                            pending_val = 0.0
                        if sector_name:
                            # Summary uploads are normalized before caching so
                            # exports and previews read the same final dataset.
                            if (
                                _is_faulty_empty_consumer_sector(sector_name)
                                or _is_extra_zain_city_13g_sector(sector_name)
                                or _is_extra_noor_mohalla_main_road_sector(sector_name)
                            ):
                                continue
                            sector_name, _ = _canonical_consumer_sector_locality(sector_name, "")
                            base = _sector_base(sector_name)
                            if base in pending_lookup:
                                pending_lookup[base] += pending_val
                            else:
                                pending_lookup[base] = pending_val
                                base_to_original[base] = sector_name

                    # Use original (shortest) name for each merged group
                    merged_lookup = {base_to_original[k]: v for k, v in pending_lookup.items()}

                    cache_path = os.path.join(app.config["UPLOAD_FOLDER"], "bill_pending_cache.json")
                    with open(cache_path, "w", encoding="utf-8") as f:
                        json.dump({"pending": merged_lookup, "filename": file.name}, f, ensure_ascii=False, indent=2)
                    msg = f"Bill summary uploaded. {len(merged_lookup)} sectors with pending amounts."
                else:
                    # Raw bill format: import to database
                    imported, duplicates = import_bill_list_dataframe(df)
                    msg = f"Bill data saved. Imported {imported:,} row(s)."
                    if duplicates:
                        msg += f" Skipped {duplicates:,} duplicate(s)."
                flash(msg)
            except Exception as exc:
                flash(f"Failed to import bill file: {exc}")
            return redirect(url_for("consumer_sector_remaining_report", season=season, year=year))

        flash("Unknown upload action.")
        return redirect(url_for("consumer_sector_remaining_report", season=season, year=year))

    # ---- GET: build and render ----
    rows = build_consumer_sector_remaining_report(year, season)

    # Grand totals
    grand_active = sum(r["active"] for r in rows)
    grand_budget = sum(r["budget"] for r in rows)
    grand_remaining = sum(r["remaining"] for r in rows)

    # Upload status context
    consumer_filename = _consumer_report_filename
    consumer_rows = 0
    if _last_consumer_summary:
        consumer_rows = _last_consumer_summary.get("total_connections", 0)

    # Count bills in database
    bill_count = 0
    bill_filename = None
    try:
        with get_db() as conn:
            bill_count = conn.execute("SELECT COUNT(*) FROM bills").fetchone()[0]
    except Exception:
        pass

    return render_template(
        "consumer_sector_remaining_report.html",
        rows=rows,
        year=year,
        season=season,
        season_label=season_label,
        grand_active=grand_active,
        grand_budget=grand_budget,
        grand_remaining=grand_remaining,
        consumer_filename=consumer_filename,
        consumer_rows=consumer_rows,
        bill_filename=bill_filename,
        bill_count=bill_count,
    )


@app.route("/consumer-sector-remaining-report/export/<fmt_type>")
def export_consumer_sector_remaining(fmt_type: str):
    """Export the Consumer Sector Remaining Report as CSV, Excel, or PDF."""
    year_param = request.args.get("year", "").strip()
    season = request.args.get("season", "").strip().lower()

    if season not in ("jan-jun", "jul-dec"):
        season = "jan-jun"
    try:
        year = int(year_param)
    except (TypeError, ValueError):
        year = datetime.now().year

    season_label = "January to June" if season == "jan-jun" else "July to December"
    remaining_header = f"{season_label} Remaining"
    file_slug = f"Consumer_Sector_Remaining_{season_label.replace(' ', '_')}_{year}"

    rows = build_consumer_sector_remaining_report(year, season)

    headers = ["SR", "Sector", "Locality", "Active", "Budget (Rs.)", remaining_header]

    # Grand totals
    grand_active = sum(r["active"] for r in rows)
    grand_budget = sum(r["budget"] for r in rows)
    grand_remaining = sum(r["remaining"] for r in rows)
    grand_row = ["", "Grand Total", "", grand_active, grand_budget, grand_remaining]

    # Build table data
    table_data = []
    for r in rows:
        table_data.append([
            r["serial"],
            r["sector"],
            r["locality"],
            r["active"],
            fmt(r["budget"]),
            fmt(r["remaining"]),
        ])
    table_data.append(["", "Grand Total", "", grand_active, fmt(grand_budget), fmt(grand_remaining)])

    # PDF
    if fmt_type == "pdf":
        buf = io.BytesIO()
        margins = 8 * mm
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=6*mm, bottomMargin=6*mm, leftMargin=6*mm, rightMargin=6*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("CSReportTitle", parent=styles["Heading1"], fontSize=18, textColor=ACCENT, alignment=1, spaceAfter=4*mm, fontName="Helvetica-Bold")
        subtitle_style = ParagraphStyle("CSReportSubtitle", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#555555"), alignment=1, spaceAfter=6*mm)
        elements = [
            Paragraph("Consumer Sector Remaining Report", title_style),
            Paragraph(f"{season_label} &middot; {year}", subtitle_style),
        ]

        page_w = landscape(A4)[0] - 2 * margins
        col_widths = [page_w * 0.05, page_w * 0.28, page_w * 0.27, page_w * 0.10, page_w * 0.15, page_w * 0.15]

        def wrap_left(v):
            return wrap_pdf_body_cells([[v]], font_size=8, left_columns={0})[0][0]

        body_rows = []
        for r in table_data:
            body_rows.append([
                r[0], wrap_left(r[1]), wrap_left(r[2]), r[3], r[4], r[5]
            ])

        full_table = [headers] + body_rows
        elements.append(_make_pdf_table(full_table, col_widths=col_widths, left_cols={1, 2}, header_font_size=9, body_font_size=8, cell_padding=5))
        doc.build(elements)
        buf.seek(0)
        return Response(buf.getvalue(), mimetype="application/pdf", headers={"Content-Disposition": f"attachment; filename={file_slug}.pdf"})

    # CSV
    if fmt_type == "csv":
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(headers)
        writer.writerows(table_data)
        return Response(out.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={file_slug}.csv"})

    # Excel
    if fmt_type == "xlsx":
        df = pd.DataFrame(table_data, columns=headers)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Consumer Sector Remaining", index=False)
        buf.seek(0)
        return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={file_slug}.xlsx"})

    flash("Unsupported export format.")
    return redirect(url_for("consumer_sector_remaining_report"))


if __name__ == "__main__":
    port = int(os.environ.get("WATER_SUPPLY_PORT", "5000"))
    host = os.environ.get("WATER_SUPPLY_HOST", "127.0.0.1")
    url = f"http://{host}:{port}"

    if os.environ.get("WATER_SUPPLY_NO_BROWSER") != "1":
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()

    print(f"Water Supply Report is running at {url}")
    print("Close this window to stop the app.")
    app.run(host=host, port=port, debug=False, use_reloader=False)
